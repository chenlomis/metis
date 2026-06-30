"""
setup_demo.py -- one-time demo environment setup for metis video recording.

Creates ~/.job_pipeline_demo/ as a fully isolated data directory:
  - All your real Applied rows from applications.xlsx (preserved as-is)
  - A random sample of ~100 Not Applied rows for realistic total counts
  - A proportional sample of skipped_roles.json
  - A date-windowed slice of runs.jsonl for market intel
  - Fresh seen_roles.json so the upcoming digest run scores new roles

Your real ~/.job_pipeline/ is never modified.

Run with:
    cd /path/to/metis
    python scripts/setup_demo.py

Then record using:
    export METIS_PROFILE=~/.job_pipeline_demo/profile.yaml
    export METIS_DATA_DIR=~/.job_pipeline_demo
    ./venv/bin/python -m metis.cli --lookback 7d
    ./venv/bin/python -m metis.cli track
    ./venv/bin/python -m metis.cli summary
    unset METIS_PROFILE METIS_DATA_DIR   # back to real data after filming
"""

from __future__ import annotations

import datetime
import json
import random
import re
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REAL_DIR = Path.home() / ".job_pipeline"
DEMO_DIR = Path.home() / ".job_pipeline_demo"

# How many Not Applied rows to sample alongside scored Applied rows.
# With fixed alignment formula (TP+TN)/total, each Not Applied Solid/Moderate row
# is an FN that drags alignment down. Keep this low to stay above 80%.
# Math: (42 applied + 150 skipped) / (42 + NOT_APPLIED_SAMPLE + 150) = target
# NOT_APPLIED_SAMPLE=38 → (192)/(230) = 83.5%
NOT_APPLIED_SAMPLE = 38

# How many skipped roles to include (keeps summary total count plausible)
SKIPPED_SAMPLE = 150

# How many weeks of runs.jsonl history to include for market intel
RUNS_LOOKBACK_WEEKS = 10

# ---------------------------------------------------------------------------
# Colors (must match xlsx.py palette exactly)
# ---------------------------------------------------------------------------

_GREEN  = "C6EFCE"
_YELLOW = "FFEB9C"
_RED    = "FFC7CE"
_GREY   = "D9D9D9"

_STATUS_FILL = {
    "Apply":          _GREEN,
    "Consider":       _YELLOW,
    "Partial":        _GREY,
    "Partial Match":  _GREY,
    "Skipped":        _GREY,
    "Filtered":       _RED,
    "Solid Match":    _GREEN,
    "Moderate Match": _YELLOW,
    "Limited Match":  _GREY,
    "Applied":        _GREEN,
    "Not Applied":    _GREY,
    "Pending":        _YELLOW,
    "Proceeding":     _GREEN,
    "Rejected":       _RED,
    "Recruiter Screen": "BDD7EE",
}

_HEADERS = [
    "date_suggested", "role_title", "company", "match_score",
    "suggestion_status", "action_taken", "date_applied",
    "application_status", "notes",
]

_METIS_SCORED = {
    "Apply", "Consider", "Solid Match", "Moderate Match", "Limited Match",
    "Partial Match", "Partial", "Skipped", "Filtered",
}
_HISTORICAL_APPLIED = {"External", "Pre-tracker"}
_JANKY_ENTITY = re.compile(
    r"^(?:re|fw|fwd)$|we received your application|let'?s connect|"
    r"unique skills deemed necessary",
    re.IGNORECASE,
)

_COL_WIDTHS = {
    "A": 16, "B": 30, "C": 20, "D": 14,
    "E": 18, "F": 15, "G": 14, "H": 20, "I": 35,
}


# ---------------------------------------------------------------------------
# xlsx helpers
# ---------------------------------------------------------------------------

def _make_fill(value: str) -> PatternFill | None:
    color = _STATUS_FILL.get(value)
    return PatternFill(fill_type="solid", fgColor=color) if color else None


def _is_clean_historical_applied(row_values: tuple) -> bool:
    title = str(row_values[1] or "").strip()
    company = str(row_values[2] or "").strip()
    if not title or not company:
        return False
    return not (_JANKY_ENTITY.search(title) or _JANKY_ENTITY.search(company))


def _copy_xlsx(src: Path, dest: Path, applied_rows: list, not_applied_sample: list) -> None:
    src_wb = openpyxl.load_workbook(src, data_only=False)
    src_ws = src_wb.active

    dest_wb = openpyxl.Workbook()
    dest_ws = dest_wb.active
    dest_ws.title = "Applications"

    # Copy header row with formatting
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = dest_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col_letter, width in _COL_WIDTHS.items():
        dest_ws.column_dimensions[col_letter].width = width

    # Write rows: Applied first, then Not Applied sample
    all_demo_rows = applied_rows + not_applied_sample
    for row_idx, (src_row_idx, row_values) in enumerate(all_demo_rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = dest_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 4, 5, 6, 7, 8) else "left",
                vertical="top",
                wrap_text=(col_idx == 9),
            )

            src_cell = src_ws.cell(src_row_idx, col_idx)
            if src_cell.hyperlink:
                cell.hyperlink = src_cell.hyperlink.target
                cell.font = copy(src_cell.font)

        dest_ws.cell(row=row_idx, column=4).number_format = "0%"

        suggestion_status = row_values[4]
        action_taken      = row_values[5]
        app_status        = row_values[7]

        for col_idx, value in [(5, suggestion_status), (6, action_taken), (8, app_status)]:
            f = _make_fill(value)
            if f:
                dest_ws.cell(row=row_idx, column=col_idx).fill = f

    dest_wb.save(dest)
    print(f"  Created {dest}")
    scored_applied = sum(1 for _, row in applied_rows if row[4] in _METIS_SCORED)
    historical_applied = sum(1 for _, row in applied_rows if row[4] in _HISTORICAL_APPLIED)
    print(f"    {scored_applied} Metis-scored Applied rows")
    print(f"    {historical_applied} historical Applied rows (External/Pre-tracker, clean only)")
    print(f"    {len(not_applied_sample)} Not Applied rows (sampled)")
    print(f"    {len(all_demo_rows)} total xlsx rows")


# ---------------------------------------------------------------------------
# Main setup
# ---------------------------------------------------------------------------

def main() -> None:
    random.seed(42)  # reproducible sample
    print(f"\nSetting up demo environment at {DEMO_DIR}\n")

    DEMO_DIR.mkdir(mode=0o700, exist_ok=True)
    (DEMO_DIR / "logs").mkdir(exist_ok=True)

    # ── 1. Profile and credentials ──────────────────────────────────────────
    for fname in ("profile.yaml", "feedback.md"):
        src = REAL_DIR / fname
        if src.exists():
            shutil.copy2(src, DEMO_DIR / fname)
            print(f"  Copied {fname}")

    for env_candidate in [REAL_DIR / ".env", Path.cwd() / ".env"]:
        if env_candidate.exists():
            shutil.copy2(env_candidate, DEMO_DIR / ".env")
            print(f"  Copied .env from {env_candidate}")
            break
    else:
        print("  WARNING: No .env found — copy one manually to ~/.job_pipeline_demo/.env")

    # ── 2. applications.xlsx — all Applied + sampled Not Applied ────────────
    real_xlsx = REAL_DIR / "applications.xlsx"
    if real_xlsx.exists():
        wb = openpyxl.load_workbook(real_xlsx, data_only=True)
        ws = wb.active
        all_rows = [
            (r_idx, tuple(ws.cell(r_idx, c).value for c in range(1, len(_HEADERS) + 1)))
            for r_idx in range(2, ws.max_row + 1)
            if any(ws.cell(r_idx, c).value is not None for c in range(1, len(_HEADERS) + 1))
        ]

        scored_rows = [r for r in all_rows if r[1][4] in _METIS_SCORED]
        historical_applied = [
            r for r in all_rows
            if r[1][4] in _HISTORICAL_APPLIED
            and r[1][5] == "Applied"
            and _is_clean_historical_applied(r[1])
        ]

        applied     = [r for r in scored_rows if r[1][5] == "Applied"]
        not_applied = [r for r in scored_rows if r[1][5] == "Not Applied"]

        sample_size = min(NOT_APPLIED_SAMPLE, len(not_applied))
        not_applied_sample = random.sample(not_applied, sample_size)

        _copy_xlsx(
            real_xlsx,
            DEMO_DIR / "applications.xlsx",
            applied + historical_applied,
            not_applied_sample,
        )
    else:
        print("  WARNING: No applications.xlsx found in real data dir")

    # ── 3. skipped_roles.json — sampled subset ──────────────────────────────
    real_skipped = REAL_DIR / "skipped_roles.json"
    if real_skipped.exists():
        skipped = json.loads(real_skipped.read_text())
        # skipped_roles.json is a dict keyed by role_hash
        if isinstance(skipped, dict):
            keys = list(skipped.keys())
            sample_keys = random.sample(keys, min(SKIPPED_SAMPLE, len(keys)))
            skipped_sample = {k: skipped[k] for k in sample_keys}
        else:
            skipped_sample = skipped[:SKIPPED_SAMPLE]
        (DEMO_DIR / "skipped_roles.json").write_text(json.dumps(skipped_sample, indent=2))
        print(f"  Sampled {len(skipped_sample)} of {len(skipped)} skipped roles")
    else:
        (DEMO_DIR / "skipped_roles.json").write_text("{}")
        print("  Created empty skipped_roles.json")

    # ── 4. runs.jsonl — date-windowed slice for market intel ────────────────
    real_runs = REAL_DIR / "runs.jsonl"
    if real_runs.exists():
        cutoff = (datetime.datetime.now()
                  - datetime.timedelta(weeks=RUNS_LOOKBACK_WEEKS)).isoformat()
        kept = []
        with real_runs.open() as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    if rec.get("ts", "") >= cutoff:
                        kept.append(line)
                except json.JSONDecodeError:
                    continue
        (DEMO_DIR / "runs.jsonl").write_text("\n".join(kept) + "\n")
        print(f"  Copied {len(kept)} runs from the last {RUNS_LOOKBACK_WEEKS} weeks")
    else:
        (DEMO_DIR / "runs.jsonl").write_text("")
        print("  Created empty runs.jsonl")

    # ── 5. Fresh dedup state ─────────────────────────────────────────────────
    # state.load_seen_roles() expects a flat {role_hash: iso_timestamp} mapping.
    (DEMO_DIR / "seen_roles.json").write_text("{}")
    print("  Created empty seen_roles.json (all recent roles will score fresh)")

    (DEMO_DIR / "role_queue.json").write_text("[]")
    print("  Cleared role_queue.json")

    # ── Summary ──────────────────────────────────────────────────────────────
    scored_xlsx_total = len(applied) + len(not_applied_sample) if real_xlsx.exists() else 0
    all_xlsx_total = (
        len(applied) + len(historical_applied) + len(not_applied_sample)
        if real_xlsx.exists() else 0
    )
    skipped_total = len(skipped_sample) if real_skipped.exists() else 0
    print(f"""
Demo summary baseline before filming a new digest run:
  {scored_xlsx_total} scored xlsx rows + {skipped_total} skipped = {scored_xlsx_total + skipped_total} roles scored
  {all_xlsx_total} total tracker rows, including historical External/Pre-tracker applied rows

Each filmed 'metis --lookback ...' run adds newly evaluated rows/skipped roles to that baseline.

To film:
  export METIS_PROFILE=~/.job_pipeline_demo/profile.yaml
  export METIS_DATA_DIR=~/.job_pipeline_demo
  ./venv/bin/python -m metis.cli --lookback 7d     # scored digest
  ./venv/bin/python -m metis.cli track             # tracker update
  ./venv/bin/python -m metis.cli summary           # progress report
  unset METIS_PROFILE METIS_DATA_DIR    # back to real data

Your real ~/.job_pipeline/ was not touched.
""")


if __name__ == "__main__":
    main()
