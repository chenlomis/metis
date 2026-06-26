"""scorerole/track.py — thin orchestrator for the track pipeline stage.

Sub-modules own the actual logic:
  track_imap.py   — IMAP fetch + email decode helpers
  track_parse.py  — email classification + entity extraction
  track_write.py  — tracker fuzzy match + xlsx write operations

This module wires them together and exposes run_track() and backfill_from_digests().
"""
from __future__ import annotations

import datetime
import email as email_lib
import imaplib
import logging
import sys
import time

log = logging.getLogger(__name__)

# Re-exports for backward-compat (tests import from scorerole.track)
from .track_parse import classify_email, _LLM_VALID_CLASSES  # noqa: F401
from .track_write import update_recruiter_screen              # noqa: F401

from .track_imap  import (
    fetch_candidate_emails,
    _safe_decode,
    _IMAP_MAX_RETRIES,
    _IMAP_RETRY_DELAY,
)
from .track_parse import parse_email
from .track_write import (
    find_tracker_row,
    update_confirmation,
    update_rejection,
    create_skipped_row,
    create_backfill_row,
    _parse_digest_html,
)


# ---------------------------------------------------------------------------
# LLM client factory
# ---------------------------------------------------------------------------

def _build_llm_client(api_key: str | None):
    """Return an Anthropic client if profile.yaml has track.llm_fallback: true.

    Returns None when the flag is absent, false, or the api_key is missing —
    so classify_email() runs phrase-only (safe for OSS users with no key).
    """
    if not api_key:
        return None
    try:
        from .profile import load_profile_yaml
        profile = load_profile_yaml() or {}
    except Exception:
        return None
    if not profile.get("track", {}).get("llm_fallback", False):
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        log.info("track: LLM fallback enabled (Haiku) for unknown emails")
        return client
    except Exception as exc:
        log.warning("track: could not build LLM client (%s) — running phrase-only", exc)
        return None


# ---------------------------------------------------------------------------
# Digest backfill
# ---------------------------------------------------------------------------

def backfill_from_digests(
    gmail_address: str,
    app_password: str,
    since_dt: datetime.datetime,
) -> int:
    """Fetch past digest emails from Gmail and write any new rows to the tracker.

    Only writes Apply/Consider roles not already in the tracker (dedup by
    normalized title+company). Returns the number of new rows added.
    """
    from email.utils import parsedate_to_datetime
    from .xlsx import write_to_tracker

    since_str    = since_dt.strftime("%d-%b-%Y")
    added_total  = 0
    n_digests    = 0

    for attempt in range(1, _IMAP_MAX_RETRIES + 1):
        try:
            with imaplib.IMAP4_SSL("imap.gmail.com") as imap:
                imap.login(gmail_address, app_password)
                imap.select("INBOX")

                _, data = imap.search(None, f'SINCE {since_str} SUBJECT "Personalized Job Alert Digest"')
                if not data or not data[0]:
                    log.info("track: no digest emails found since %s", since_str)
                    return 0

                digest_ids = data[0].split()
                n_digests  = len(digest_ids)
                log.info("track: found %d digest email(s) to backfill from", n_digests)

                for msg_id in digest_ids:
                    _, raw = imap.fetch(msg_id, "(RFC822)")
                    if not raw or not raw[0]:
                        continue
                    raw_bytes = raw[0][1] if isinstance(raw[0], tuple) else raw[0]
                    msg = email_lib.message_from_bytes(raw_bytes)

                    date_header = msg.get("Date", "")
                    try:
                        email_date = parsedate_to_datetime(date_header).date().isoformat()
                    except Exception:
                        email_date = datetime.date.today().isoformat()

                    html_body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/html":
                                payload = part.get_payload(decode=True)
                                if payload:
                                    html_body = _safe_decode(payload, part.get_content_charset())
                                    break
                    else:
                        if msg.get_content_type() == "text/html":
                            payload = msg.get_payload(decode=True)
                            if payload:
                                html_body = _safe_decode(payload, msg.get_content_charset())

                    if not html_body:
                        log.debug("track: digest %s has no HTML body — skipping", email_date)
                        continue

                    jobs = _parse_digest_html(html_body, email_date)
                    if jobs:
                        log.info("track: digest %s → %d role(s) parsed", email_date, len(jobs))
                        write_to_tracker(jobs, run_date=email_date)
                        added_total += len(jobs)
                    else:
                        log.warning("track: digest %s → 0 roles parsed (HTML structure mismatch?)", email_date)
            break
        except OSError as e:
            if attempt < _IMAP_MAX_RETRIES:
                log.warning(
                    "track: IMAP connect failed (attempt %d/%d): %s — retrying in %ds…",
                    attempt, _IMAP_MAX_RETRIES, e, _IMAP_RETRY_DELAY,
                )
                time.sleep(_IMAP_RETRY_DELAY)
            else:
                log.error("track: IMAP connect failed after %d attempts: %s — skipping backfill",
                          _IMAP_MAX_RETRIES, e)
                return 0

    log.info("track: digest backfill complete — %d role(s) eligible across %d digest(s)",
             added_total, n_digests)
    return added_total


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_track(
    gmail_address: str,
    app_password: str,
    since_dt: datetime.datetime,
    dry_run: bool = False,
    api_key: str | None = None,
) -> None:
    """Parse confirmation/rejection emails and update the Applications tracker."""
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    from .xlsx  import TRACKER_PATH
    from .state import lookup_skipped_role, promote_skipped_role

    llm_client = _build_llm_client(api_key)

    log.info("track: step 1 — backfilling tracker from digest emails…")
    backfill_from_digests(gmail_address, app_password, since_dt)

    applied_companies: set[str] = set()
    if TRACKER_PATH.exists():
        _wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
        _ws = _wb.active
        for _r in range(2, _ws.max_row + 1):
            if _ws.cell(_r, 6).value == "Applied":
                co = _ws.cell(_r, 3).value
                if co:
                    applied_companies.add(str(co))
        _wb.close()
    log.info("track: %d applied companies loaded for direct-outreach search", len(applied_companies))

    emails = fetch_candidate_emails(gmail_address, app_password, since_dt, applied_companies=applied_companies)
    if not emails:
        log.info("track: no candidate emails found in lookback window.")
        return

    parsed_emails = [parse_email(e, llm_client=llm_client) for e in emails]

    actionable = [p for p in parsed_emails
                  if p["classification"] in ("confirmation", "rejection", "recruiter_screen")
                  and p["company"]]
    unknown    = [p for p in parsed_emails if p["classification"] == "unknown"]

    log.info("track: %d actionable (%d confirmation, %d rejection, %d recruiter_screen), %d unknown",
             len(actionable),
             sum(1 for p in actionable if p["classification"] == "confirmation"),
             sum(1 for p in actionable if p["classification"] == "rejection"),
             sum(1 for p in actionable if p["classification"] == "recruiter_screen"),
             len(unknown))

    if unknown:
        for u in unknown:
            log.warning("track: unclassified — '%s' (company=%s)", u["subject"], u["company"])

    if dry_run:
        for p in parsed_emails:
            print(f"[{p['classification'].upper():12}] {p['company'] or '?':25} | {p['role'] or '?'}")
        return

    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        from .xlsx import _write_header, _set_column_widths
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        _write_header(ws)
        _set_column_widths(ws)
        ws.freeze_panes = "A2"
        log.info("track: created new tracker at %s", TRACKER_PATH)

    changed = 0

    for parsed in actionable:
        company = parsed["company"]
        role    = parsed["role"]
        kind    = parsed["classification"]

        row_idx = find_tracker_row(ws, company, role)

        if row_idx is None:
            if kind == "confirmation":
                skipped_meta = lookup_skipped_role(company, role or "")
                if skipped_meta:
                    log.info("track: backfill (skipped) — %s / %s", company, role)
                    create_skipped_row(ws, parsed, skipped_meta)
                    promote_skipped_role(company, role or "")
                else:
                    log.info("track: backfill (pre-tracker) — %s / %s", company, role or "?")
                    create_backfill_row(ws, parsed)
                changed += 1
            else:
                log.info("track: skip rejection (no tracker row) — %s", company)
            continue

        current_action = ws.cell(row_idx, 6).value
        current_status = ws.cell(row_idx, 8).value

        if kind == "confirmation" and current_action != "Applied":
            log.info("track: ✓ confirmation — %s / %s → Applied + Pending", company, role or "?")
            update_confirmation(ws, row_idx, parsed["date"])
            changed += 1
        elif kind == "confirmation" and current_action == "Applied":
            log.info("track: skip confirmation (already Applied) — %s", company)
        elif kind == "rejection" and current_status != "Rejected":
            log.info("track: ✗ rejection  — %s / %s → Rejected", company, role or "?")
            update_rejection(ws, row_idx)
            changed += 1
        elif kind == "rejection" and current_status == "Rejected":
            log.info("track: skip rejection (already Rejected) — %s", company)
        elif kind == "recruiter_screen" and current_status not in ("Recruiter Screen", "Rejected"):
            log.info("track: ★ recruiter screen — %s / %s → Recruiter Screen", company, role or "?")
            update_recruiter_screen(ws, row_idx)
            changed += 1
        elif kind == "recruiter_screen":
            log.info("track: skip recruiter screen (already %s) — %s", current_status, company)

    from .xlsx import _sort_rows_by_date
    _sort_rows_by_date(ws)

    if changed:
        wb.save(TRACKER_PATH)
        TRACKER_PATH.chmod(0o600)
        log.info("track: updated %d row(s) in %s", changed, TRACKER_PATH)
        print(f"  Tracker → {TRACKER_PATH}")
        if sys.stdout.isatty():
            import subprocess
            subprocess.Popen(["open", str(TRACKER_PATH)])
    else:
        wb.save(TRACKER_PATH)
        TRACKER_PATH.chmod(0o600)
        log.info("track: no updates needed.")
