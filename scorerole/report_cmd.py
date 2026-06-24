"""scorerole report — generate and deliver the Scorerole market report.

Usage:
    scorerole report                      # generate + send to email
    scorerole report --lookback 60d       # scope market intel to 60 days (default: 30d)
    scorerole report --output report.html # save as HTML instead of sending
    scorerole report --output report.pdf  # save as PDF instead of sending
"""

from __future__ import annotations

import datetime
import math
import os
import smtplib
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Design tokens — Ink & Slate palette (matches Job Alert Digest)
# ---------------------------------------------------------------------------
FONT      = "-apple-system, BlinkMacSystemFont, 'Segoe UI', 'Helvetica Neue', Arial, sans-serif"
C_HEADING = "#1f2118"
C_MUTED   = "#72716d"
C_SUBTLE  = "#aaaaaa"
APPLY_BG  = "#eef2ee"; APPLY_NUM  = "#2d5a2d"
CONSID_BG = "#faeeda"; CONSID_NUM = "#854f0b"
TOTAL_BG  = "#f5f5f3"; TOTAL_NUM  = "#1f2118"
RED_BG    = "#f2eeee"; RED_NUM    = "#8b2e2e"
CARD_BG   = "#ffffff"
BORDER    = "#eeece5"
BODY_BG   = "#f0f0ef"

# suggestion_status values — support both old and new naming
_SOLID    = {"Apply", "Solid", "Solid Match"}
_MODERATE = {"Consider", "Moderate", "Moderate Match"}
_PARTIAL  = {"Skipped", "Partial", "Partial Match"}
_SCORED   = _SOLID | _MODERATE | _PARTIAL


# ---------------------------------------------------------------------------
# Data layer
# ---------------------------------------------------------------------------

def _coerce_date(val) -> datetime.date | None:
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        try:
            return datetime.date.fromisoformat(str(val)[:10])
        except ValueError:
            return None
    return None


def load_report_data(tracker_path: Path) -> dict:
    """Read applications.xlsx and compute all metrics for the report."""
    if not tracker_path.exists():
        return {
            "roles_scored": 0, "total_applied": 0, "time_saved_h": 0.0,
            "alignment": 0.0, "TP": 0, "TN": 0, "align_base": 0,
            "solid": 0, "moderate": 0, "partial": 0,
            "pending": 0, "screens": 0, "rejections": 0, "daily": {},
        }
    wb = openpyxl.load_workbook(tracker_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = [
        {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        for r in range(2, ws.max_row + 1)
    ]
    wb.close()

    scored     = [r for r in rows if r.get("suggestion_status") in _SCORED]
    solid_rows = [r for r in scored if r.get("suggestion_status") in _SOLID]
    mod_rows   = [r for r in scored if r.get("suggestion_status") in _MODERATE]
    part_rows  = [r for r in scored if r.get("suggestion_status") in _PARTIAL]
    applied    = [r for r in rows if r.get("action_taken") == "Applied"]

    # Alignment (confusion matrix):
    #   TP = Solid suggestion + user Applied
    #   TN = Moderate suggestion + user NOT Applied
    align_base = solid_rows + mod_rows
    TP = sum(1 for r in solid_rows if r.get("action_taken") == "Applied")
    TN = sum(1 for r in mod_rows   if r.get("action_taken") != "Applied")
    base      = len(align_base)
    alignment = (TP + TN) / base if base else 0.0

    pending    = [r for r in applied if r.get("application_status") == "Pending"]
    screens    = [r for r in applied if r.get("application_status") == "Recruiter Screen"]
    rejections = [r for r in applied if r.get("application_status") == "Rejected"]

    # Daily application volume
    daily: dict[datetime.date, int] = defaultdict(int)
    for r in applied:
        d = _coerce_date(r.get("date_applied"))
        if d:
            daily[d] += 1

    return {
        "roles_scored":  len(scored),
        "total_applied": len(applied),
        "time_saved_h":  round(len(scored) / 60, 1),
        "alignment":     alignment,
        "TP": TP, "TN": TN, "align_base": base,
        "solid":    len(solid_rows),
        "moderate": len(mod_rows),
        "partial":  len(part_rows),
        "pending":    len(pending),
        "screens":    len(screens),
        "rejections": len(rejections),
        "daily": dict(daily),
    }


# ---------------------------------------------------------------------------
# HTML rendering helpers
# ---------------------------------------------------------------------------

def _pct(n: int, total: int) -> str:
    if not total:
        return "—"
    return f"{n} / {round(n / total * 100)}%"


def _section_h(t: str) -> str:
    return (
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:6px;">'
        f'<tr><td style="font-family:{FONT};font-size:14px;font-weight:600;color:{C_HEADING};'
        f'padding-bottom:6px;border-bottom:1px solid {BORDER};">{t}</td></tr></table>'
    )


def _sub_h(t: str) -> str:
    return (
        f'<p style="font-family:{FONT};font-size:10px;font-weight:600;text-transform:uppercase;'
        f'letter-spacing:0.05em;color:{C_SUBTLE};margin:0 0 8px 0;">{t}</p>'
    )


def _spacer(h: int = 20) -> str:
    return f'<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td height="{h}"></td></tr></table>'


def _tiles(items: list) -> str:
    cells = ""
    for label, num, sub, bg, color in items:
        sub_html = (
            f'<div style="font-family:{FONT};font-size:10px;color:{color};margin-top:2px;opacity:0.6;">{sub}</div>'
            if sub else ""
        )
        cells += (
            f'<td width="25%" style="padding:0 4px;" valign="top">'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="background:{bg};border-radius:6px;table-layout:fixed;">'
            f'<tr><td style="padding:10px 10px 8px;text-align:center;">'
            f'<div style="font-family:{FONT};font-size:28px;font-weight:700;color:{color};line-height:1.1;">{num}</div>'
            f'<div style="font-family:{FONT};font-size:11px;font-weight:500;color:{color};margin-top:3px;opacity:0.75;">{label}</div>'
            f'{sub_html}'
            f'</td></tr></table></td>'
        )
    return (
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        f'style="table-layout:fixed;border-collapse:separate;border-spacing:0;">'
        f'<tr>{cells}</tr></table>'
    )


def _green_th(cols: list, aligns: list | None = None) -> str:
    if aligns is None:
        aligns = ["left"] * len(cols)
    cells = "".join(
        f'<td style="font-family:{FONT};font-size:10px;font-weight:600;color:{APPLY_NUM};'
        f'background:{APPLY_BG};padding:7px 10px;text-transform:uppercase;'
        f'letter-spacing:0.04em;text-align:{a};">{c}</td>'
        for c, a in zip(cols, aligns)
    )
    return f"<tr>{cells}</tr>"


def _td(content: str, width: str | None = None, align: str = "left",
        muted: bool = False, size: int = 13) -> str:
    color = C_MUTED if muted else C_HEADING
    w = f'width="{width}"' if width else ""
    return (
        f'<td {w} style="font-family:{FONT};font-size:{size}px;font-weight:400;color:{color};'
        f'padding:7px 10px;line-height:1.4;text-align:{align};border-bottom:1px solid {BORDER};">'
        f'{content}</td>'
    )


def _fit_pill(label: str) -> str:
    if "Strong" in label:
        return f'<span style="background:#eef2ee;color:#2d5a2d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">🟢 Strong</span>'
    if "Partial" in label:
        return f'<span style="background:#faeeda;color:#854f0b;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">🟡 Partial</span>'
    return f'<span style="background:#f2eeee;color:#8b2e2e;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">🔴 Gap</span>'


def _trend_pill(label: str) -> str:
    if "Trending" in label:
        return f'<span style="background:#eef2ee;color:#2d5a2d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Trending ↑</span>'
    if "Niche" in label:
        return f'<span style="background:#f5f5f3;color:#72716d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Niche</span>'
    return f'<span style="background:#f5f5f3;color:#72716d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Established</span>'


def _bar_chart(daily: dict) -> str:
    if not daily:
        return '<p style="font-family:{FONT};font-size:12px;color:{C_SUBTLE};">No application data.</p>'

    start_d = min(daily.keys())
    end_d   = max(daily.keys())
    all_days: list[datetime.date] = []
    d = start_d
    while d <= end_d:
        all_days.append(d)
        d += datetime.timedelta(days=1)

    max_v = max(daily.values(), default=1)
    H     = 8
    col_w = 15

    bar_rows = ""
    for level in range(H, 0, -1):
        cells = ""
        for day in all_days:
            val = daily.get(day, 0)
            bh  = max(1, math.ceil(val / max_v * H)) if val > 0 else 0
            filled   = val > 0 and level <= bh
            top_cell = val > 0 and level == bh
            bg     = APPLY_NUM if filled else "transparent"
            radius = "border-radius:2px 2px 0 0;" if top_cell else ""
            cells += (
                f'<td width="{col_w}" style="width:{col_w}px;height:10px;'
                f'background:{bg};{radius}padding:0;"></td>'
            )
        bar_rows += f"<tr>{cells}</tr>\n"

    label_cells = ""
    for i, day in enumerate(all_days):
        if day == start_d or day.day == 1:
            lbl = day.strftime("%b %-d")
        elif i % 4 == 0:
            lbl = str(day.day)
        else:
            lbl = ""
        label_cells += (
            f'<td width="{col_w}" style="width:{col_w}px;font-family:{FONT};font-size:9px;'
            f'color:{C_SUBTLE};text-align:center;padding-top:3px;white-space:nowrap;'
            f'overflow:hidden;">{lbl}</td>'
        )

    peak_day  = max(daily, key=daily.get)
    peak_val  = daily[peak_day]
    span_days = (end_d - start_d).days + 1

    chart = (
        f'<table cellpadding="0" cellspacing="2" border="0" '
        f'style="border-collapse:separate;border-spacing:2px 0;">'
        f'{bar_rows}<tr>{label_cells}</tr></table>'
        f'<p style="font-family:{FONT};font-size:10px;color:{C_SUBTLE};'
        f'margin:4px 0 0;text-align:right;">'
        f'{start_d.strftime("%b %-d")} – {end_d.strftime("%b %-d")} · '
        f'{span_days} days · peak {peak_day.strftime("%b %-d")} ({peak_val})</p>'
    )
    return chart


# ---------------------------------------------------------------------------
# Market intelligence (§5–§6) — LLM-generated, placeholder until wired
# ---------------------------------------------------------------------------

_MARKET_INTEL_PLACEHOLDER = """
<tr>
  <td colspan="3" style="font-family:{FONT};font-size:12px;color:{C_MUTED};
      padding:16px 10px;text-align:center;font-style:italic;">
    Market intelligence analysis coming in a future release.
  </td>
</tr>
""".format(FONT=FONT, C_MUTED=C_MUTED)


# ---------------------------------------------------------------------------
# Full HTML assembly
# ---------------------------------------------------------------------------

def render_html(data: dict) -> str:
    now_h = datetime.datetime.now().hour
    if now_h < 12:
        greeting = "Good morning"
    elif now_h < 18:
        greeting = "Good afternoon"
    else:
        greeting = "Good evening"

    daily = data["daily"]
    if daily:
        start_d = min(daily.keys())
        end_d   = max(daily.keys())
        date_range = f'{start_d.strftime("%b %-d")} &ndash; {end_d.strftime("%b %-d, %Y")}'
    else:
        date_range = "No data"

    n_scored  = data["roles_scored"]
    n_applied = data["total_applied"]
    alignment = data["alignment"]
    align_pct = f"{alignment:.1%}"
    align_bg  = APPLY_BG  if alignment >= 0.80 else CONSID_BG
    align_fg  = APPLY_NUM if alignment >= 0.80 else CONSID_NUM

    # Alignment banner (shown when < 80%)
    diag_banner = ""
    if alignment < 0.80:
        diag_banner = (
            f'{_spacer(8)}'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0">'
            f'<tr><td style="border-left:3px solid {CONSID_NUM};background:{CONSID_BG};'
            f'padding:8px 12px;border-radius:0 4px 4px 0;">'
            f'<span style="font-family:{FONT};font-size:12px;color:{CONSID_NUM};">'
            f'Recommendation alignment is {align_pct} — below the 80% threshold. '
            f'Your application choices diverged from scorer suggestions more than expected. '
            f'Consider reviewing your profile weights.</span>'
            f'</td></tr></table>'
        )

    solid    = data["solid"]
    moderate = data["moderate"]
    partial  = data["partial"]
    pending  = data["pending"]
    screens  = data["screens"]
    rejects  = data["rejections"]

    html = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:{BODY_BG};font-family:{FONT};">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:{BODY_BG};">
<tr><td align="center" style="padding:24px 16px;">
<table width="600" cellpadding="0" cellspacing="0" border="0" style="background:{CARD_BG};border-radius:10px;overflow:hidden;">

<!-- HEADER -->
<tr><td style="padding:28px 28px 20px;">
  <p style="font-family:{FONT};font-size:18px;font-weight:600;color:{C_HEADING};margin:0 0 4px;line-height:1.3;">{greeting}, Lomis</p>
  <p style="font-family:{FONT};font-size:13px;color:{C_MUTED};margin:0;line-height:1.5;">Scorerole report &middot; {date_range} &middot; {n_scored} roles evaluated</p>
</td></tr>

<!-- §2 EXECUTIVE ROI -->
<tr><td style="padding:0 28px 0;">
{_section_h("Executive ROI")}
{_tiles([
    ("Roles Scored",  str(n_scored),                  None, TOTAL_BG,  TOTAL_NUM),
    ("Total Applied", str(n_applied),                 None, TOTAL_BG,  TOTAL_NUM),
    ("Time Saved",    f"{data['time_saved_h']}h",     None, APPLY_BG,  APPLY_NUM),
    ("Alignment",     align_pct,                      None, align_bg,  align_fg),
])}
{diag_banner}
{_spacer(6)}
<p style="font-family:{FONT};font-size:10px;color:{C_SUBTLE};margin:0;">
  Time saved estimated at 1 min/role vs. manual JD review &middot;
  Alignment = (TP + TN) / scored &middot; base = {data['align_base']} (Solid + Moderate rows only)
</p>
</td></tr>

{_spacer(20)}

<!-- §3 SCORE EVAL -->
<tr><td style="padding:0 28px 0;">
{_section_h("Score Eval Distribution")}
{_tiles([
    ("Total Scored",   str(n_scored), "100%",              TOTAL_BG,  TOTAL_NUM),
    ("Solid Match",    str(solid),    _pct(solid,    n_scored), APPLY_BG,  APPLY_NUM),
    ("Moderate Match", str(moderate), _pct(moderate, n_scored), CONSID_BG, CONSID_NUM),
    ("Partial Match",  str(partial),  _pct(partial,  n_scored), TOTAL_BG,  TOTAL_NUM),
])}
</td></tr>

{_spacer(20)}

<!-- §4 APPLICATION PIPELINE -->
<tr><td style="padding:0 28px 0;">
{_section_h("Application Pipeline")}
{_tiles([
    ("Total Applied", str(n_applied), "100%",                TOTAL_BG,  TOTAL_NUM),
    ("Pending",       str(pending),   _pct(pending,  n_applied), CONSID_BG, CONSID_NUM),
    ("Screens",       str(screens),   _pct(screens,  n_applied), APPLY_BG,  APPLY_NUM),
    ("Rejections",    str(rejects),   _pct(rejects,  n_applied), RED_BG,    RED_NUM),
])}
{_spacer(16)}
{_sub_h("Daily Application Volume")}
{_bar_chart(daily)}
</td></tr>

{_spacer(20)}

<!-- §5 CORE STRENGTHS -->
<tr><td style="padding:0 28px 0;">
{_section_h("Core Strengths Alignment")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};">
{_green_th(["Profile Signal", "Evidence Found in JDs", "Companies in Batch Hiring This"])}
{_MARKET_INTEL_PLACEHOLDER}
</table>
</td></tr>

{_spacer(20)}

<!-- §6 MARKET LANDSCAPE -->
<tr><td style="padding:0 28px 0;">
{_section_h("Market Landscape")}
{_sub_h("Table A — Skills & Domains")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};">
{_green_th(["Skill / Domain", "Role Freq", "Trend", "Profile Fit"], ["left","center","center","center"])}
{_MARKET_INTEL_PLACEHOLDER}
</table>
{_spacer(14)}
{_sub_h("Table B — Verticals & Markets")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};">
{_green_th(["Company Vertical", "Role Freq", "Domain Expertise", "Profile Fit"], ["left","center","center","center"])}
{_MARKET_INTEL_PLACEHOLDER}
</table>
</td></tr>

{_spacer(28)}

<!-- FOOTER -->
<tr><td style="padding:0 28px 28px;">
  <p style="font-family:{FONT};font-size:10px;color:{C_SUBTLE};margin:0;line-height:1.6;border-top:1px solid {BORDER};padding-top:14px;">
    Generated by scorerole &middot; {datetime.date.today().strftime("%b %-d, %Y")} &middot;
    {n_scored} roles evaluated
  </p>
</td></tr>

</table></td></tr></table>
</body></html>"""

    return html


# ---------------------------------------------------------------------------
# Output — send email or save file
# ---------------------------------------------------------------------------

def _send_email(html: str, gmail_address: str, app_password: str, preview: bool) -> None:
    prefix = "[DRAFT PREVIEW] " if preview else ""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"{prefix}Scorerole Market Report — {datetime.date.today()}"
    msg["From"]    = gmail_address
    msg["To"]      = gmail_address
    msg.attach(MIMEText(html, "html"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(gmail_address, app_password)
        smtp.sendmail(gmail_address, gmail_address, msg.as_string())


def _save_file(html: str, output_path: Path) -> None:
    suffix = output_path.suffix.lower()
    if suffix == ".pdf":
        try:
            from weasyprint import HTML as WP_HTML
        except ImportError:
            raise SystemExit(
                "weasyprint is required for PDF output: pip install weasyprint\n"
                "On macOS also run: brew install pango"
            )
        WP_HTML(string=html).write_pdf(str(output_path))
    else:
        output_path.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run_report(
    tracker_path: Path,
    gmail_address: str,
    app_password: str,
    output: str | None = None,
    preview: bool = True,
) -> None:
    """Generate the Scorerole report and send or save it."""
    data = load_report_data(tracker_path)
    html = render_html(data)

    if output:
        out_path = Path(output)
        _save_file(html, out_path)
        print(f"Report saved to {out_path.resolve()}")
    else:
        _send_email(html, gmail_address, app_password, preview=preview)
        label = "draft preview" if preview else "report"
        print(f"Scorerole {label} sent to {gmail_address}")
