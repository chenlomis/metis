"""tracker.py — write scored roles to the Applications xlsx tracker.

Called implicitly at the end of each scorerole run (after digest delivery).
Only Apply and Consider roles are written as new rows. Skipped roles are
persisted separately in skipped_roles.json via state.save_skipped_roles().

Dedup: rows are identified by normalized title+company (exact key match).
This is sufficient for the pipeline write path — fuzzy matching is deferred
to the confirmation-email parser (Phase 2), which needs to match ATS titles
that may differ from LinkedIn titles.
"""
from __future__ import annotations

import re, datetime, logging, os
from pathlib import Path

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

_DEFAULT_TRACKER = Path.home() / ".job_pipeline" / "applications.xlsx"
TRACKER_PATH = Path(os.getenv("TRACKER_PATH", str(_DEFAULT_TRACKER)))

# Column order — must match _COL_* indices and header row
_HEADERS = [
    "date_suggested",
    "role_title",
    "company",
    "match_score",
    "suggestion_status",
    "action_taken",
    "date_applied",
    "application_status",
    "notes",
]

# 1-based column indices (openpyxl is 1-indexed)
_COL_DATE_SUGGESTED    = 1
_COL_ROLE_TITLE        = 2
_COL_COMPANY           = 3
_COL_MATCH_SCORE       = 4
_COL_SUGGESTION_STATUS = 5
_COL_ACTION_TAKEN      = 6
_COL_DATE_APPLIED      = 7
_COL_APP_STATUS        = 8
_COL_NOTES             = 9

# ---------------------------------------------------------------------------
# Colors (openpyxl PatternFill hex values — no leading #)
# ---------------------------------------------------------------------------

_GREEN  = "C6EFCE"   # suggestion_status=Apply, action_taken=Applied, app_status=Proceeding
_YELLOW = "FFEB9C"   # suggestion_status=Consider, app_status=Pending
_RED    = "FFC7CE"   # app_status=Rejected
_GREY   = "D9D9D9"   # action_taken=Not Applied, suggestion_status=Skipped

_STATUS_FILL = {
    # suggestion_status
    "Apply":        _GREEN,
    "Consider":     _YELLOW,
    "Skipped":      _GREY,
    # action_taken
    "Applied":      _GREEN,
    "Not Applied":  _GREY,
    # application_status
    "Pending":      _YELLOW,
    "Proceeding":   _GREEN,
    "Rejected":     _RED,
}

# ---------------------------------------------------------------------------
# Input validation — guard against LinkedIn extractor mis-parsing
# ---------------------------------------------------------------------------

# Any real job title contains at least one of these role-level words.
# Company names and locations will not match.
_JOB_TITLE_KEYWORDS = re.compile(
    r"\b(?:manager|director|lead|head|officer|president|vp|principal|"
    r"staff|senior|associate|engineer|analyst|scientist|designer|"
    r"architect|specialist|coordinator|owner|operations|strategy|"
    r"product|program|project|general\s+manager|gm)\b",
    re.IGNORECASE,
)

# Company field looks like a location when the entire value is a city, state, or region.
# Uses ^…$ anchors to avoid flagging "San Francisco Health" as a location.
_COMPANY_IS_LOCATION = re.compile(
    r"^(?:"
    r"san\s+francisco|new\s+york(?:\s+city)?|los\s+angeles|seattle|"
    r"austin|boston|chicago|denver|atlanta|miami|new\s+jersey|"
    r"bay\s+area|silicon\s+valley|remote|united\s+states|"
    r"greater\s+\w+(?:\s+\w+)?\s+area|"
    r"[a-z\s]+,\s*(?:CA|NY|WA|TX|MA|IL|CO|GA|FL|OR|VA|NC|AZ|OH|PA|"
    r"NJ|MN|MI|MO|IN|TN|UT|MD|WI|SC|NV|CT|LA|AL|AR|IA|KS|KY|ME|MS|"
    r"MT|NE|NH|NM|ND|OK|RI|SD|VT|WV|WY|DC|HI|AK|ID|DE)"
    r")$",
    re.IGNORECASE,
)


def _is_plausible_job_row(title: str, company: str) -> bool:
    """Return False when title or company look like mis-parsed LinkedIn fields.

    Rejects rows where:
    - title contains no job-role keywords (likely a company name or location)
    - company field is a city, state, or region string (likely the location field)
    """
    if not _JOB_TITLE_KEYWORDS.search(title):
        return False
    if _COMPANY_IS_LOCATION.match(company.strip()):
        return False
    return True


def _sort_rows_by_date(ws) -> None:
    """Sort data rows by date_suggested (col A) descending — newest first.

    Reads all data rows into memory, clears them, then rewrites in sorted order.
    Header row (row 1) is untouched. Preserves fills, fonts, hyperlinks.
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from copy import copy

    max_row = ws.max_row
    if max_row < 3:
        return  # 0 or 1 data rows — nothing to sort

    # Capture every data row as a list of (value, style snapshot)
    data_rows = []
    for row_idx in range(2, max_row + 1):
        row_cells = []
        for col_idx in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row_idx, col_idx)
            row_cells.append({
                "value":      cell.value,
                "hyperlink":  cell.hyperlink,
                "fill":       copy(cell.fill),
                "font":       copy(cell.font),
                "number_fmt": cell.number_format,
                "alignment":  copy(cell.alignment),
            })
        data_rows.append(row_cells)

    # Sort by date in column A descending — rows without a date go to the bottom
    def _sort_key(row):
        val = row[0]["value"]
        return val if isinstance(val, str) and val else ""

    data_rows.sort(key=_sort_key, reverse=True)

    # Rewrite — clear then restore
    for row_idx, row_cells in enumerate(data_rows, start=2):
        for col_idx, snap in enumerate(row_cells, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value        = snap["value"]
            cell.fill         = snap["fill"]
            cell.font         = snap["font"]
            cell.number_format = snap["number_fmt"]
            cell.alignment    = snap["alignment"]
            if snap["hyperlink"]:
                cell.hyperlink = snap["hyperlink"]


def _norm_key(title: str, company: str) -> str:
    """Normalized dedup key — same logic as state._role_hash but human-readable."""
    return re.sub(r"[^a-z0-9]", "", (title + company).lower())


def _make_fill(hex_color: str):
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_row_styles(ws, row_idx: int, suggestion_status: str, action_taken: str) -> None:
    """Apply conditional fill colors to mutable enum cells."""
    from openpyxl.styles import PatternFill

    cells_and_values = [
        (ws.cell(row_idx, _COL_SUGGESTION_STATUS), suggestion_status),
        (ws.cell(row_idx, _COL_ACTION_TAKEN),      action_taken),
    ]
    for cell, value in cells_and_values:
        color = _STATUS_FILL.get(value)
        if color:
            cell.fill = _make_fill(color)


def _set_hyperlink(cell, url: str, display: str) -> None:
    """Set cell value as a clickable hyperlink."""
    from openpyxl.styles import Font
    cell.value = display
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


def _build_row_values(job: dict, run_date_str: str) -> list:
    """Return a list of 9 values in _HEADERS column order."""
    verdict = job.get("eval", {}).get("verdict", "consider")
    suggestion_status = "Apply" if verdict == "apply" else "Consider"
    score = job.get("eval", {}).get("score")
    match_score = score / 100.0 if score is not None else None  # store as decimal for % format

    return [
        run_date_str,           # date_suggested
        job.get("title", ""),   # role_title (will be overwritten with hyperlink)
        job.get("company", ""),
        match_score,
        suggestion_status,
        "Not Applied",          # action_taken — default
        None,                   # date_applied
        None,                   # application_status
        None,                   # notes
    ]


def _write_header(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(fill_type="solid", fgColor="2F4F8F")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.auto_filter.ref = ws.dimensions


def _set_column_widths(ws) -> None:
    widths = {
        "A": 14,   # date_suggested
        "B": 40,   # role_title
        "C": 22,   # company
        "D": 12,   # match_score
        "E": 18,   # suggestion_status
        "F": 15,   # action_taken
        "G": 14,   # date_applied
        "H": 20,   # application_status
        "I": 30,   # notes
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _existing_keys(ws) -> dict[str, int]:
    """Return {norm_key: row_index} for all data rows (skips header row 1)."""
    keys: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        title   = ws.cell(row_idx, _COL_ROLE_TITLE).value or ""
        company = ws.cell(row_idx, _COL_COMPANY).value or ""
        # hyperlink cells store the display text as .value
        keys[_norm_key(str(title), str(company))] = row_idx
    return keys


def _append_job_row(ws, job: dict, run_date_str: str) -> None:
    from openpyxl.styles import Alignment

    row_values = _build_row_values(job, run_date_str)
    next_row = ws.max_row + 1

    for col_idx, value in enumerate(row_values, start=1):
        cell = ws.cell(next_row, col_idx, value)
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == _COL_NOTES))

    # Overwrite role_title cell with hyperlink
    url = job.get("url") or job.get("apply_url", "")
    if url:
        _set_hyperlink(ws.cell(next_row, _COL_ROLE_TITLE), url, job.get("title", ""))

    # Format match_score as percentage
    ws.cell(next_row, _COL_MATCH_SCORE).number_format = "0%"

    # Conditional fill colors
    suggestion_status = row_values[_COL_SUGGESTION_STATUS - 1]
    action_taken      = row_values[_COL_ACTION_TAKEN - 1]
    _apply_row_styles(ws, next_row, suggestion_status, action_taken)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_to_tracker(jobs: list[dict], run_date: str | None = None) -> None:
    """Append new Apply/Consider rows to the Applications xlsx tracker.

    Skips roles already present (dedup by normalized title+company key).
    Creates the spreadsheet with header + formatting if it doesn't exist yet.

    Args:
        jobs:     Ranked job dicts with 'eval' populated (output of rank_jobs()).
        run_date: ISO date string for date_suggested (defaults to today).
    """
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping tracker write. Run: pip install openpyxl")
        return

    run_date_str = run_date or datetime.date.today().isoformat()

    # Filter to only Apply + Consider
    eligible = [j for j in jobs if j.get("eval", {}).get("verdict") in ("apply", "consider")]
    if not eligible:
        log.info("Tracker: no Apply/Consider roles to write.")
        return

    TRACKER_PATH.parent.mkdir(mode=0o700, parents=True, exist_ok=True)

    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        _write_header(ws)
        _set_column_widths(ws)
        ws.freeze_panes = "A2"
        log.info("Tracker: created new spreadsheet at %s", TRACKER_PATH)

    existing = _existing_keys(ws)
    added = 0
    skipped = 0

    for job in eligible:
        title   = job.get("title", "")
        company = job.get("company", "")
        if not _is_plausible_job_row(title, company):
            log.warning(
                "Tracker: skipping likely-misparsed row — title=%r  company=%r  "
                "(title lacks job-role keywords or company looks like a location)",
                title, company,
            )
            skipped += 1
            continue
        key = _norm_key(title, company)
        if key in existing:
            skipped += 1
            continue
        _append_job_row(ws, job, run_date_str)
        existing[key] = ws.max_row   # prevent duplicates within the same run
        added += 1

    if added:
        _sort_rows_by_date(ws)
        wb.save(TRACKER_PATH)
        TRACKER_PATH.chmod(0o600)
        log.info("Tracker: wrote %d new row(s) to %s (%d duplicate(s) skipped)",
                 added, TRACKER_PATH, skipped)
    else:
        log.info("Tracker: all %d role(s) already present — nothing written.", skipped)
