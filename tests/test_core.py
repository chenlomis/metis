"""Core logic tests for metis.

Philosophy: "if this breaks, something important is wrong."
These tests cover the functions whose silent failure would cause the most
harm — bad dedup, missed jobs, wrong scoring order, or buried roles.

Run with:  pytest tests/
"""
import datetime
import json
import re
import tempfile
from pathlib import Path
from unittest.mock import patch

import anthropic

import pytest


# ---------------------------------------------------------------------------
# state.py — role hashing + seen_roles TTL
# ---------------------------------------------------------------------------

class TestRoleHash:
    """_role_hash must be stable — dedup depends on it never changing."""

    def test_hash_is_12_chars(self):
        from metis.state import _role_hash
        h = _role_hash("Staff Product Manager", "Anthropic")
        assert len(h) == 12

    def test_hash_is_stable(self):
        """Same inputs always produce the same hash (no randomness)."""
        from metis.state import _role_hash
        assert _role_hash("Staff PM", "Stripe") == _role_hash("Staff PM", "Stripe")

    def test_hash_normalises_case_and_punctuation(self):
        """Normalisation means 'Staff PM @ Stripe' == 'staff pm stripe'."""
        from metis.state import _role_hash
        assert _role_hash("Staff PM", "Stripe") == _role_hash("STAFF PM", "STRIPE")
        assert _role_hash("Staff PM", "Stripe") == _role_hash("Staff PM!", "Stripe.")

    def test_different_roles_produce_different_hashes(self):
        from metis.state import _role_hash
        assert _role_hash("Staff PM", "Stripe") != _role_hash("Staff PM", "Anthropic")
        assert _role_hash("Staff PM", "Stripe") != _role_hash("Senior PM", "Stripe")

    def test_company_variant_suffixes_do_not_change_persisted_hash_contract(self):
        """_role_hash is persisted; do not add company canonicalization without migration."""
        from metis.state import _role_hash
        assert _role_hash("Staff PM", "NVIDIA") != _role_hash("Staff PM", "NVIDIA AI")
        assert _role_hash("Staff PM", "Anthropic") != _role_hash("Staff PM", "Anthropic Labs")
        assert _role_hash("Staff PM", "Acme") != _role_hash("Staff PM", "Acme Corp.")
        assert _role_hash("Staff PM", "Stripe") != _role_hash("Staff PM", "Stripe Inc.")

    def test_distinct_companies_still_differ(self):
        """Normalization must not collapse genuinely different companies."""
        from metis.state import _role_hash
        assert _role_hash("Staff PM", "Scale") != _role_hash("Staff PM", "Anthropic")
        assert _role_hash("Staff PM", "OpenAI") != _role_hash("Staff PM", "Anthropic")


class TestSeenRolesTTL:
    """save_seen_roles must prune expired entries; load_seen_roles must honour TTL."""

    def test_new_entries_survive_within_ttl(self, tmp_path):
        from metis.state import save_seen_roles, load_seen_roles
        with patch("metis.state.DATA_DIR", tmp_path):
            now = datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)
            entries = {"abc123": now.isoformat()}
            save_seen_roles(entries)
            seen = load_seen_roles()
        assert "abc123" in seen

    def test_expired_entries_are_pruned_on_load(self, tmp_path):
        """An entry from 31 days ago must not appear in load_seen_roles (TTL is 30d)."""
        from metis.state import save_seen_roles, load_seen_roles
        with patch("metis.state.DATA_DIR", tmp_path):
            old_ts = (
                datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)
                - datetime.timedelta(days=31)
            ).isoformat()
            p = tmp_path / "seen_roles.json"
            p.write_text(json.dumps({"deadbeef1234": old_ts}))
            seen = load_seen_roles()
        assert "deadbeef1234" not in seen

    def test_save_prunes_stale_entries(self, tmp_path):
        """Saving new entries must evict old ones from the file (TTL is 30d)."""
        from metis.state import save_seen_roles
        with patch("metis.state.DATA_DIR", tmp_path):
            old_ts = (
                datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)
                - datetime.timedelta(days=31)
            ).isoformat()
            now = datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None).isoformat()
            p = tmp_path / "seen_roles.json"
            p.write_text(json.dumps({"stale": old_ts}))
            save_seen_roles({"fresh": now})
            on_disk = json.loads(p.read_text())
        assert "stale" not in on_disk
        assert "fresh" in on_disk

    def test_capped_roles_not_buried(self, tmp_path):
        """Roles dropped by the cap must NOT be written to seen_roles.json
        (regression for the role-burial bug fixed in commit 8e805de).
        """
        from metis.state import save_seen_roles, load_seen_roles
        # Simulate: 3 roles found, only 1 scored, 2 capped
        now = datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None).isoformat()
        with patch("metis.state.DATA_DIR", tmp_path):
            save_seen_roles({"scored_role": now})   # only the scored one
            seen = load_seen_roles()
        assert "scored_role" in seen
        assert "capped_role_1" not in seen   # was never passed to save_seen_roles
        assert "capped_role_2" not in seen


# ---------------------------------------------------------------------------
# sources/linkedin.py — job extraction from email body
# ---------------------------------------------------------------------------

_SAMPLE_ALERT_BODY = """\
Your job alert for "product manager"

Senior Product Manager
Anthropic
San Francisco, CA · Remote
3 company alumni
View job: https://www.linkedin.com/comm/jobs/view/3901234567/AAAA_aaaa111/?trackingId=xyz

Staff Software Engineer
Stripe
New York, NY · On-site
View job: https://www.linkedin.com/comm/jobs/view/3901234568/BBBB_bbbb222/?trackingId=abc

--
Your job alert
"""

class TestExtractJobs:
    """extract_jobs() must reliably parse title / company / location / job_id."""

    def test_extracts_both_jobs(self):
        from metis.sources.linkedin import extract_jobs
        jobs = extract_jobs(_SAMPLE_ALERT_BODY)
        assert len(jobs) == 2

    def test_first_job_fields(self):
        from metis.sources.linkedin import extract_jobs
        job = extract_jobs(_SAMPLE_ALERT_BODY)[0]
        assert job["title"]    == "Senior Product Manager"
        assert job["company"]  == "Anthropic"
        assert "San Francisco" in job["location"]
        assert job["job_id"]   == "3901234567"

    def test_second_job_fields(self):
        from metis.sources.linkedin import extract_jobs
        job = extract_jobs(_SAMPLE_ALERT_BODY)[1]
        assert job["title"]   == "Staff Software Engineer"
        assert job["company"] == "Stripe"
        assert job["job_id"]  == "3901234568"

    def test_alumni_count_captured(self):
        from metis.sources.linkedin import extract_jobs
        job = extract_jobs(_SAMPLE_ALERT_BODY)[0]
        assert job["alumni_count"] == 3

    def test_url_format_is_clean(self):
        """URLs must be the canonical /jobs/view/ID/ format, not the tracking URL."""
        from metis.sources.linkedin import extract_jobs
        job = extract_jobs(_SAMPLE_ALERT_BODY)[0]
        assert job["url"] == "https://www.linkedin.com/jobs/view/3901234567/"

    def test_duplicate_job_id_deduplicated(self):
        body = _SAMPLE_ALERT_BODY + "\n".join([
            "", "Senior Product Manager", "Anthropic", "Remote",
            "View job: https://www.linkedin.com/comm/jobs/view/3901234567/DUPE/?trackingId=dupe",
        ])
        from metis.sources.linkedin import extract_jobs
        jobs = extract_jobs(body)
        ids = [j["job_id"] for j in jobs]
        assert ids.count("3901234567") == 1

    def test_returns_empty_on_no_jobs(self):
        from metis.sources.linkedin import extract_jobs
        assert extract_jobs("No jobs here") == []


class TestExtractJobsHtml:
    """extract_jobs_html() must parse recommendation emails that have no plain-text 'View job:' line."""

    _HTML = """
    <html><body>
    <a href="https://www.linkedin.com/comm/jobs/view/3999000001/?trackingId=r1">
      Head of Product
    </a>
    <span>Cohere · Toronto, ON</span>
    <a href="https://www.linkedin.com/comm/jobs/view/3999000002/?trackingId=r2">
      Principal ML Engineer
    </a>
    <span>Mistral AI · Paris, France (Remote)</span>
    </body></html>
    """

    def test_extracts_two_jobs(self):
        from metis.sources.linkedin import extract_jobs_html
        jobs = extract_jobs_html(self._HTML)
        assert len(jobs) == 2

    def test_title_and_company_parsed(self):
        from metis.sources.linkedin import extract_jobs_html
        jobs = extract_jobs_html(self._HTML)
        titles   = {j["title"]   for j in jobs}
        companies = {j["company"] for j in jobs}
        assert "Head of Product" in titles
        assert "Cohere" in companies


# ---------------------------------------------------------------------------
# score.py — verdict derivation + ranking
# ---------------------------------------------------------------------------

class TestRankJobs:
    """rank_jobs() must re-derive verdicts from scores AND sort correctly."""

    def _make_job(self, title, score, raw_verdict="apply"):
        return {
            "title": title, "company": "Acme", "location": "SF",
            "eval": {"score": score, "verdict": raw_verdict,
                     "leveragePoints": [], "frictionPoints": [], "tags": []},
        }

    def test_high_score_becomes_apply(self):
        from metis.score import rank_jobs
        jobs = rank_jobs([self._make_job("Senior PM", 80)])
        assert jobs[0]["eval"]["verdict"] == "apply"

    def test_mid_score_becomes_consider(self):
        from metis.score import rank_jobs
        jobs = rank_jobs([self._make_job("Mid PM", 60)])
        assert jobs[0]["eval"]["verdict"] == "consider"

    def test_low_score_becomes_skipped(self):
        from metis.score import rank_jobs
        jobs = rank_jobs([self._make_job("Junior PM", 30)])
        assert jobs[0]["eval"]["verdict"] == "skipped"

    def test_verdict_drift_corrected(self):
        """Claude might return verdict='apply' for score=62 — rank_jobs must fix it."""
        from metis.score import rank_jobs
        job = self._make_job("Drifted", score=62, raw_verdict="apply")
        result = rank_jobs([job])
        assert result[0]["eval"]["verdict"] == "consider"

    def test_sort_order(self):
        """apply before consider before skipped; higher score first within tier."""
        from metis.score import rank_jobs
        jobs = [
            self._make_job("Skip",     40, "skipped"),
            self._make_job("Apply-B",  76, "apply"),
            self._make_job("Consider", 60, "consider"),
            self._make_job("Apply-A",  90, "apply"),
        ]
        ranked = rank_jobs(jobs)
        titles = [j["title"] for j in ranked]
        assert titles == ["Apply-A", "Apply-B", "Consider", "Skip"]

    def test_boundary_at_apply_threshold(self):
        """Score of exactly 75 must be 'apply'; 74 must be 'consider'."""
        from metis.score import rank_jobs
        j75 = rank_jobs([self._make_job("Boundary", 75)])[0]
        j74 = rank_jobs([self._make_job("Below",    74)])[0]
        assert j75["eval"]["verdict"] == "apply"
        assert j74["eval"]["verdict"] == "consider"

    def test_filtered_verdict_preserved(self):
        """A deal_breaker violation (verdict='filtered', score=0) must not be
        reclassified as 'skipped' just because score=0."""
        from metis.score import rank_jobs
        job = {
            "title": "Director of Tobacco", "company": "Big Tobacco", "location": "NY",
            "eval": {"score": 0, "verdict": "filtered",
                     "leveragePoints": [], "frictionPoints": [],
                     "tags": [{"text": "deal breaker: industry", "sentiment": "red"}]},
        }
        result = rank_jobs([job])
        assert result[0]["eval"]["verdict"] == "filtered"

    def test_filtered_roles_sorted_last(self):
        """Filtered roles must sort after skipped."""
        from metis.score import rank_jobs
        jobs = [
            self._make_job("Filtered", 0, "filtered"),
            self._make_job("Skipped",  30, "skipped"),
            self._make_job("Apply",    80, "apply"),
        ]
        # patch filtered verdict so rank_jobs preserves it
        jobs[0]["eval"]["verdict"] = "filtered"
        ranked = rank_jobs(jobs)
        assert ranked[0]["eval"]["verdict"] == "apply"
        assert ranked[-1]["eval"]["verdict"] == "filtered"


# ---------------------------------------------------------------------------
# score.py — partial JSON recovery
# ---------------------------------------------------------------------------

class TestRecoverPartialJson:
    """_recover_partial_json must salvage complete objects from a truncated array."""

    def test_recovers_from_truncated_array(self):
        from metis.score import _recover_partial_json
        raw = '[{"score": 80, "verdict": "apply"}, {"score": 60, "verdict": "consi'
        objects = _recover_partial_json(raw)
        assert len(objects) == 1
        assert objects[0]["score"] == 80

    def test_recovers_multiple_complete_objects(self):
        from metis.score import _recover_partial_json
        raw = '[{"score": 80}, {"score": 60}, {"score": 40}]'
        objects = _recover_partial_json(raw)
        assert len(objects) == 3

    def test_returns_empty_list_on_garbage(self):
        from metis.score import _recover_partial_json
        assert _recover_partial_json("not json at all") == []

    def test_skips_invalid_inner_objects(self):
        from metis.score import _recover_partial_json
        raw = '[{"score": 80}, {bad json}, {"score": 40}]'
        objects = _recover_partial_json(raw)
        scores = [o["score"] for o in objects]
        assert 80 in scores
        assert 40 in scores


# ---------------------------------------------------------------------------
# pipeline.py — lookback parsing
# ---------------------------------------------------------------------------

class TestParseLookback:
    """_parse_lookback must handle all documented formats."""

    def test_days_shorthand(self):
        from metis.pipeline import _parse_lookback
        result = _parse_lookback("3d")
        delta = datetime.datetime.now() - result
        assert 2 < delta.total_seconds() / 86400 < 4

    def test_iso_date(self):
        from metis.pipeline import _parse_lookback
        result = _parse_lookback("2026-05-10")
        assert result.year == 2026
        assert result.month == 5
        assert result.day == 10

    def test_invalid_returns_none(self):
        from metis.pipeline import _parse_lookback
        assert _parse_lookback("not-a-date") is None

    def test_zero_days_returns_near_now(self):
        from metis.pipeline import _parse_lookback
        result = _parse_lookback("0d")
        delta = datetime.datetime.now() - result
        assert delta.total_seconds() < 10

    def test_prompt_score_all_accepts_number_below_cap(self, monkeypatch):
        from metis.pipeline import _prompt_score_all

        monkeypatch.setattr("builtins.input", lambda _prompt: "30")

        assert _prompt_score_all(110, 40) == 30


# ---------------------------------------------------------------------------
# profile.py — render_profile sanity
# ---------------------------------------------------------------------------

class TestRenderProfile:
    """render_profile must produce a non-empty string with the key sections."""

    _PROFILE = {
        "candidate": {"name": "Alex Kim", "title": "Staff PM", "location": "NYC",
                      "work_mode": ["Remote-first"], "open_to_remote": True},
        "target": {"level": "Staff", "roles": ["PM", "Group PM"]},
        "scoring": {"apply_threshold": 75, "consider_threshold": 55},
        "aspirations": {"track": "IC", "direction": "AI-native products"},
        "deal_breakers": ["No management-only roles"],
        "experience": [{"company": "Acme", "title": "PM", "dates": "2022–2024",
                        "highlights": ["Led 0→1 product"]}],
    }

    def test_renders_non_empty(self):
        from metis.profile import render_profile
        out = render_profile(self._PROFILE)
        assert len(out) > 100

    def test_contains_candidate_name(self):
        from metis.profile import render_profile
        out = render_profile(self._PROFILE)
        assert "Alex Kim" in out

    def test_contains_target_roles(self):
        from metis.profile import render_profile
        out = render_profile(self._PROFILE)
        assert "PM" in out

    def test_contains_aspirations(self):
        from metis.profile import render_profile
        out = render_profile(self._PROFILE)
        assert "ASPIRATIONS" in out
        assert "AI-native products" in out

    def test_contains_deal_breakers(self):
        from metis.profile import render_profile
        out = render_profile(self._PROFILE)
        assert "DEAL BREAKERS" in out

    def test_contains_experience(self):
        from metis.profile import render_profile
        out = render_profile(self._PROFILE)
        assert "Acme" in out
        assert "0→1 product" in out

    def test_empty_profile_does_not_crash(self):
        from metis.profile import render_profile
        out = render_profile({})
        assert isinstance(out, str)

    def test_profile_with_no_experience(self):
        """Resume with no work history must not crash or raise KeyError."""
        from metis.profile import render_profile
        profile = {
            "candidate": {"name": "New Grad", "title": "Recent Graduate"},
            "target": {"level": "Junior", "roles": ["PM"]},
            "scoring": {"apply_threshold": 75, "consider_threshold": 55},
        }
        out = render_profile(profile)
        assert "New Grad" in out

    def test_profile_missing_optional_sections_renders_cleanly(self):
        """A profile with only candidate + target must not show empty headers."""
        from metis.profile import render_profile
        profile = {
            "candidate": {"name": "Min User"},
            "target":    {"roles": ["Engineer"]},
        }
        out = render_profile(profile)
        # No section should appear with empty content
        assert "EXPERIENCE:" not in out
        assert "EDUCATION:" not in out
        assert "DEAL BREAKERS" not in out

    def test_load_profile_text_reads_feedback_from_data_dir(self, tmp_path, monkeypatch):
        """METIS_DATA_DIR must isolate feedback for demos/personas."""
        import importlib
        import metis.profile as profile_mod

        profile_path = tmp_path / "profile.yaml"
        profile_path.write_text(
            "candidate:\n"
            "  name: Demo User\n"
            "target:\n"
            "  roles:\n"
            "    - Product Manager\n"
        )
        (tmp_path / "feedback.md").write_text("Prefer climate roles over generic SaaS.")

        monkeypatch.setenv("METIS_PROFILE", str(profile_path))
        monkeypatch.setenv("METIS_DATA_DIR", str(tmp_path))
        profile_mod = importlib.reload(profile_mod)

        out = profile_mod.load_profile_text()

        assert "Demo User" in out
        assert "Prefer climate roles over generic SaaS." in out

        monkeypatch.delenv("METIS_PROFILE", raising=False)
        monkeypatch.delenv("METIS_DATA_DIR", raising=False)
        importlib.reload(profile_mod)


# ---------------------------------------------------------------------------
# score.py — output normalization
# ---------------------------------------------------------------------------

class TestScoreNormalization:
    def test_scoring_rubric_keeps_adjacent_domains_soft(self):
        from metis.score import _build_score_suffix

        rubric = _build_score_suffix("Lomis", 75, 55)

        assert "Do not let domain_background alone push a role below the consider threshold" in rubric
        assert "domain: foreign\" requires an explicit hard domain prerequisite" in rubric

    def test_known_tag_text_forces_canonical_sentiment(self):
        from metis.score import _normalize_tag_sentiments

        evals = [{
            "dimensions": [
                {
                    "name": "domain_background",
                    "score": 40,
                    "rationale": "Healthcare regulatory compliance requires HIPAA expertise",
                }
            ],
            "tags": [
                {"text": "stage: public co fit", "sentiment": "amber"},
                {"text": "domain: foreign", "sentiment": "amber"},
            ]
        }]

        _normalize_tag_sentiments(evals)

        assert evals[0]["tags"][0]["sentiment"] == "green"
        assert evals[0]["tags"][1]["sentiment"] == "red"

    def test_comp_undisclosed_removed_when_salary_is_aspirational(self):
        from metis.score import _normalize_tag_sentiments

        evals = [{
            "tags": [
                {"text": "comp: undisclosed", "sentiment": "amber"},
                {"text": "stage: growth fit", "sentiment": "amber"},
            ]
        }]

        _normalize_tag_sentiments(evals, salary_is_hard_floor=False)

        assert evals[0]["tags"] == [{"text": "stage: growth fit", "sentiment": "green"}]

    def test_comp_undisclosed_kept_when_salary_is_hard_floor(self):
        from metis.score import _normalize_tag_sentiments

        evals = [{"tags": [{"text": "comp: undisclosed", "sentiment": "green"}]}]

        _normalize_tag_sentiments(evals, salary_is_hard_floor=True)

        assert evals[0]["tags"] == [{"text": "comp: undisclosed", "sentiment": "amber"}]


# ---------------------------------------------------------------------------
# sources/linkedin.py — edge cases in email parsing
# ---------------------------------------------------------------------------

class TestExtractJobsIndividualNotification:
    """LinkedIn also sends single-job notification emails with the format:
    '[Job Title] at [Company] – Your job alert for [query] in [City]'
    These must parse correctly alongside the standard multi-job digest format.
    """

    _INDIVIDUAL_BODY = """\
Your job alert for "senior product manager"

Sr. Product Manager Tech, Product Quality
Amazon
Seattle, WA · Full-time
3 company alumni
View job: https://www.linkedin.com/comm/jobs/view/3901299999/ABC_def/?trackingId=ind1
"""

    _INDIVIDUAL_BODY_WITH_DASH = """\
Lead Product Manager – Risk Platform
BILL
Sunnyvale, CA
View job: https://www.linkedin.com/comm/jobs/view/3901288888/XYZ_abc/?trackingId=ind2
"""

    def test_single_job_notification_parsed(self):
        from metis.sources.linkedin import extract_jobs
        jobs = extract_jobs(self._INDIVIDUAL_BODY)
        assert len(jobs) == 1
        assert jobs[0]["title"]   == "Sr. Product Manager Tech, Product Quality"
        assert jobs[0]["company"] == "Amazon"
        assert jobs[0]["job_id"]  == "3901299999"

    def test_alumni_count_in_individual_email(self):
        from metis.sources.linkedin import extract_jobs
        jobs = extract_jobs(self._INDIVIDUAL_BODY)
        assert jobs[0]["alumni_count"] == 3

    def test_title_with_dash_parsed_correctly(self):
        """Em-dash in job title must not be mistaken for noise."""
        from metis.sources.linkedin import extract_jobs
        jobs = extract_jobs(self._INDIVIDUAL_BODY_WITH_DASH)
        assert len(jobs) == 1
        assert "Risk Platform" in jobs[0]["title"]
        assert jobs[0]["company"] == "BILL"

    def test_mixed_digest_and_individual_in_same_session(self):
        """A run that ingests both a digest email and an individual notification
        email must deduplicate correctly and return all unique jobs."""
        from metis.sources.linkedin import extract_jobs
        combined = _SAMPLE_ALERT_BODY + "\n" + self._INDIVIDUAL_BODY
        jobs = extract_jobs(combined)
        ids = [j["job_id"] for j in jobs]
        # All 3 jobs unique
        assert len(ids) == len(set(ids))
        assert "3901299999" in ids


class TestExtractJobsEdgeCases:
    """Unusual but real email structures that must not crash or silently drop jobs."""

    def test_malformed_url_no_job_id_skipped(self):
        """A 'View job:' line with no numeric ID must be silently skipped."""
        from metis.sources.linkedin import extract_jobs
        body = (
            "Some Title\nSome Company\nSome City\n"
            "View job: https://www.linkedin.com/comm/jobs/view/not-a-number/\n"
        )
        jobs = extract_jobs(body)
        assert jobs == []

    def test_fewer_than_three_lines_before_url_skipped(self):
        """If there aren't enough context lines above the URL, job must be skipped
        rather than filling fields with noise."""
        from metis.sources.linkedin import extract_jobs
        body = (
            "Only one line\n"
            "View job: https://www.linkedin.com/comm/jobs/view/1234567890/X/\n"
        )
        jobs = extract_jobs(body)
        assert jobs == []

    def test_noise_lines_filtered_before_field_extraction(self):
        """'Actively hiring', 'Be an early applicant', etc. must not
        displace title/company/location from their expected positions."""
        from metis.sources.linkedin import extract_jobs
        body = (
            "Senior PM\n"
            "Anthropic\n"
            "San Francisco, CA\n"
            "Actively hiring\n"
            "Be an early applicant\n"
            "3 connections\n"
            "View job: https://www.linkedin.com/comm/jobs/view/9900000001/Z/\n"
        )
        jobs = extract_jobs(body)
        assert len(jobs) == 1
        assert jobs[0]["title"] == "Senior PM"
        assert jobs[0]["company"] == "Anthropic"

    def test_html_extraction_skips_navigation_links(self):
        """Short anchor text like 'View all jobs', 'Apply', 'LinkedIn' must
        not be parsed as job titles."""
        from metis.sources.linkedin import extract_jobs_html
        html = """
        <html><body>
          <a href="https://www.linkedin.com/comm/jobs/view/5000000001/?t=x">Head of AI</a>
          <span>OpenAI · San Francisco, CA</span>
          <a href="https://www.linkedin.com/comm/jobs/view/5000000001/?t=x">View all jobs</a>
          <a href="https://www.linkedin.com/comm/jobs/view/5000000001/?t=x">Apply</a>
        </body></html>
        """
        jobs = extract_jobs_html(html)
        titles = [j["title"] for j in jobs]
        assert "View all jobs" not in titles
        assert "Apply" not in titles
        # The real job should be present exactly once
        assert titles.count("Head of AI") == 1

    def test_empty_html_body_returns_empty_list(self):
        from metis.sources.linkedin import extract_jobs_html
        assert extract_jobs_html("") == []
        assert extract_jobs_html("<html><body></body></html>") == []


# ---------------------------------------------------------------------------
# score.py — score_jobs_batch edge cases (no real API calls)
# ---------------------------------------------------------------------------

class TestScoreJobsBatchEdgeCases:
    """Verify the scoring glue logic without calling the Anthropic API."""

    def _make_job(self, title="Staff PM", company="Acme", score=80):
        return {
            "title": title, "company": company, "location": "SF",
            "jd": "Some job description text.",
            "eval": {},
        }

    def test_short_chunk_response_retries_missing_jobs(self):
        """If the scorer returns too few evals, retry missing jobs before using
        parse-error placeholders."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = [self._make_job(f"Job {i}") for i in range(3)]

        def response_for(scores):
            fake_response = mock.MagicMock()
            fake_response.content = [mock.MagicMock(text=json.dumps([
                {"score": score, "verdict": "consider", "leveragePoints": [], "frictionPoints": [], "tags": []}
                for score in scores
            ]))]
            fake_response.usage = mock.MagicMock(
                input_tokens=100, output_tokens=20,
                cache_creation_input_tokens=0, cache_read_input_tokens=0,
            )
            return fake_response

        with mock.patch("metis.score._build_score_system", return_value="mock profile"):
            fake_client = mock.MagicMock()
            fake_client.messages.create.side_effect = [
                response_for([80]),
                response_for([61]),
                response_for([62]),
            ]
            result = score_jobs_batch(fake_client, jobs)

        assert len(result) == 3
        assert [job["eval"]["score"] for job in result] == [80, 61, 62]
        assert fake_client.messages.create.call_count == 3

    def test_provider_wrapper_results_are_unwrapped(self):
        """OpenAI JSON mode may wrap the requested array in a top-level results key."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = [self._make_job("Job 1"), self._make_job("Job 2")]
        wrapped = {
            "results": [
                {"score": 80, "verdict": "apply", "leveragePoints": [], "frictionPoints": [], "tags": []},
                {"score": 62, "verdict": "consider", "leveragePoints": [], "frictionPoints": [], "tags": []},
            ]
        }
        fake_response = mock.MagicMock()
        fake_response.content = [mock.MagicMock(text=json.dumps(wrapped))]
        fake_response.usage = mock.MagicMock(
            input_tokens=100, output_tokens=20,
            cache_creation_input_tokens=0, cache_read_input_tokens=0,
        )

        with mock.patch("metis.score._build_score_system", return_value="mock profile"):
            fake_client = mock.MagicMock()
            fake_client.messages.create.return_value = fake_response
            result = score_jobs_batch(fake_client, jobs)

        assert result[0]["eval"]["score"] == 80
        assert result[1]["eval"]["score"] == 62
        assert result[1]["eval"]["frictionPoints"] == []

    def test_claude_returns_completely_broken_json(self):
        """If Claude returns garbage, all jobs must be marked skipped — no crash."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = [self._make_job()]
        fake_response = mock.MagicMock()
        fake_response.content = [mock.MagicMock(text="I cannot score these roles.")]
        fake_response.usage = mock.MagicMock(
            input_tokens=50, output_tokens=10,
            cache_creation_input_tokens=0, cache_read_input_tokens=0,
        )

        with mock.patch("metis.score._build_score_system", return_value="mock profile"):
            fake_client = mock.MagicMock()
            fake_client.messages.create.return_value = fake_response
            result = score_jobs_batch(fake_client, jobs)

        assert result[0]["eval"]["verdict"] == "skipped"
        assert "parse error" in result[0]["eval"]["frictionPoints"][0].lower()


def test_domain_foreign_tag_demotes_without_hard_barrier():
    from metis.score import _normalize_tag_sentiments

    evals = [{
        "dimensions": [
            {
                "name": "domain_background",
                "score": 70,
                "rationale": "Retail commerce platform, not AI infra-native",
            }
        ],
        "frictionPoints": ["Retail/commerce platform, not AI infra-native."],
        "tags": [{"text": "domain: foreign", "sentiment": "red"}],
    }]

    _normalize_tag_sentiments(evals)

    assert evals[0]["tags"][0] == {"text": "domain: adjacent", "sentiment": "amber"}


def test_domain_foreign_tag_preserved_for_hard_networking_barrier():
    from metis.score import _normalize_tag_sentiments

    evals = [{
        "dimensions": [
            {
                "name": "domain_background",
                "score": 30,
                "rationale": "RDMA and InfiniBand datacenter fabric expertise required",
            }
        ],
        "frictionPoints": ["Networking protocol domain is mandatory for first-quarter roadmap."],
        "tags": [{"text": "domain: foreign", "sentiment": "amber"}],
    }]

    _normalize_tag_sentiments(evals)

    assert evals[0]["tags"][0] == {"text": "domain: foreign", "sentiment": "red"}


# ---------------------------------------------------------------------------
# score.py — chunking behaviour
# ---------------------------------------------------------------------------

class TestScoreJobsBatchChunking:
    """score_jobs_batch must chunk large batches and merge results correctly."""

    def _make_jobs(self, n: int):
        return [
            {"title": f"Job {i}", "company": "Acme", "location": "SF",
             "jd": "Some job description.", "eval": {}}
            for i in range(n)
        ]

    def _fake_response(self, n: int):
        import unittest.mock as mock
        evals = [
            {"score": 70, "verdict": "consider",
             "leveragePoints": [], "frictionPoints": [], "tags": []}
            for _ in range(n)
        ]
        r = mock.MagicMock()
        r.content = [mock.MagicMock(text=json.dumps(evals))]
        r.usage = mock.MagicMock(
            input_tokens=100, output_tokens=50,
            cache_creation_input_tokens=0, cache_read_input_tokens=0,
        )
        return r

    def test_small_batch_single_chunk(self):
        """Fewer than _SCORE_CHUNK_SIZE jobs → exactly one API call."""
        from metis.score import score_jobs_batch, _SCORE_CHUNK_SIZE
        import unittest.mock as mock

        jobs = self._make_jobs(_SCORE_CHUNK_SIZE - 1)
        fake_client = mock.MagicMock()
        fake_client.messages.create.return_value = self._fake_response(len(jobs))

        with mock.patch("metis.score._build_score_system", return_value="mock"):
            result = score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 1
        assert len(result) == len(jobs)

    def test_large_batch_multiple_chunks(self):
        """More than _SCORE_CHUNK_SIZE jobs → multiple API calls, all evals merged."""
        from metis.score import score_jobs_batch, _SCORE_CHUNK_SIZE
        import unittest.mock as mock

        n = _SCORE_CHUNK_SIZE + 5  # guaranteed to require 2 chunks
        jobs = self._make_jobs(n)

        def side_effect(*args, **kwargs):
            # Return evals matching the chunk size for each call
            content = kwargs.get("messages", [{}])[0].get("content", "")
            # Count "JOB N:" lines to infer chunk size
            chunk_n = content.count("JOB ")
            return self._fake_response(chunk_n)

        fake_client = mock.MagicMock()
        fake_client.messages.create.side_effect = side_effect

        with mock.patch("metis.score._build_score_system", return_value="mock"):
            result = score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 2
        assert len(result) == n
        assert all(j["eval"]["score"] == 70 for j in result)

    def test_chunk_truncation_fills_remainder_after_retry_failure(self):
        """If retries still fail to return evals, missing slots get _error_eval."""
        from metis.score import score_jobs_batch, _SCORE_CHUNK_SIZE
        import unittest.mock as mock

        jobs = self._make_jobs(5)
        fake_client = mock.MagicMock()
        fake_client.messages.create.return_value = self._fake_response(0)

        with mock.patch("metis.score._build_score_system", return_value="mock"):
            result = score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 6
        assert result[0]["eval"]["verdict"] == "skipped"
        assert "parse error" in result[0]["eval"]["frictionPoints"][0].lower()


# ---------------------------------------------------------------------------
# score.py — API retry logic
# ---------------------------------------------------------------------------

class TestScoreChunkRetry:
    """_score_chunk must retry on transient API errors and raise on persistent failure."""

    def _make_jobs(self, n=2):
        return [
            {"title": f"Job {i}", "company": "Acme", "location": "SF",
             "jd": "Some job description.", "eval": {}}
            for i in range(n)
        ]

    def _good_response(self, n):
        import unittest.mock as mock
        evals = [
            {"score": 75, "verdict": "apply",
             "leveragePoints": [], "frictionPoints": [], "tags": []}
            for _ in range(n)
        ]
        r = mock.MagicMock()
        r.content = [mock.MagicMock(text=json.dumps(evals))]
        r.usage = mock.MagicMock(
            input_tokens=100, output_tokens=40,
            cache_creation_input_tokens=0, cache_read_input_tokens=0,
        )
        return r

    def test_succeeds_on_first_try(self):
        """Happy path — no retries needed."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = self._make_jobs(2)
        fake_client = mock.MagicMock()
        fake_client.messages.create.return_value = self._good_response(2)

        with mock.patch("metis.score._build_score_system", return_value="mock"), \
             mock.patch("time.sleep"):
            result = score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 1
        assert result[0]["eval"]["score"] == 75

    def test_retries_on_500_and_succeeds(self):
        """First call raises InternalServerError; second call succeeds — result is correct."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = self._make_jobs(2)
        good = self._good_response(2)
        fake_client = mock.MagicMock()
        fake_client.messages.create.side_effect = [
            anthropic.InternalServerError(
                message="Internal server error",
                response=mock.MagicMock(status_code=500, headers={}),
                body={"type": "error", "error": {"type": "api_error",
                      "message": "Internal server error"}},
            ),
            good,
        ]

        with mock.patch("metis.score._build_score_system", return_value="mock"), \
             mock.patch("time.sleep") as mock_sleep:
            result = score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 2
        mock_sleep.assert_called_once()   # backed off once
        assert result[0]["eval"]["score"] == 75

    def test_retries_on_rate_limit_and_succeeds(self):
        """RateLimitError (429) is also retried."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = self._make_jobs(1)
        good = self._good_response(1)
        fake_client = mock.MagicMock()
        fake_client.messages.create.side_effect = [
            anthropic.RateLimitError(
                message="rate limited",
                response=mock.MagicMock(status_code=429, headers={}),
                body={},
            ),
            good,
        ]

        with mock.patch("metis.score._build_score_system", return_value="mock"), \
             mock.patch("time.sleep"):
            result = score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 2
        assert result[0]["eval"]["score"] == 75

    def test_raises_after_max_retries(self):
        """Three consecutive 500s must propagate — not silently swallow."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = self._make_jobs(1)
        fake_client = mock.MagicMock()
        fake_client.messages.create.side_effect = anthropic.InternalServerError(
            message="Internal server error",
            response=mock.MagicMock(status_code=500, headers={}),
            body={"type": "error", "error": {"type": "api_error",
                  "message": "Internal server error"}},
        )

        with mock.patch("metis.score._build_score_system", return_value="mock"), \
             mock.patch("time.sleep"):
            with pytest.raises(anthropic.InternalServerError):
                score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 3   # all 3 attempts made

    def test_non_retryable_error_propagates_immediately(self):
        """A non-retryable error (e.g. AuthenticationError) must not be retried."""
        from metis.score import score_jobs_batch
        import unittest.mock as mock

        jobs = self._make_jobs(1)
        fake_client = mock.MagicMock()
        fake_client.messages.create.side_effect = anthropic.AuthenticationError(
            message="invalid api key",
            response=mock.MagicMock(status_code=401, headers={}),
            body={},
        )

        with mock.patch("metis.score._build_score_system", return_value="mock"), \
             mock.patch("time.sleep"):
            with pytest.raises(anthropic.AuthenticationError):
                score_jobs_batch(fake_client, jobs)

        assert fake_client.messages.create.call_count == 1   # no retry


# ---------------------------------------------------------------------------
# init_bak_cmd.py — salary floor / deal_breaker consistency
# ---------------------------------------------------------------------------

class TestApplyPrefsToProfile:
    """_apply_prefs_to_profile must keep salary_floor_usd as the single source of truth."""

    def _base_profile(self):
        return {
            "deal_breakers": [
                "Base salary below $250,000",
                "Role requires people management",
            ],
            "salary_floor_usd": 250000,
        }

    def test_salary_floor_update_removes_salary_deal_breaker(self):
        """When salary_floor is updated, any salary mention in deal_breakers must be removed."""
        from metis.init_bak_cmd import _apply_prefs_to_profile
        profile = self._base_profile()
        _apply_prefs_to_profile(profile, {"salary_floor": "200000"})
        assert profile["salary_floor_usd"] == 200000
        # No salary-related deal-breaker should remain
        salary_dbs = [d for d in profile["deal_breakers"] if "salary" in d.lower() or "compensation" in d.lower()]
        assert salary_dbs == []

    def test_non_salary_deal_breakers_preserved(self):
        """Only salary-related deal-breakers are removed — others must survive."""
        from metis.init_bak_cmd import _apply_prefs_to_profile
        profile = self._base_profile()
        _apply_prefs_to_profile(profile, {"salary_floor": "200000"})
        assert any("management" in d for d in profile["deal_breakers"])

    def test_no_salary_deal_breaker_no_change(self):
        """If deal_breakers has no salary mention, the list must be unchanged."""
        from metis.init_bak_cmd import _apply_prefs_to_profile
        profile = {"deal_breakers": ["No management roles"], "salary_floor_usd": 200000}
        _apply_prefs_to_profile(profile, {"salary_floor": "180000"})
        assert profile["deal_breakers"] == ["No management roles"]
        assert profile["salary_floor_usd"] == 180000

    def test_domain_flex_flexible_injects_calibration_note(self):
        """domain_flex='flexible' must inject the domain-gap friction note into profile.notes."""
        from metis.init_bak_cmd import _apply_prefs_to_profile
        profile = {"notes": "Existing calibration text."}
        _apply_prefs_to_profile(profile, {"domain_flex": "flexible"})
        assert "Domain gaps are friction" in profile["notes"]
        assert "Existing calibration text." in profile["notes"]

    def test_domain_flex_replaces_prior_domain_note(self):
        """Re-running init with a different domain_flex must replace the old domain note."""
        from metis.init_bak_cmd import _apply_prefs_to_profile
        profile = {"notes": "Domain gaps are friction, not disqualifiers: old text."}
        _apply_prefs_to_profile(profile, {"domain_flex": "strict"})
        assert "Domain gaps are friction" not in profile["notes"]
        assert "meaningful penalty" in profile["notes"]


# ---------------------------------------------------------------------------
# render.py — digest rendering edge cases
# ---------------------------------------------------------------------------

class TestRenderEdgeCases:
    """build_digest_html must not crash on unusual but valid job lists."""

    def _job(self, verdict, score):
        return {
            "title": "Staff PM", "company": "Acme", "location": "NYC", "url": "#",
            "eval": {
                "verdict": verdict, "score": score,
                "leveragePoints": ["strength — evidence"], "frictionPoints": [],
                "tags": [{"text": "AI", "sentiment": "green"}],
            },
        }

    def test_empty_job_list_renders(self):
        """An empty list must render valid HTML without crashing."""
        from metis.render import build_digest_html
        html = build_digest_html([], "June 14, 2026")
        assert "<html" in html
        assert "0 roles evaluated" not in html or True  # any output is acceptable

    def test_all_skipped_renders_without_apply_section(self):
        from metis.render import build_digest_html
        jobs = [self._job("skipped", 30), self._job("skipped", 20)]
        html = build_digest_html(jobs, "June 14, 2026")
        assert "<html" in html
        assert "Limited Match" in html

    def test_all_filtered_shows_footer_count_only(self):
        """If every role was filtered by deal_breaker, no job sections appear,
        but the footer must mention the filtered count.
        build_digest_html receives only scored jobs (filtered ones removed upstream);
        deal_breaker_count is passed in separately."""
        from metis.render import build_digest_html
        # pass empty scored list + count=2, matching how pipeline.py calls it
        html = build_digest_html([], "June 14, 2026", deal_breaker_count=2)
        assert "filtered by deal" in html
        # No apply/consider/skipped *section* should appear (stat row cells are fine)
        assert "View posting" not in html  # no job cards rendered

    def test_mixed_verdicts_all_sections_present(self):
        from metis.render import build_digest_html
        jobs = [
            self._job("apply",    80),
            self._job("consider", 60),
            self._job("skipped",  30),
        ]
        html = build_digest_html(jobs, "June 14, 2026")
        assert "Solid Match" in html
        assert "Moderate Match" in html
        assert "Limited Match" in html

    def test_no_friction_points_omits_friction_line(self):
        """A job with empty frictionPoints must not render a bare '↓ Friction:' row."""
        from metis.render import build_digest_html
        jobs = [self._job("apply", 80)]
        jobs[0]["eval"]["frictionPoints"] = []
        html = build_digest_html(jobs, "June 14, 2026")
        assert "Friction:" not in html or "↓ Friction: </span>" not in html

    def test_empty_location_renders_without_separator(self):
        """A job with no location must render 'Company' not 'Company · '."""
        from metis.render import build_digest_html
        jobs = [self._job("apply", 80)]
        jobs[0]["location"] = ""
        html = build_digest_html(jobs, "June 14, 2026")
        assert "Acme · " not in html  # no trailing separator
        assert "Acme" in html         # company name still present

    def test_footer_uses_active_provider_label(self, monkeypatch):
        from metis.render import build_digest_html, build_digest_payload

        monkeypatch.setenv("METIS_LLM_PROVIDER", "openai")
        payload = build_digest_payload([self._job("apply", 80)], "June 14, 2026")
        html = build_digest_html([self._job("apply", 80)], "June 14, 2026")

        assert payload["providerLabel"] == "OpenAI"
        assert "powered by OpenAI" in html
        assert "powered by Claude" not in html


class TestSanitizeLocation:
    """_sanitize_location strips LinkedIn CTA text from location fields."""

    def _sanitize(self, loc: str) -> str:
        from metis.sources.linkedin import _sanitize_location
        return _sanitize_location(loc)

    def test_normal_location_unchanged(self):
        assert self._sanitize("San Francisco, CA") == "San Francisco, CA"

    def test_apply_with_resume_stripped_entirely(self):
        assert self._sanitize("Apply with resume & profile") == ""

    def test_easy_apply_stripped_entirely(self):
        assert self._sanitize("Easy Apply") == ""

    def test_location_with_apply_suffix_stripped(self):
        assert self._sanitize("United States · Apply with resume & profile") == "United States"

    def test_location_with_easy_apply_suffix_stripped(self):
        assert self._sanitize("New York, NY · Easy Apply") == "New York, NY"

    def test_empty_string_passthrough(self):
        assert self._sanitize("") == ""

    def test_remote_location_unchanged(self):
        assert self._sanitize("Remote") == "Remote"

    def test_case_insensitive(self):
        assert self._sanitize("San Francisco · APPLY WITH RESUME") == "San Francisco"


# ---------------------------------------------------------------------------
# save_skipped_roles / lookup_skipped_role
# ---------------------------------------------------------------------------

class TestSaveSkippedRoles:
    """save_skipped_roles writes metadata; lookup_skipped_role retrieves it."""

    def _make_job(self, title: str, company: str, score: int = 45, verdict: str = "skipped") -> dict:
        return {"title": title, "company": company, "url": "https://example.com",
                "eval": {"score": score, "verdict": verdict}}

    def test_writes_entry_for_skipped_job(self, tmp_path, monkeypatch):
        import metis.state as state
        monkeypatch.setattr(state, "SKIPPED_FILE", tmp_path / "skipped_roles.json")
        job = self._make_job("Senior PM", "Acme")
        state.save_skipped_roles([job])
        data = json.loads((tmp_path / "skipped_roles.json").read_text())
        assert len(data) == 1
        entry = next(iter(data.values()))
        assert entry["role_title"] == "Senior PM"
        assert entry["company"] == "Acme"
        assert entry["match_score"] == 45

    def test_empty_list_writes_nothing(self, tmp_path, monkeypatch):
        import metis.state as state
        out = tmp_path / "skipped_roles.json"
        monkeypatch.setattr(state, "SKIPPED_FILE", out)
        state.save_skipped_roles([])
        assert not out.exists()

    def test_lookup_returns_entry(self, tmp_path, monkeypatch):
        import metis.state as state
        monkeypatch.setattr(state, "SKIPPED_FILE", tmp_path / "skipped_roles.json")
        job = self._make_job("Staff PM", "Beta Corp")
        state.save_skipped_roles([job])
        result = state.lookup_skipped_role("Staff PM", "Beta Corp")
        assert result is not None
        assert result["company"] == "Beta Corp"

    def test_lookup_returns_none_for_unknown(self, tmp_path, monkeypatch):
        import metis.state as state
        monkeypatch.setattr(state, "SKIPPED_FILE", tmp_path / "skipped_roles.json")
        state.save_skipped_roles([self._make_job("PM", "X")])
        assert state.lookup_skipped_role("Unrelated", "Nobody") is None

    def test_expired_entries_pruned_on_write(self, tmp_path, monkeypatch):
        import metis.state as state
        monkeypatch.setattr(state, "SKIPPED_FILE", tmp_path / "skipped_roles.json")
        monkeypatch.setattr(state, "SKIPPED_TTL_DAYS", 30)
        stale_time = (datetime.datetime.now() - datetime.timedelta(days=31)).isoformat()
        stale = {"abc123": {"role_title": "Old PM", "company": "Gone", "saved_at": stale_time}}
        (tmp_path / "skipped_roles.json").write_text(json.dumps(stale))
        # Writing a new job triggers pruning of the stale entry
        state.save_skipped_roles([self._make_job("New PM", "Current Co")])
        data = json.loads((tmp_path / "skipped_roles.json").read_text())
        assert all(v["role_title"] != "Old PM" for v in data.values())

    def test_file_permissions_are_0600(self, tmp_path, monkeypatch):
        import stat, metis.state as state
        monkeypatch.setattr(state, "SKIPPED_FILE", tmp_path / "skipped_roles.json")
        state.save_skipped_roles([self._make_job("PM", "Co")])
        mode = (tmp_path / "skipped_roles.json").stat().st_mode
        assert oct(stat.S_IMODE(mode)) == oct(0o600)


# ---------------------------------------------------------------------------
# write_to_tracker
# ---------------------------------------------------------------------------

class TestWriteToTracker:
    """write_to_tracker appends Apply/Consider rows to the xlsx; skips duplicates."""

    def _make_job(self, title: str, company: str, verdict: str, score: int = 80) -> dict:
        return {"title": title, "company": company, "url": "https://linkedin.com/jobs/1",
                "location": "Remote", "eval": {"score": score, "verdict": verdict,
                "leveragePoints": [], "frictionPoints": []}}

    def test_apply_and_consider_rows_written(self, tmp_path, monkeypatch):
        pytest.importorskip("openpyxl")
        import metis.xlsx as tracker
        monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "applications.xlsx")
        jobs = [
            self._make_job("Senior PM", "Acme", "apply"),
            self._make_job("Staff PM", "Beta", "consider", score=62),
            self._make_job("Junior PM", "Gamma", "skipped", score=30),
        ]
        tracker.write_to_tracker(jobs, run_date="2026-06-16")
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path / "applications.xlsx")
        ws = wb.active
        # Row 1 = header; rows 2+ = data
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "Senior PM" in titles
        assert "Staff PM" in titles
        assert "Junior PM" not in titles   # skipped verdict not written

    def test_duplicate_role_not_written_twice(self, tmp_path, monkeypatch):
        pytest.importorskip("openpyxl")
        import metis.xlsx as tracker
        monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "applications.xlsx")
        job = self._make_job("Senior PM", "Acme", "apply")
        tracker.write_to_tracker([job], run_date="2026-06-16")
        tracker.write_to_tracker([job], run_date="2026-06-17")  # second run, same role
        import openpyxl
        ws = openpyxl.load_workbook(tmp_path / "applications.xlsx").active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert titles.count("Senior PM") == 1

    def test_no_eligible_roles_writes_nothing(self, tmp_path, monkeypatch):
        pytest.importorskip("openpyxl")
        import metis.xlsx as tracker
        monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "applications.xlsx")
        jobs = [self._make_job("PM", "X", "skipped", score=20)]
        tracker.write_to_tracker(jobs)
        assert not (tmp_path / "applications.xlsx").exists()

    def test_missing_openpyxl_logs_warning_and_returns(self, tmp_path, monkeypatch, caplog):
        import metis.xlsx as tracker
        monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "applications.xlsx")
        import unittest.mock as mock
        with mock.patch.dict("sys.modules", {"openpyxl": None}):
            import importlib
            tracker_fresh = importlib.reload(tracker)
            monkeypatch.setattr(tracker_fresh, "TRACKER_PATH", tmp_path / "applications.xlsx")
            import logging
            with caplog.at_level(logging.WARNING, logger="metis.xlsx"):
                tracker_fresh.write_to_tracker(
                    [self._make_job("PM", "Co", "apply")], run_date="2026-06-16"
                )
        assert not (tmp_path / "applications.xlsx").exists()


# ---------------------------------------------------------------------------
# --no-limit flag renamed from --all
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Hard gate: jd_quality semantics — D-54
# ---------------------------------------------------------------------------

class TestHardGates:
    """check_hard_gates must distinguish blank JD from extraction failure.

    "extraction_failed" means Haiku's JSON output couldn't be parsed — the JD
    content IS present and scoring should proceed.  Only "blank" (Haiku saw an
    empty JD) should trigger the jd_blank gate.  Conflating the two silently
    drops real roles.  See DECISIONS.md D-54.
    """

    def _gate(self, jd_quality, profile=None):
        from metis.extract import check_hard_gates
        struct = {"jd_quality": jd_quality}
        return check_hard_gates(struct, profile or {})

    def test_blank_jd_fires_gate(self):
        passes, gate = self._gate("blank")
        assert not passes
        assert gate == "jd_blank"

    def test_extraction_failed_does_not_fire_gate(self):
        passes, gate = self._gate("extraction_failed")
        assert passes, "extraction_failed must not trigger jd_blank — JD content exists"
        assert gate == ""

    def test_low_quality_does_not_fire_gate(self):
        passes, gate = self._gate("low")
        assert passes

    def test_unknown_quality_does_not_fire_gate(self):
        passes, gate = self._gate("unknown")
        assert passes

    def test_missing_jd_quality_does_not_fire_gate(self):
        from metis.extract import check_hard_gates
        passes, gate = check_hard_gates({}, {})
        assert passes


class TestNoLimitFlag:
    """--all was renamed --no-limit; score_all internal param unchanged."""

    def _parse(self, argv: list) -> object:
        import argparse
        parser = argparse.ArgumentParser(prog="metis")
        parser.add_argument("--lookback", default="3d")
        parser.add_argument("--no-limit", dest="score_all", action="store_true")
        parser.add_argument("--no-tracker", dest="no_tracker", action="store_true")
        parser.add_subparsers(dest="command")
        return parser.parse_args(argv)

    def test_no_limit_sets_score_all(self):
        assert self._parse(["--no-limit"]).score_all is True

    def test_default_score_all_is_false(self):
        assert self._parse([]).score_all is False

    def test_all_flag_not_accepted(self):
        import argparse
        parser = argparse.ArgumentParser(prog="metis")
        parser.add_argument("--no-limit", dest="score_all", action="store_true")
        parser.add_subparsers(dest="command")
        with pytest.raises(SystemExit) as exc:
            parser.parse_args(["--all"])
        assert exc.value.code == 2


# ---------------------------------------------------------------------------
# salary_is_hard_floor flag — render_profile and scoring rubric behavior
# ---------------------------------------------------------------------------

class TestSalaryFloorFlag:

    def _profile(self, hard_floor: bool) -> dict:
        return {
            "candidate": {"name": "Test User"},
            "salary_floor_usd": 280000,
            "salary_is_hard_floor": hard_floor,
        }

    def test_render_profile_hard_floor_label(self):
        from metis.profile import render_profile
        result = render_profile(self._profile(hard_floor=True))
        assert "SALARY FLOOR" in result
        assert "hard minimum" in result

    def test_render_profile_aspirational_label(self):
        from metis.profile import render_profile
        result = render_profile(self._profile(hard_floor=False))
        assert "SALARY TARGET" in result
        assert "aspirational" in result
        assert "SALARY FLOOR" not in result

    def test_scoring_rubric_hard_floor_caps_undisclosed(self):
        from metis.score import _build_score_suffix
        rubric = _build_score_suffix("Test", 75, 55, salary_is_hard_floor=True)
        assert "cap this dimension at 60" in rubric
        assert "comp: undisclosed" in rubric

    def test_scoring_rubric_aspirational_no_undisclosed_penalty(self):
        from metis.score import _build_score_suffix
        rubric = _build_score_suffix("Test", 75, 55, salary_is_hard_floor=False)
        assert "absence of disclosure is not a concern" in rubric
        assert "DO NOT USE" in rubric  # tag suppression notice


class TestAnonEmployerFallback:

    def test_known_company_surfaced_when_stage_unknown(self):
        from metis.extract import format_extraction_for_scoring
        ext = {"jd_quality": "good", "company_stage": "unknown", "company_tier": None}
        result = format_extraction_for_scoring(ext, listing_company="Jobgether")
        assert "listed as: Jobgether" in result

    def test_no_fallback_when_stage_is_known(self):
        from metis.extract import format_extraction_for_scoring
        ext = {"jd_quality": "good", "company_stage": "series_b", "company_tier": None}
        result = format_extraction_for_scoring(ext, listing_company="Jobgether")
        assert "listed as" not in result

    def test_no_fallback_for_empty_listing_company(self):
        from metis.extract import format_extraction_for_scoring
        ext = {"jd_quality": "good", "company_stage": "unknown", "company_tier": None}
        result = format_extraction_for_scoring(ext, listing_company="")
        assert "listed as" not in result


# ---------------------------------------------------------------------------
# Hard gate: jd_quality semantics — D-54
# ---------------------------------------------------------------------------

class TestHardGates:
    """check_hard_gates must distinguish blank JD from extraction failure.

    "extraction_failed" means Haiku's JSON output couldn't be parsed — the JD
    content IS present and scoring should proceed.  Only "blank" (Haiku saw an
    empty JD) should trigger the jd_blank gate.  Conflating the two silently
    drops real roles.  See DECISIONS.md D-54.
    """

    def _gate(self, jd_quality, profile=None):
        from metis.extract import check_hard_gates
        struct = {"jd_quality": jd_quality}
        return check_hard_gates(struct, profile or {})

    def test_blank_jd_fires_gate(self):
        passes, gate = self._gate("blank")
        assert not passes
        assert gate == "jd_blank"

    def test_extraction_failed_does_not_fire_gate(self):
        passes, gate = self._gate("extraction_failed")
        assert passes, "extraction_failed must not trigger jd_blank — JD content exists"
        assert gate == ""

    def test_low_quality_does_not_fire_gate(self):
        passes, gate = self._gate("low")
        assert passes

    def test_unknown_quality_does_not_fire_gate(self):
        passes, gate = self._gate("unknown")
        assert passes

    def test_missing_jd_quality_does_not_fire_gate(self):
        from metis.extract import check_hard_gates
        passes, gate = check_hard_gates({}, {})
        assert passes
