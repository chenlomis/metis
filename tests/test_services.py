from __future__ import annotations

import datetime
import json
from pathlib import Path
import ast
import logging
from concurrent.futures import ThreadPoolExecutor


def test_get_metis_status_reports_missing_setup(tmp_path):
    from metis.services import get_metis_status

    result = get_metis_status(
        data_dir=tmp_path,
        env={"METIS_LLM_PROVIDER": "anthropic"},
    )

    assert result["configured"] is False
    assert result["missing"] == ["profile", "llm_api_key", "email_access"]
    assert result["checks"]["profile_exists"] is False
    assert result["checks"]["email_access_configured"] is False
    assert "metis init" in result["next_steps"][0]


def test_get_metis_status_accepts_explicit_paths_and_oauth(tmp_path):
    from metis.services import get_metis_status

    profile_path = tmp_path / "designer.yaml"
    tracker_path = tmp_path / "tracker.xlsx"
    profile_path.write_text("candidate:\n  name: Test User\n", encoding="utf-8")
    tracker_path.write_text("placeholder", encoding="utf-8")
    (tmp_path / "email_provider.json").write_text(
        json.dumps({"provider": "gmail_oauth"}),
        encoding="utf-8",
    )
    (tmp_path / "gmail_token.json").write_text("{}", encoding="utf-8")

    result = get_metis_status(
        data_dir=tmp_path,
        profile_path=profile_path,
        tracker_path=tracker_path,
        env={"ANTHROPIC_API_KEY": "test-key"},
    )

    assert result["configured"] is True
    assert result["missing"] == []
    assert result["paths"]["profile"] == str(profile_path)
    assert result["email"]["oauth_provider"] == "gmail_oauth"
    assert result["email"]["oauth_configured"] is True


def test_get_metis_status_warns_on_profile_data_dir_mismatch(tmp_path):
    from metis.services import get_metis_status

    data_dir = tmp_path / "pm"
    other_dir = tmp_path / "designer"
    data_dir.mkdir()
    other_dir.mkdir()
    profile_path = other_dir / "profile.yaml"
    tracker_path = other_dir / "applications.xlsx"
    profile_path.write_text("candidate:\n  name: Designer\n", encoding="utf-8")
    tracker_path.write_text("placeholder", encoding="utf-8")

    result = get_metis_status(
        data_dir=data_dir,
        profile_path=profile_path,
        tracker_path=tracker_path,
        env={
            "ANTHROPIC_API_KEY": "test-key",
            "GMAIL_ADDRESS": "user@example.com",
            "GMAIL_APP_PASSWORD": "test-password",
        },
    )

    assert result["configured"] is True
    assert any("profile_path is outside data_dir" in warning for warning in result["warnings"])
    assert any("tracker_path is outside data_dir" in warning for warning in result["warnings"])


def test_list_recommended_roles_dedupes_and_marks_tailoring(tmp_path):
    from metis.services import get_role_details, list_recommended_roles

    runs = [
        {
            "ts": "2026-07-01T10:00:00",
            "role_hash": "role_old",
            "title": "Product Manager",
            "company": "Acme",
            "eval": {"verdict": "apply", "score": 91},
        },
        {
            "ts": "2026-07-01T11:00:00",
            "role_hash": "role_skip",
            "title": "Sales Manager",
            "company": "Beta",
            "eval": {"verdict": "skipped", "score": 42},
        },
        {
            "ts": "2026-07-01T12:00:00",
            "role_hash": "role_old",
            "title": "Product Manager",
            "company": "Acme",
            "url": "https://example.com/jobs/1",
            "eval": {
                "verdict": "consider",
                "score": 73,
                "frictionPoints": ["Domain is adjacent"],
                "leveragePoints": ["Strong customer context"],
            },
        },
    ]
    (tmp_path / "runs.jsonl").write_text(
        "\n".join(json.dumps(row) for row in runs),
        encoding="utf-8",
    )

    result = list_recommended_roles(data_dir=tmp_path)

    assert result["count"] == 1
    role = result["roles"][0]
    assert role["role_id"] == "role_old"
    assert role["score"] == 73
    assert role["verdict"] == "consider"
    assert role["tailoring_available"] is True

    detail = get_role_details("role_old", data_dir=tmp_path)
    assert detail is not None
    assert detail["ts"] == "2026-07-01T12:00:00"


def test_list_recommended_roles_sorts_apply_before_consider_by_score(tmp_path):
    from metis.services import list_recommended_roles

    runs = [
        {
            "role_hash": "consider_high",
            "title": "Senior Product Manager",
            "company": "Beta",
            "eval": {"verdict": "consider", "score": 88},
        },
        {
            "role_hash": "apply_low",
            "title": "Staff Product Manager",
            "company": "Acme",
            "eval": {"verdict": "apply", "score": 77},
        },
        {
            "role_hash": "apply_high",
            "title": "Principal Product Manager",
            "company": "Gamma",
            "eval": {"verdict": "apply", "score": 91},
        },
    ]
    (tmp_path / "runs.jsonl").write_text(
        "\n".join(json.dumps(row) for row in runs),
        encoding="utf-8",
    )

    result = list_recommended_roles(data_dir=tmp_path)

    assert [role["role_id"] for role in result["roles"]] == [
        "apply_high",
        "apply_low",
        "consider_high",
    ]


def test_list_recommended_roles_defaults_to_last_run_when_available(tmp_path):
    from metis.services import list_recommended_roles

    runs = [
        {
            "role_hash": "old_role",
            "title": "Principal Product Manager",
            "company": "OldCo",
            "eval": {"verdict": "apply", "score": 95},
        },
        {
            "role_hash": "latest_role",
            "title": "Staff Product Manager",
            "company": "NewCo",
            "eval": {"verdict": "apply", "score": 82},
        },
    ]
    (tmp_path / "runs.jsonl").write_text(
        "\n".join(json.dumps(row) for row in runs),
        encoding="utf-8",
    )
    (tmp_path / "last_run.json").write_text(
        json.dumps({
            "roles": [
                {
                    "title": "Staff Product Manager",
                    "company": "NewCo",
                    "score": 82,
                    "verdict": "apply",
                }
            ]
        }),
        encoding="utf-8",
    )

    result = list_recommended_roles(data_dir=tmp_path)

    assert [role["role_id"] for role in result["roles"]] == ["latest_role"]


def test_list_scoring_feedback_parses_recent_entries(tmp_path):
    from metis.services import list_scoring_feedback

    (tmp_path / "feedback.md").write_text(
        "# Scoring Feedback\n\n"
        "<!-- id:fb_1 | run:run_a | roles:acme | dims:culture_values -->\n"
        "## [user] 2026-07-01\n\n"
        "Score AI infrastructure roles higher.\n\n"
        "<!-- id:fb_2 | run:run_b | roles:beta,gamma | dims:domain_background -->\n"
        "## [user] 2026-07-02\n\n"
        "Penalize generic B2C marketplaces.\n",
        encoding="utf-8",
    )

    result = list_scoring_feedback(data_dir=tmp_path, limit=1)

    assert result["count"] == 1
    entry = result["entries"][0]
    assert entry["feedback_id"] == "fb_2"
    assert entry["date"] == "2026-07-02"
    assert entry["summary"] == "Penalize generic B2C marketplaces."
    assert entry["metadata"]["roles"] == ["beta", "gamma"]


def test_list_application_activity_reads_tracker_rows(tmp_path):
    from openpyxl import Workbook

    from metis.services import list_application_activity

    tracker = tmp_path / "applications.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "date_suggested",
        "role_title",
        "company",
        "match_score",
        "suggestion_status",
        "action_taken",
        "date_applied",
        "application_status",
        "notes",
    ])
    ws.append([
        "2026-07-02",
        "Product Manager",
        "Acme",
        88,
        "Solid Match",
        "Applied",
        "2026-07-03",
        "Pending",
        "Referral",
    ])
    wb.save(tracker)

    result = list_application_activity(tracker_path=tracker)

    assert result["count"] == 1
    assert result["applications"][0]["company"] == "Acme"
    assert result["applications"][0]["application_status"] == "Pending"


def test_record_scoring_feedback_appends_entry_and_log(tmp_path):
    from metis.services import list_scoring_feedback, record_scoring_feedback

    result = record_scoring_feedback(
        "Score AI infrastructure platform roles higher.",
        data_dir=tmp_path,
        run_id="run_20260703",
        roles=["snowflake", "twilio"],
        dims=["domain_background"],
        feedback_id="fb_test_123",
        today=datetime.date(2026, 7, 3),
    )

    assert result["saved"] is True
    assert result["feedback_id"] == "fb_test_123"
    content = (tmp_path / "feedback.md").read_text(encoding="utf-8")
    assert "<!-- id:fb_test_123 | run:run_20260703 | roles:snowflake,twilio | dims:domain_background -->" in content
    assert "## [user] 2026-07-03" in content
    assert "Score AI infrastructure platform roles higher." in content

    log_record = json.loads((tmp_path / "feedback_log.jsonl").read_text(encoding="utf-8"))
    assert log_record["feedback_id"] == "fb_test_123"
    assert log_record["roles"] == ["snowflake", "twilio"]
    assert log_record["source"] == "service"

    listed = list_scoring_feedback(data_dir=tmp_path)
    assert listed["entries"][0]["feedback_id"] == "fb_test_123"


def test_record_scoring_feedback_rejects_empty_text(tmp_path):
    from metis.services import record_scoring_feedback

    try:
        record_scoring_feedback("   ", data_dir=tmp_path)
    except ValueError as exc:
        assert "cannot be empty" in str(exc)
    else:
        raise AssertionError("Expected empty feedback to raise ValueError")


def test_record_scoring_feedback_sanitizes_comment_metadata(tmp_path):
    from metis.services import record_scoring_feedback

    record_scoring_feedback(
        "Prefer vertical AI roles.",
        data_dir=tmp_path,
        feedback_id="fb_meta",
        run_id="run_1",
        roles=["Acme -->\n## injected", "Beta|Gamma"],
        dims=["domain,background", "<culture>"],
        today=datetime.date(2026, 7, 4),
    )

    content = (tmp_path / "feedback.md").read_text(encoding="utf-8")
    comment = next(line for line in content.splitlines() if line.startswith("<!-- id:fb_meta"))

    assert comment.startswith("<!-- id:fb_meta")
    assert "-->" == comment[-3:]
    assert "## injected" not in comment
    assert "<culture>" not in comment
    assert "Beta/Gamma" in comment


def test_record_scoring_feedback_serializes_concurrent_writes(tmp_path):
    from metis.services import list_scoring_feedback, record_scoring_feedback

    def write_one(idx: int) -> None:
        record_scoring_feedback(
            f"Calibration note {idx}",
            data_dir=tmp_path,
            feedback_id=f"fb_{idx:02d}",
            today=datetime.date(2026, 7, 4),
        )

    with ThreadPoolExecutor(max_workers=8) as pool:
        list(pool.map(write_one, range(20)))

    content = (tmp_path / "feedback.md").read_text(encoding="utf-8")
    log_lines = (tmp_path / "feedback_log.jsonl").read_text(encoding="utf-8").splitlines()
    listed = list_scoring_feedback(data_dir=tmp_path, limit=25)

    assert content.count("Calibration note") == 20
    assert len(log_lines) == 20
    assert listed["count"] == 20


def test_run_job_search_reports_missing_configuration(tmp_path):
    from metis.services import run_job_search

    result = run_job_search(data_dir=tmp_path, env={})

    assert result["ran"] is False
    assert result["status"] == "missing_configuration"
    assert set(result["missing"]) == {"profile", "ANTHROPIC_API_KEY", "gmail_credentials"}


def test_run_job_search_returns_invalid_input_for_bad_lookback(tmp_path):
    from metis.services import run_job_search

    profile = tmp_path / "profile.yaml"
    profile.write_text("candidate:\n  name: Test User\n", encoding="utf-8")

    result = run_job_search(
        data_dir=tmp_path,
        profile_path=profile,
        lookback="garbage",
        env={
            "ANTHROPIC_API_KEY": "test-key",
            "GMAIL_ADDRESS": "user@example.com",
            "GMAIL_APP_PASSWORD": "test-password",
        },
    )

    assert result["ran"] is False
    assert result["status"] == "invalid_input"
    assert "lookback" in result["error"]


def test_run_job_search_requires_confirmation_for_write_mode(tmp_path):
    from metis.services import run_job_search

    profile = tmp_path / "profile.yaml"
    profile.write_text("candidate:\n  name: Test User\n", encoding="utf-8")

    result = run_job_search(
        data_dir=tmp_path,
        profile_path=profile,
        dry_run=False,
        env={
            "ANTHROPIC_API_KEY": "test-key",
            "GMAIL_ADDRESS": "user@example.com",
            "GMAIL_APP_PASSWORD": "test-password",
        },
    )

    assert result["ran"] is False
    assert result["status"] == "confirmation_required"


def test_run_job_search_invokes_pipeline_with_explicit_runtime(tmp_path, monkeypatch):
    from metis.services import run_job_search
    from metis import pipeline, profile, state, xlsx

    profile_path = tmp_path / "profile.yaml"
    tracker_path = tmp_path / "tracker.xlsx"
    profile_path.write_text("candidate:\n  name: Test User\n", encoding="utf-8")
    calls = []

    def fake_run_pipeline(since_dt, score_all=False, dry_run=False):
        calls.append({
            "since": since_dt,
            "score_all": score_all,
            "dry_run": dry_run,
            "data_dir": state.DATA_DIR,
            "profile": profile.YAML_PATH,
            "tracker": xlsx.TRACKER_PATH,
            "gmail": pipeline.GMAIL_ADDRESS,
        })
        print("pipeline preview")

    monkeypatch.setattr(pipeline, "run_pipeline", fake_run_pipeline)

    result = run_job_search(
        data_dir=tmp_path,
        profile_path=profile_path,
        tracker_path=tracker_path,
        lookback="2d",
        score_all=True,
        dry_run=True,
        env={
            "ANTHROPIC_API_KEY": "test-key",
            "GMAIL_ADDRESS": "user@example.com",
            "GMAIL_APP_PASSWORD": "test-password",
        },
    )

    assert result["ran"] is True
    assert result["status"] == "completed"
    assert result["dry_run"] is True
    assert result["score_all"] is True
    assert result["stdout"] == "pipeline preview"
    assert len(calls) == 1
    assert calls[0]["dry_run"] is True
    assert calls[0]["data_dir"] == tmp_path
    assert calls[0]["profile"] == profile_path
    assert calls[0]["tracker"] == tracker_path
    assert calls[0]["gmail"] == "user@example.com"


def test_run_job_search_accepts_oauth_email_state(tmp_path, monkeypatch):
    from metis.services import run_job_search
    from metis import pipeline

    profile_path = tmp_path / "profile.yaml"
    profile_path.write_text("candidate:\n  name: Test User\n", encoding="utf-8")
    (tmp_path / "email_provider.json").write_text(json.dumps({"provider": "gmail_oauth"}), encoding="utf-8")
    (tmp_path / "gmail_token.json").write_text("{}", encoding="utf-8")
    calls = []

    def fake_run_pipeline(since_dt, score_all=False, dry_run=False):
        calls.append((since_dt, score_all, dry_run))

    monkeypatch.setattr(pipeline, "run_pipeline", fake_run_pipeline)

    result = run_job_search(
        data_dir=tmp_path,
        profile_path=profile_path,
        dry_run=True,
        env={"ANTHROPIC_API_KEY": "test-key"},
    )

    assert result["ran"] is True
    assert result["status"] == "completed"
    assert len(calls) == 1


def test_run_job_search_returns_failed_status_for_pipeline_exit(tmp_path, monkeypatch):
    from metis.services import run_job_search
    from metis import pipeline

    profile_path = tmp_path / "profile.yaml"
    profile_path.write_text("candidate:\n  name: Test User\n", encoding="utf-8")

    def fake_run_pipeline(*args, **kwargs):
        print("scored before failure")
        raise SystemExit("SMTP failed")

    monkeypatch.setattr(pipeline, "run_pipeline", fake_run_pipeline)

    result = run_job_search(
        data_dir=tmp_path,
        profile_path=profile_path,
        dry_run=False,
        confirm_send=True,
        env={
            "ANTHROPIC_API_KEY": "test-key",
            "GMAIL_ADDRESS": "user@example.com",
            "GMAIL_APP_PASSWORD": "test-password",
        },
    )

    assert result["ran"] is True
    assert result["status"] == "failed"
    assert result["error_type"] == "SystemExit"
    assert result["error"] == "SMTP failed"
    assert result["stdout"] == "scored before failure"


def test_linkedin_fetch_uses_oauth_fetcher_without_imap_credentials(monkeypatch):
    from metis.sources import email_fetcher, linkedin

    monkeypatch.setattr(linkedin, "_GMAIL_ADDRESS_ENV", "")
    monkeypatch.setattr(linkedin, "_GMAIL_APP_PASSWORD_ENV", "")
    monkeypatch.setattr(email_fetcher, "get_provider", lambda: "gmail_oauth")
    monkeypatch.setattr(
        email_fetcher,
        "fetch_emails_from_sender",
        lambda sender, since_dt: [{"text": f"Body from {sender}", "html": "", "subject": "LinkedIn", "date": "2026-07-04T00:00:00"}],
    )

    threads = linkedin.fetch_linkedin_alerts_since(datetime.datetime(2026, 7, 1))

    assert len(threads) == 3
    assert threads[0]["subject"] == "LinkedIn"
    assert "jobalerts-noreply@linkedin.com" in threads[0]["body"]


def test_mcp_server_exposes_contract_tool_names():
    import metis.mcp_server as mcp_server

    source = Path(mcp_server.__file__).read_text(encoding="utf-8")
    tree = ast.parse(source)
    tool_names = {
        node.name
        for node in ast.walk(tree)
        if isinstance(node, ast.FunctionDef)
        and any(
            isinstance(decorator, ast.Call)
            and isinstance(decorator.func, ast.Attribute)
            and decorator.func.attr == "tool"
            for decorator in node.decorator_list
        )
    }

    assert tool_names == {
        "get_metis_status",
        "run_job_search",
        "list_recommended_roles",
        "get_role_details",
        "record_scoring_feedback",
        "list_scoring_feedback",
        "track_applications",
        "list_application_activity",
        "generate_progress_summary",
    }
    assert not (tool_names & {"init", "schedule", "reset", "feedback", "metis"})


def test_track_applications_requires_current_tracking_credentials(tmp_path):
    from metis.services import track_applications

    result = track_applications(data_dir=tmp_path, env={}, dry_run=True)

    assert result["ran"] is False
    assert result["status"] == "missing_credentials"
    assert "GMAIL_ADDRESS" in result["message"]


def test_track_applications_rejects_non_positive_lookback(tmp_path):
    from metis.services import track_applications

    result = track_applications(data_dir=tmp_path, lookback_days=0, dry_run=True)

    assert result["ran"] is False
    assert result["status"] == "invalid_input"
    assert "positive integer" in result["message"]


def test_track_applications_dry_run_returns_preview_without_writing(tmp_path, monkeypatch):
    from metis.services import track_applications
    from metis import track

    monkeypatch.setattr(track, "backfill_from_digests", lambda *args, **kwargs: 0)
    monkeypatch.setattr(track, "_build_llm_client", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        track,
        "fetch_candidate_emails",
        lambda *args, **kwargs: [{"subject": "Thanks", "sender": "jobs@example.com", "body": ""}],
    )
    monkeypatch.setattr(
        track,
        "parse_email",
        lambda *_args, **_kwargs: {
            "classification": "confirmation",
            "company": "Acme",
            "role": "Product Manager",
            "date": "2026-07-04",
        },
    )

    result = track_applications(
        data_dir=tmp_path,
        gmail_address="user@example.com",
        app_password="test-password",
        since_dt=datetime.datetime(2026, 7, 1),
        dry_run=True,
    )

    assert result["ran"] is True
    assert result["status"] == "completed"
    assert result["warnings"] == []
    assert result["dry_run"] is True
    assert result["rows_before"] == 0
    assert result["rows_after"] == 0
    assert "CONFIRMATION" in result["stdout"]
    assert not (tmp_path / "applications.xlsx").exists()


def test_track_applications_surfaces_scan_warnings(tmp_path, monkeypatch):
    from metis.services import track_applications
    from metis import track

    def fake_fetch(*args, **kwargs):
        logging.getLogger("metis.track").warning("track: IMAP connect failed after 3 attempts: test network")
        return []

    monkeypatch.setattr(track, "_build_llm_client", lambda *args, **kwargs: None)
    monkeypatch.setattr(track, "fetch_candidate_emails", fake_fetch)

    result = track_applications(
        data_dir=tmp_path,
        gmail_address="user@example.com",
        app_password="test-password",
        since_dt=datetime.datetime(2026, 7, 1),
        dry_run=True,
    )

    assert result["ran"] is True
    assert result["status"] == "completed_with_warnings"
    assert result["rows_before"] == 0
    assert result["rows_after"] == 0
    assert result["rows_changed"] == 0
    assert result["warnings"] == ["track: IMAP connect failed after 3 attempts: test network"]


def test_track_applications_returns_failed_status_for_parser_crash(tmp_path, monkeypatch):
    from metis.services import track_applications
    from metis import track

    monkeypatch.setattr(track, "backfill_from_digests", lambda *args, **kwargs: 0)
    monkeypatch.setattr(track, "_build_llm_client", lambda *args, **kwargs: None)

    def crash_fetch(*args, **kwargs):
        raise RuntimeError("malformed email payload")

    monkeypatch.setattr(track, "fetch_candidate_emails", crash_fetch)

    result = track_applications(
        data_dir=tmp_path,
        gmail_address="user@example.com",
        app_password="test-password",
        since_dt=datetime.datetime(2026, 7, 1),
        dry_run=True,
    )

    assert result["ran"] is True
    assert result["status"] == "failed"
    assert result["error_type"] == "RuntimeError"
    assert result["error"] == "malformed email payload"
    assert result["rows_changed"] == 0


def test_track_applications_updates_tracker_with_explicit_path(tmp_path, monkeypatch):
    from openpyxl import Workbook, load_workbook

    from metis.services import track_applications
    from metis import track

    tracker = tmp_path / "custom_tracker.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "date_suggested",
        "role_title",
        "company",
        "match_score",
        "suggestion_status",
        "action_taken",
        "date_applied",
        "application_status",
        "notes",
    ])
    ws.append([
        "2026-07-01",
        "Product Manager",
        "Acme",
        0.88,
        "Solid Match",
        "Not Applied",
        None,
        None,
        None,
    ])
    wb.save(tracker)

    monkeypatch.setattr(track, "backfill_from_digests", lambda *args, **kwargs: 0)
    monkeypatch.setattr(track, "_build_llm_client", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        track,
        "fetch_candidate_emails",
        lambda *args, **kwargs: [{"subject": "Thanks", "sender": "jobs@example.com", "body": ""}],
    )
    monkeypatch.setattr(
        track,
        "parse_email",
        lambda *_args, **_kwargs: {
            "classification": "confirmation",
            "company": "Acme",
            "role": "Product Manager",
            "date": "2026-07-04",
        },
    )

    result = track_applications(
        data_dir=tmp_path,
        tracker_path=tracker,
        gmail_address="user@example.com",
        app_password="test-password",
        since_dt=datetime.datetime(2026, 7, 1),
        dry_run=False,
    )

    assert result["ran"] is True
    assert result["rows_before"] == 1
    assert result["rows_after"] == 1
    assert result["rows_changed"] == 1
    assert result["applications"][0]["action_taken"] == "Applied"
    assert result["applications"][0]["date_applied"] == "2026-07-04"
    assert result["applications"][0]["application_status"] == "Pending"

    saved = load_workbook(tracker, data_only=True).active
    assert saved.cell(2, 6).value == "Applied"
    assert saved.cell(2, 8).value == "Pending"


def test_generate_progress_summary_returns_tracker_metrics(tmp_path):
    from openpyxl import Workbook

    from metis.services import generate_progress_summary

    tracker = tmp_path / "applications.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "date_suggested",
        "role_title",
        "company",
        "match_score",
        "suggestion_status",
        "action_taken",
        "date_applied",
        "application_status",
        "notes",
    ])
    ws.append(["2026-07-03", "Principal PM", "Acme", 0.91, "Solid Match", "Applied", "2026-07-04", "Pending", None])
    ws.append(["2026-07-03", "Staff PM", "Beta", 0.84, "Solid Match", "Not Applied", None, None, None])
    ws.append(["2026-07-03", "Senior PM", "Gamma", 0.66, "Moderate Match", "Not Applied", None, None, None])
    ws.append(["2026-07-03", "Rejected PM", "Delta", 0.70, "Moderate Match", "Applied", "2026-07-04", "Rejected", None])
    wb.save(tracker)
    (tmp_path / "skipped_roles.json").write_text(
        json.dumps({"skip_1": {"role_title": "Skip"}}),
        encoding="utf-8",
    )

    result = generate_progress_summary(data_dir=tmp_path, tracker_path=tracker)

    assert result["metrics"]["roles_scored"] == 5
    assert result["metrics"]["total_applied"] == 2
    assert result["metrics"]["true_positive"] == 2
    assert result["metrics"]["true_negative"] == 1
    assert result["metrics"]["alignment_base"] == 5
    assert result["score_distribution"] == {"solid": 2, "moderate": 2, "partial": 1}
    assert result["application_pipeline"]["pending"] == 1
    assert result["application_pipeline"]["rejections"] == 1
    assert result["application_pipeline"]["daily"] == {"2026-07-04": 2}
    assert [item["company"] for item in result["not_applied_recommendations"]] == ["Beta", "Gamma"]


def test_generate_progress_summary_includes_market_intel_from_runs(tmp_path):
    from metis.services import generate_progress_summary

    now = datetime.datetime.now().isoformat(timespec="seconds")
    (tmp_path / "runs.jsonl").write_text(
        json.dumps({
            "ts": now,
            "company": "Acme AI",
            "eval": {
                "verdict": "apply",
                "leveragePoints": [
                    "LLM roadmap and agentic workflow platform ownership.",
                    "API and developer infrastructure strategy.",
                ],
            },
            "extraction": {
                "primary_execution_stack": ["roadmap", "technical_specs", "ml_ai"],
                "product_surface": ["api", "platform"],
                "company_tier": "large_private",
                "customer_type": "b2b",
                "inferred_structural_level": "staff",
                "salary_disclosed": True,
                "salary_min": 220000,
                "salary_max": 280000,
            },
        }) + "\n",
        encoding="utf-8",
    )

    result = generate_progress_summary(data_dir=tmp_path, lookback_days=30)
    intel = result["market_intel"]

    assert intel["profile_signals"]
    assert {item["signal"] for item in intel["profile_signals"]} >= {
        "AI / LLM Product Strategy",
        "Technical Depth",
    }
    assert {item["name"] for item in intel["skills_in_demand"]} >= {
        "Product Roadmap & Strategy",
        "Technical Specifications",
        "ML / AI Systems",
    }
    assert intel["verticals_and_markets"][0]["name"] == "Enterprise SaaS (Large Private)"
    assert intel["role_level_distribution"][0]["label"] == "Staff / Lead (target band)"
    assert intel["comp_snapshot"]["by_level"][0]["range"] == "$220K–$280K"
