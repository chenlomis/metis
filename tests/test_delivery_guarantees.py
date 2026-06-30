"""Tests for delivery ordering guarantees and track→xlsx integration.

Two critical invariants:

T-07  SMTP failure must NOT write seen_roles.json.
      Roles stay "unseen" so they re-score on the next run.
      If seen_roles is written before SMTP we silently lose roles.

T-08  track→xlsx integration: when track.py classifies an email as
      "confirmation", the matching row in the Applications tracker must
      flip action_taken → "Applied" and set date_applied.
"""
import datetime
import smtplib
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_job(verdict="apply", title="Staff PM", company="Stripe", score=82):
    return {
        "title":   title,
        "company": company,
        "url":     "https://stripe.com/jobs/1",
        "eval": {
            "verdict":         verdict,
            "score":           score,
            "dimensions":      [],
            "leveragePoints":  [],
            "frictionPoints":  [],
            "tags":            [],
            "gate":            None,
            "summary":         None,
        },
    }


def _make_tracker(tmp_path: Path, rows: list[dict]) -> Path:
    """Write a minimal Applications xlsx with the given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = [
        "date_suggested", "role_title", "company", "match_score",
        "suggestion_status", "action_taken", "date_applied",
        "application_status", "notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row_idx, 1, row.get("date_suggested", "2026-06-01"))
        ws.cell(row_idx, 2, row.get("role_title", ""))
        ws.cell(row_idx, 3, row.get("company", ""))
        ws.cell(row_idx, 4, row.get("match_score", 0.80))
        ws.cell(row_idx, 5, row.get("suggestion_status", "Solid Match"))
        ws.cell(row_idx, 6, row.get("action_taken", "Not Applied"))
        ws.cell(row_idx, 7, row.get("date_applied"))
        ws.cell(row_idx, 8, row.get("application_status"))
        ws.cell(row_idx, 9, row.get("notes"))
    p = tmp_path / "applications.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# T-07  SMTP failure → seen_roles NOT written
# ---------------------------------------------------------------------------

class TestT07SmtpFailureDoesNotWriteSeenRoles:
    """_stage_deliver must raise SystemExit(1) on SMTP failure without
    persisting seen_roles.json.  If the write happens first the roles are
    lost from the next run's scoring window."""

    def _run_stage(self, tmp_path, smtp_exc):
        """Invoke _stage_deliver with a mocked SMTP that raises smtp_exc."""
        jobs = [_make_job("apply"), _make_job("consider", title="Senior PM", company="Anthropic")]
        new_role_timestamps = {"abc123": "2026-06-28T10:00:00", "def456": "2026-06-28T10:00:01"}

        with (
            patch("metis.pipeline.render_html", return_value="<html>digest</html>"),
            patch("metis.pipeline.send_digest", side_effect=smtp_exc),
            patch("metis.pipeline.save_seen_roles") as mock_save,
            patch("metis.pipeline.save_skipped_roles"),
            patch("metis.pipeline.save_last_run"),
            patch("metis.xlsx.write_to_tracker"),
        ):
            from metis.pipeline import _stage_deliver
            with pytest.raises(SystemExit) as exc_info:
                _stage_deliver(jobs, n_filtered=0, new_role_timestamps=new_role_timestamps)

        assert exc_info.value.code == 1, "should exit with code 1, not 0"
        mock_save.assert_not_called()

    def test_auth_failure_does_not_write_seen_roles(self, tmp_path):
        self._run_stage(tmp_path, smtplib.SMTPAuthenticationError(535, b"auth failed"))

    def test_smtp_error_does_not_write_seen_roles(self, tmp_path):
        self._run_stage(tmp_path, smtplib.SMTPException("connection reset"))

    def test_generic_exception_does_not_write_seen_roles(self, tmp_path):
        self._run_stage(tmp_path, OSError("network unreachable"))

    def test_success_does_write_seen_roles(self, tmp_path):
        """Sanity: on successful SMTP delivery seen_roles IS persisted."""
        jobs = [_make_job("apply")]
        new_role_timestamps = {"abc123": "2026-06-28T10:00:00"}

        with (
            patch("metis.pipeline.render_html", return_value="<html>digest</html>"),
            patch("metis.pipeline.send_digest"),
            patch("metis.pipeline.save_seen_roles") as mock_save,
            patch("metis.pipeline.save_skipped_roles"),
            patch("metis.pipeline.save_last_run"),
            patch("metis.xlsx.write_to_tracker"),
        ):
            from metis.pipeline import _stage_deliver
            _stage_deliver(jobs, n_filtered=0, new_role_timestamps=new_role_timestamps)

        mock_save.assert_called_once_with(new_role_timestamps)

    def test_dry_run_does_not_write_delivery_state(self):
        """Dry-run renders the digest preview but must not send or persist state."""
        jobs = [_make_job("apply")]
        new_role_timestamps = {"abc123": "2026-06-28T10:00:00"}

        with (
            patch("metis.pipeline.render_html", return_value="<html>digest</html>"),
            patch("metis.pipeline.send_digest") as mock_send,
            patch("metis.pipeline.save_seen_roles") as mock_seen,
            patch("metis.pipeline.save_skipped_roles") as mock_skipped,
            patch("metis.pipeline.save_last_run") as mock_last_run,
            patch("metis.xlsx.write_to_tracker") as mock_tracker,
        ):
            from metis.pipeline import _stage_deliver
            _stage_deliver(
                jobs,
                n_filtered=0,
                new_role_timestamps=new_role_timestamps,
                dry_run=True,
            )

        mock_send.assert_not_called()
        mock_seen.assert_not_called()
        mock_skipped.assert_not_called()
        mock_last_run.assert_not_called()
        mock_tracker.assert_not_called()

    def test_dry_run_cap_does_not_write_role_queue(self):
        """Dry-run should not clear or overwrite role_queue.json."""
        jobs = [_make_job("apply")]

        with patch("metis.pipeline.save_role_queue") as mock_queue:
            from metis.pipeline import _stage_cap
            _stage_cap(jobs, score_all=False, client=MagicMock(), dry_run=True)

        mock_queue.assert_not_called()


# ---------------------------------------------------------------------------
# T-08  track → xlsx: confirmation email flips row to Applied
# ---------------------------------------------------------------------------

class TestTrackXlsxIntegration:
    """When track.py classifies an email as 'confirmation', the matching
    tracker row must flip action_taken → 'Applied' and record date_applied.

    Tests the update_confirmation() path through find_tracker_row().
    No IMAP or network calls — all input is injected directly.
    """

    def _reload_ws(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb.active

    def test_confirmation_sets_applied(self, tmp_path):
        tracker = _make_tracker(tmp_path, [
            {"role_title": "Staff Product Manager", "company": "Stripe",
             "action_taken": "Not Applied"},
        ])

        from metis.track_write import find_tracker_row, update_confirmation
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        row_idx = find_tracker_row(ws, "Stripe", "Staff Product Manager")
        assert row_idx is not None, "find_tracker_row should match the Stripe row"
        update_confirmation(ws, row_idx, "2026-06-28")
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 6).value == "Applied",     "action_taken should be 'Applied'"
        assert ws.cell(2, 7).value == "2026-06-28",  "date_applied should be set"
        assert ws.cell(2, 8).value == "Pending",     "application_status should be 'Pending'"

    def test_rejection_sets_rejected(self, tmp_path):
        tracker = _make_tracker(tmp_path, [
            {"role_title": "Staff PM", "company": "Anthropic",
             "action_taken": "Applied", "application_status": "Pending"},
        ])

        from metis.track_write import find_tracker_row, update_rejection
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        row_idx = find_tracker_row(ws, "Anthropic", "Staff PM")
        assert row_idx is not None
        update_rejection(ws, row_idx)
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 8).value == "Rejected", "application_status should be 'Rejected'"

    def test_fuzzy_match_tolerates_ats_title_drift(self, tmp_path):
        """ATS titles often differ from LinkedIn titles — fuzzy match must bridge the gap."""
        tracker = _make_tracker(tmp_path, [
            {"role_title": "Staff Product Manager, Platform", "company": "Stripe Inc.",
             "action_taken": "Not Applied"},
        ])

        from metis.track_write import find_tracker_row
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        # ATS subject might say "Staff PM - Stripe" — shorter, variant company name
        row_idx = find_tracker_row(ws, "Stripe", "Staff Product Manager")

        assert row_idx is not None, (
            "fuzzy match should bridge 'Stripe Inc.' → 'Stripe' and "
            "'Staff Product Manager, Platform' → 'Staff Product Manager'"
        )

    def test_no_match_returns_none(self, tmp_path):
        """A confirmation for a company not in the tracker returns None without error."""
        tracker = _make_tracker(tmp_path, [
            {"role_title": "Staff PM", "company": "Stripe", "action_taken": "Not Applied"},
        ])

        from metis.track_write import find_tracker_row
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        row_idx = find_tracker_row(ws, "Anthropic", "Staff PM")

        assert row_idx is None, "Anthropic is not in the tracker — should return None"

    def test_confirmation_preserves_existing_date_applied(self, tmp_path):
        """A second confirmation email must not overwrite the original application date."""
        tracker = _make_tracker(tmp_path, [
            {"role_title": "Staff PM", "company": "Stripe",
             "action_taken": "Applied", "date_applied": "2026-06-01"},
        ])

        from metis.track_write import find_tracker_row, update_confirmation
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        row_idx = find_tracker_row(ws, "Stripe", "Staff PM")
        assert row_idx is not None
        update_confirmation(ws, row_idx, "2026-06-28")   # later date — should be ignored
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 7).value == "2026-06-01", \
            "existing date_applied must not be overwritten by a follow-up confirmation"

    def test_company_only_confirmation_can_match_existing_row(self, tmp_path):
        tracker = _make_tracker(tmp_path, [
            {"role_title": "Staff Product Manager, AI Platform", "company": "Weights & Biases",
             "action_taken": "Not Applied"},
        ])

        from metis.track_write import find_tracker_row, update_confirmation
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        row_idx = find_tracker_row(ws, "Weights & Biases", None)
        assert row_idx is not None
        update_confirmation(ws, row_idx, "2026-06-29")
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 6).value == "Applied"
        assert ws.cell(2, 7).value == "2026-06-29"

    def test_external_backfill_uses_clean_placeholder_title(self, tmp_path):
        tracker = _make_tracker(tmp_path, [])

        from metis.track_write import create_backfill_row
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        create_backfill_row(ws, {
            "company": "Docker",
            "role": None,
            "date": "2026-06-29",
        })
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 2).value == "External application"
        assert ws.cell(2, 3).value == "Docker"
        assert ws.cell(2, 5).value == "External"

    def test_external_backfill_never_uses_company_as_role_title(self, tmp_path):
        tracker = _make_tracker(tmp_path, [])

        from metis.track_write import create_backfill_row
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        create_backfill_row(ws, {
            "company": "Anthropic",
            "role": "Anthropic",
            "date": "2026-06-29",
        })
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 2).value == "External application"
        assert ws.cell(2, 3).value == "Anthropic"

    def test_external_backfill_alignment_matches_tracker_convention(self, tmp_path):
        tracker = _make_tracker(tmp_path, [])

        from metis.track_write import create_backfill_row
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        create_backfill_row(ws, {
            "company": "Docker",
            "role": None,
            "date": "2026-06-29",
        })
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        centered_cols = (1, 4, 5, 6, 7, 8)
        left_cols = (2, 3, 9)
        assert all(ws.cell(2, col).alignment.horizontal == "center" for col in centered_cols)
        assert all(ws.cell(2, col).alignment.horizontal == "left" for col in left_cols)

    def test_external_backfill_uses_parsed_url_as_hyperlink(self, tmp_path):
        tracker = _make_tracker(tmp_path, [])

        from metis.track_write import create_backfill_row
        wb = openpyxl.load_workbook(tracker)
        ws = wb.active
        create_backfill_row(ws, {
            "company": "Docker",
            "role": "Staff Product Manager",
            "url": "https://www.linkedin.com/jobs/view/123/",
            "date": "2026-06-29",
        })
        wb.save(tracker)

        ws = self._reload_ws(tracker)
        assert ws.cell(2, 2).hyperlink.target == "https://www.linkedin.com/jobs/view/123/"


def test_parse_email_rejects_reply_subject_noise():
    from metis.track_parse import parse_email

    parsed = parse_email({
        "subject": "Re: Let's connect",
        "body": "Thanks for your message.",
        "sender": "Jane Person <jane@gmail.com>",
        "date": "2026-06-29",
    })

    assert parsed["company"] is None
    assert parsed["role"] is None


def test_parse_email_drops_company_as_role_title():
    from metis.track_parse import parse_email

    parsed = parse_email({
        "subject": "Thanks for applying to Weights & Biases",
        "body": "Thanks for applying to Weights & Biases.",
        "sender": "Weights & Biases Talent <careers@wandb.com>",
        "date": "2026-06-29",
    })

    assert parsed["company"] == "Weights & Biases"
    assert parsed["role"] is None


def test_parse_email_preserves_url_metadata():
    from metis.track_parse import parse_email

    parsed = parse_email({
        "subject": "Thanks for applying to Docker",
        "body": "Thanks for applying to Docker.",
        "sender": "Docker Talent <careers@docker.com>",
        "date": "2026-06-29",
        "url": "https://www.linkedin.com/jobs/view/123/",
    })

    assert parsed["url"] == "https://www.linkedin.com/jobs/view/123/"


def test_track_decode_header_value_import_path():
    from metis.track import _decode_header_value

    assert _decode_header_value("Metis Progress Report") == "Metis Progress Report"
