"""scorerole/track_write.py — tracker fuzzy match and xlsx write operations.

Owns: find_tracker_row(), update_*(), create_*(), _parse_digest_html().
No IMAP, no classification logic lives here.
"""
from __future__ import annotations

import json
import logging
import re

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fuzzy matching against tracker rows
# ---------------------------------------------------------------------------

def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _similarity(a: str, b: str) -> float:
    from difflib import SequenceMatcher
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


def find_tracker_row(ws, company: str, role: str | None) -> int | None:
    """Return 1-based row index of the best-matching row, or None.

    Matching strategy:
      1. Company must match above COMPANY_THRESHOLD (fuzzy)
      2. If role is provided, title must match above ROLE_THRESHOLD (fuzzy)
      3. Return the highest-scoring match
    """
    COMPANY_THRESHOLD = 0.70
    ROLE_THRESHOLD    = 0.55

    best_row   = None
    best_score = 0.0

    for row_idx in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row_idx, 3).value or "")
        row_title   = str(ws.cell(row_idx, 2).value or "")

        company_score = _similarity(company, row_company)
        if company_score < COMPANY_THRESHOLD:
            continue

        if role:
            role_score = _similarity(role, row_title)
            if role_score < ROLE_THRESHOLD:
                continue
            combined = (company_score + role_score) / 2
        else:
            combined = company_score

        if combined > best_score:
            best_score = combined
            best_row   = row_idx

    return best_row


# ---------------------------------------------------------------------------
# Tracker updates
# ---------------------------------------------------------------------------

def _apply_status_fill(ws, row_idx: int, col: int, value: str) -> None:
    from openpyxl.styles import PatternFill
    _STATUS_FILL = {
        "Applied":          "C6EFCE",
        "Not Applied":      "D9D9D9",
        "Pending":          "FFEB9C",
        "Proceeding":       "C6EFCE",
        "Rejected":         "FFC7CE",
        "Skipped":          "D9D9D9",
        "Recruiter Screen": "BDD7EE",
    }
    color = _STATUS_FILL.get(value)
    if color:
        ws.cell(row_idx, col).fill = PatternFill(fill_type="solid", fgColor=color)


def update_confirmation(ws, row_idx: int, date_applied: str) -> None:
    """Flip action_taken → Applied and set date_applied + application_status.

    Preserves an existing date_applied — the earliest confirmation date is
    canonical and should not be overwritten by a follow-up email.
    """
    ws.cell(row_idx, 6).value = "Applied"
    if not ws.cell(row_idx, 7).value:
        ws.cell(row_idx, 7).value = date_applied
    ws.cell(row_idx, 8).value = "Pending"
    _apply_status_fill(ws, row_idx, 6, "Applied")
    _apply_status_fill(ws, row_idx, 8, "Pending")


def update_rejection(ws, row_idx: int) -> None:
    """Set application_status → Rejected."""
    ws.cell(row_idx, 8).value = "Rejected"
    _apply_status_fill(ws, row_idx, 8, "Rejected")


def update_recruiter_screen(ws, row_idx: int) -> None:
    """Set action_taken → Applied (if not already) and application_status → Recruiter Screen."""
    if ws.cell(row_idx, 6).value != "Applied":
        ws.cell(row_idx, 6).value = "Applied"
        _apply_status_fill(ws, row_idx, 6, "Applied")
    ws.cell(row_idx, 8).value = "Recruiter Screen"
    _apply_status_fill(ws, row_idx, 8, "Recruiter Screen")


def _write_row_from_email(ws, parsed: dict, suggestion_status: str,
                          date_suggested: str | None = None,
                          match_score: float | None = None,
                          url: str = "") -> None:
    """Shared helper: append one row to ws from parsed email data."""
    from openpyxl.styles import Alignment
    from .xlsx import _set_hyperlink

    next_row = ws.max_row + 1
    values = [
        date_suggested or parsed["date"],
        parsed.get("role") or "",
        parsed.get("company") or "",
        match_score,
        suggestion_status,
        "Applied",
        parsed["date"],
        "Pending",
        None,
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(next_row, col_idx, value).alignment = Alignment(vertical="top")

    if url:
        _set_hyperlink(ws.cell(next_row, 2), url, values[1])

    ws.cell(next_row, 4).number_format = "0%"
    for col, val in [(5, suggestion_status), (6, "Applied"), (8, "Pending")]:
        _apply_status_fill(ws, next_row, col, val)


def create_skipped_row(ws, parsed: dict, skipped_meta: dict) -> None:
    """Create a new row for a skipped role the user applied to anyway."""
    score = skipped_meta.get("match_score")
    _write_row_from_email(
        ws, parsed,
        suggestion_status="Skipped",
        date_suggested=skipped_meta.get("date_suggested"),
        match_score=(score / 100.0) if score else None,
        url=skipped_meta.get("url", ""),
    )
    next_row = ws.max_row
    if skipped_meta.get("role_title"):
        ws.cell(next_row, 2).value = skipped_meta["role_title"]
    if skipped_meta.get("company"):
        ws.cell(next_row, 3).value = skipped_meta["company"]


def create_backfill_row(ws, parsed: dict) -> None:
    """Create a row for a confirmed application with no prior tracker entry.

    Used for roles applied to before the tracker existed, or applied to outside
    the scorerole digest. suggestion_status='Pre-tracker' marks these as backfills.
    """
    _write_row_from_email(ws, parsed, suggestion_status="Pre-tracker")


# ---------------------------------------------------------------------------
# Digest HTML parsing
# ---------------------------------------------------------------------------

def _parse_digest_html(html: str, email_date: str) -> list[dict]:
    """Extract Apply/Consider job rows from a rendered digest HTML email.

    Returns a list of job dicts compatible with write_to_tracker():
    {'title', 'company', 'url', 'eval': {'score', 'verdict'}}
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")

    # Prefer embedded JSON data island (added to future digests)
    tag = soup.find("script", {"type": "application/json", "id": "scorerole-data"})
    if tag and tag.string:
        try:
            payload = json.loads(tag.string)
            jobs = []
            for j in payload.get("jobs", []):
                verdict = j.get("verdict", "").lower()
                if verdict not in ("apply", "consider"):
                    continue
                jobs.append({
                    "title":   j.get("title", ""),
                    "company": j.get("company", ""),
                    "url":     j.get("postingUrl", ""),
                    "eval":    {"score": j.get("score", 0), "verdict": verdict},
                })
            return jobs
        except Exception:
            pass  # fall through to HTML parsing

    # HTML fallback: parse job cards by "View posting" links
    def _nstyle(tag) -> str:
        return re.sub(r"\s+", "", (tag.get("style") or "").lower())

    _MUTED = ("72716d", "888780", "726f6a", "6b6b6b", "757575", "666666")

    jobs = []
    for anchor in soup.find_all("a"):
        if "View posting" not in anchor.get_text():
            continue
        url = anchor.get("href", "")
        card = anchor.find_parent(
            "table",
            style=lambda s: s and "border-radius:8px" in s.replace(" ", "")
        )
        if not card:
            continue

        title_td = card.find(
            lambda t: t.name in ("td", "div", "p", "span")
            and "15px" in _nstyle(t)
            and any(w in _nstyle(t) for w in ("font-weight:500", "font-weight:600",
                                               "font-weight:bold"))
        )
        title = title_td.get_text(strip=True) if title_td else ""

        co_tag = card.find(
            lambda t: t.name in ("p", "div", "td", "span")
            and any(c in _nstyle(t) for c in _MUTED)
            and t is not title_td
        )
        if not co_tag or not co_tag.get_text(strip=True):
            co_tag = card.find(
                lambda t: t.name in ("p", "div", "td", "span")
                and "·" in t.get_text()
                and t is not title_td
                and (title_td is None or t not in title_td.parents)
                and len(t.get_text(strip=True)) < 120
            )
        company_raw = co_tag.get_text(strip=True) if co_tag else ""
        company = company_raw.split("·")[0].strip()

        score_span = card.find(
            lambda t: t.name == "span"
            and "border-radius:20px" in (t.get("style") or "").replace(" ", "")
            and re.search(r"\d+%", t.get_text())
        )
        score = 0
        if score_span:
            m = re.search(r"(\d+)%", score_span.get_text())
            if m:
                score = int(m.group(1))

        style = (score_span.get("style", "") if score_span else "").lower()
        if "eef2ee" in style or "eaf3de" in style:
            verdict = "apply"
        elif "f4f0e8" in style or "faeeda" in style:
            verdict = "consider"
        else:
            verdict = "consider"

        if title and company:
            jobs.append({
                "title":   title,
                "company": company,
                "url":     url,
                "eval":    {"score": score, "verdict": verdict},
            })

    return jobs
