from __future__ import annotations

import datetime as dt
import json
import os
import re
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, urlparse

from .application_state import data_dir, load_application_state, update_application_state
from .application_profile import application_value, load_application_profile
from .profile import load_profile_yaml
from .state import RUNS_PATH, _role_hash

log = __import__("logging").getLogger(__name__)


_SELECT_ALL = "__metis_select_all__"
_DESELECT_ALL = "__metis_deselect_all__"
_CANCEL = "__metis_cancel__"
_FINAL_STATUSES = {"applied", "applied_confirmed", "rejected", "recruiter_screen"}
_SUCCESS_PATTERNS = (
    re.compile(r"application (?:has been )?submitted", re.I),
    re.compile(r"thanks? for (?:your )?application", re.I),
    re.compile(r"we(?:'|’)ve received your application", re.I),
    re.compile(r"application received", re.I),
)


@dataclass(frozen=True)
class ApplicationCandidate:
    role_key: str
    role: dict[str, Any]
    record_path: Path | None
    resume_path: Path
    tailored: bool
    resume_kind: str = ""
    workflow_status: str = ""


class LinkedInAuthenticationError(RuntimeError):
    """The selected Chrome profile does not have an authenticated LinkedIn session."""


def _has_linkedin_session(context: Any) -> bool:
    """Use LinkedIn's auth cookie instead of unreliable page copy as session evidence."""
    try:
        return any(
            cookie.get("name") == "li_at" and bool(cookie.get("value"))
            for cookie in context.cookies("https://www.linkedin.com")
        )
    except Exception:
        return False


def _probe_linkedin_session(context: Any, page: Any) -> bool:
    """Verify the session functionally when CDP cannot expose a Keychain cookie."""
    if _has_linkedin_session(context):
        return True
    try:
        page.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(1_500)
        path = urlparse(page.url).path.lower()
        return path.startswith("/feed") and not any(
            marker in path for marker in ("/login", "/signup", "/authwall")
        )
    except Exception:
        return False


def _role_key(role: dict[str, Any]) -> str:
    return str(role.get("role_hash") or _role_hash(role.get("title", ""), role.get("company", "")))


def detect_ats(url: str) -> str | None:
    host = (urlparse(url).hostname or "").lower()
    if "greenhouse.io" in host:
        return "greenhouse"
    if "ashbyhq.com" in host:
        return "ashby"
    if host == "lever.co" or host.endswith(".lever.co"):
        return "lever"
    return None


# Recruiter aggregators: hide the employer, so search never resolves to the employer ATS.
# Listed in the spec as always-blocked. Also added to _SEARCH_BLOCKED_HOSTS so their URLs
# are ignored when parsing Google/DDG result pages.
_AGGREGATOR_DOMAINS = {
    "jobgether.com", "theladders.com", "ladders.com",
    "harnham.com", "hired.com",
}
# Company-name substrings that identify aggregators (used when the role URL is LinkedIn-hosted)
_AGGREGATOR_COMPANY_SUBSTRINGS = {
    "jobgether", "the ladders", "ladders.com", "harnham", "hired.com",
}

_SEARCH_BLOCKED_HOSTS = {
    "linkedin.com", "www.linkedin.com", "indeed.com", "www.indeed.com",
    "glassdoor.com", "www.glassdoor.com", "ziprecruiter.com", "www.ziprecruiter.com",
    "builtin.com", "www.builtin.com", "simplify.jobs", "www.simplify.jobs",
    *_AGGREGATOR_DOMAINS,
}


def _is_external_job_url(url: str) -> bool:
    """Allow employer/ATS pages while rejecting search, social, and aggregator links."""
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    if parsed.scheme not in {"http", "https"} or not host:
        return False
    if host in _SEARCH_BLOCKED_HOSTS or any(host.endswith(f".{item}") for item in _SEARCH_BLOCKED_HOSTS):
        return False
    return "google." not in host


def _start_url(role: dict[str, Any]) -> str:
    url = str(role.get("apply_url") or role.get("url") or "")
    if detect_ats(url) == "ashby":
        parsed = urlparse(url)
        path = parsed.path.rstrip("/")
        if not path.endswith("/application"):
            path += "/application"
        return parsed._replace(path=path).geturl()
    return url


def _fallback_resume() -> Path | None:
    raw = application_value("default_resume")
    if not raw:
        return None
    path = Path(raw).expanduser()
    return path if path.is_file() and path.suffix.lower() == ".docx" else None


def _is_application_ready(role: dict[str, Any]) -> bool:
    return detect_ats(_start_url(role)) in {"greenhouse", "lever", "ashby"}


def _tracker_applied_roles(root: Path) -> set[tuple[str, str]]:
    """Return normalized tracker roles already marked Applied."""
    path = Path(os.getenv("TRACKER_PATH", str(root / "applications.xlsx")))
    if not path.exists():
        return set()
    try:
        import openpyxl

        sheet = openpyxl.load_workbook(path, read_only=True, data_only=True).active
        return {
            (_normalize_match(str(row[2] or "")), _normalize_match(str(row[1] or "")))
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if len(row) >= 6 and str(row[5] or "").strip().lower() == "applied"
        }
    except Exception:
        return set()


def _tracker_pending_roles(root: Path) -> list[dict[str, Any]]:
    """Load tracker recommendations so older unapplied roles remain selectable."""
    path = Path(os.getenv("TRACKER_PATH", str(root / "applications.xlsx")))
    if not path.exists():
        return []
    try:
        import openpyxl

        sheet = openpyxl.load_workbook(path, read_only=False, data_only=False).active
        roles = []
        for row_index in range(2, sheet.max_row + 1):
            action = str(sheet.cell(row_index, 6).value or "").strip().lower()
            if action == "applied":
                continue
            title_cell = sheet.cell(row_index, 2)
            title = str(title_cell.value or "").strip()
            company = str(sheet.cell(row_index, 3).value or "").strip()
            raw_score = sheet.cell(row_index, 4).value
            score = float(raw_score or 0)
            if score <= 1:
                score *= 100
            url = title_cell.hyperlink.target if title_cell.hyperlink else ""
            if not title or not company or score < 75:
                continue
            roles.append({
                "role_hash": _role_hash(title, company), "title": title, "company": company,
                "url": url, "ts": str(sheet.cell(row_index, 1).value or ""),
                "eval": {"score": round(score), "verdict": "apply"},
            })
        return roles
    except Exception:
        return []


def _normalize_match(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def load_application_candidates(
    root: Path | None = None, *, include_applied: bool = False, resume_override: Path | None = None,
) -> list[ApplicationCandidate]:
    explicit_root = root is not None
    root = root or data_dir()
    runs_path = root / "runs.jsonl" if explicit_root else RUNS_PATH
    state = load_application_state(root)
    tracker_applied = _tracker_applied_roles(root)
    fallback = resume_override or _fallback_resume()
    candidates: dict[str, ApplicationCandidate] = {}
    for record_path in sorted((root / "resume_tailor").glob("*/*/tailoring_record.json")):
        try:
            record = json.loads(record_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        role = record.get("role") or {}
        role_key = _role_key(role)
        if (_normalize_match(str(role.get("company") or "")), _normalize_match(str(role.get("title") or ""))) in tracker_applied:
            continue
        saved = state.get(role_key) or {}
        saved_application_url = str(saved.get("application_url") or "")
        if _is_external_job_url(saved_application_url):
            role = {**role, "apply_url": saved_application_url}
        application_company = str(saved.get("application_company") or "")
        if application_company and application_company != str(role.get("company") or ""):
            role = {**role, "source_company": role.get("company"), "company": application_company}
        if not _start_url(role):
            continue
        if not include_applied and (state.get(role_key) or {}).get("status") in _FINAL_STATUSES:
            continue
        clean = Path(str((record.get("artifacts") or {}).get("clean_resume") or "")).expanduser()
        tailored = clean.is_file() and clean.suffix.lower() == ".docx"
        resume = clean if tailored else fallback
        if resume is None:
            continue
        fit = str((((record.get("plan") or {}).get("employer_lens") or {}).get("fit_assessment") or "")).lower()
        if fit.startswith("not_recommended"):
            continue
        if str(role.get("apply_mode") or "").lower() == "closed":
            continue
        candidate = ApplicationCandidate(
            role_key, role, record_path, resume, tailored,
            "tailored" if tailored else "default", str(saved.get("status") or "not started"),
        )
        previous = candidates.get(role_key)
        if previous is None or previous.record_path is None or record_path.stat().st_mtime >= previous.record_path.stat().st_mtime:
            candidates[role_key] = candidate
    if runs_path.exists():
        recent_roles: list[dict[str, Any]] = []
        for line in runs_path.read_text(encoding="utf-8").splitlines():
            try:
                role = json.loads(line)
            except (TypeError, json.JSONDecodeError):
                continue
            if (role.get("eval") or {}).get("verdict") not in {"apply", "consider"} or not _start_url(role):
                continue
            if str(role.get("apply_mode") or "").lower() == "closed":
                continue
            recent_roles.append(role)
        for role in recent_roles:
            role_key = _role_key(role)
            if (_normalize_match(str(role.get("company") or "")), _normalize_match(str(role.get("title") or ""))) in tracker_applied:
                continue
            if role_key in candidates:
                existing = candidates[role_key]
                # Tailoring owns the resume artifact; the evaluated-role trace
                # owns current score and evaluation-date metadata.
                candidates[role_key] = replace(existing, role={**existing.role, **role})
                continue
            if fallback is None:
                continue
            if not include_applied and (state.get(role_key) or {}).get("status") in _FINAL_STATUSES:
                continue
            saved = state.get(role_key) or {}
            saved_application_url = str(saved.get("application_url") or "")
            if _is_external_job_url(saved_application_url):
                role = {**role, "apply_url": saved_application_url}
            application_company = str(saved.get("application_company") or "")
            if application_company and application_company != str(role.get("company") or ""):
                role = {**role, "source_company": role.get("company"), "company": application_company}
            candidates[role_key] = ApplicationCandidate(
                role_key, role, None, fallback, False, "default", str(saved.get("status") or "not started"),
            )
    if fallback:
        for role in _tracker_pending_roles(root):
            role_key = _role_key(role)
            if role_key in candidates:
                continue
            saved = state.get(role_key) or {}
            if not include_applied and saved.get("status") in _FINAL_STATUSES:
                continue
            saved_application_url = str(saved.get("application_url") or "")
            if _is_external_job_url(saved_application_url):
                role = {**role, "apply_url": saved_application_url}
            candidates[role_key] = ApplicationCandidate(
                role_key, role, None, fallback, False, "default", str(saved.get("status") or "not started"),
            )
    return sorted(
        candidates.values(),
        key=lambda item: (
            int((item.role.get("eval") or {}).get("score") or 0),
            item.record_path.stat().st_mtime if item.record_path else 0,
        ),
        reverse=True,
    )


def _label(candidate: ApplicationCandidate) -> str:
    role = candidate.role
    score = int((role.get("eval") or {}).get("score") or 0)
    company = " ".join(str(role.get("company") or "").split())[:22]
    title = " ".join(str(role.get("title") or "").split())[:48]
    if candidate.record_path:
        day = dt.datetime.fromtimestamp(candidate.record_path.stat().st_mtime).strftime("%b %d")
    else:
        day = str(role.get("ts") or "")[:10] or "recent"
    suffix = "tailored" if candidate.tailored else "default"
    return f"{score:>3}% | {company:<22} | {title:<48} | {day} · {suffix}"


_MATCH_STOP_WORDS = {
    "a", "an", "and", "at", "for", "in", "of", "on", "product", "manager",
    "senior", "staff", "principal", "lead", "technical", "the",
}


def _search_queries(role: dict[str, Any]) -> list[str]:
    title = " ".join(str(role.get("title") or "").split())
    company = " ".join(str(role.get("company") or "").split())
    exact = f'"{title}" "{company}" jobs apply'
    relaxed_title = " ".join(
        word for word in title.split()
        if re.sub(r"[^a-z]", "", word.lower()) not in {"senior", "staff", "principal", "lead"}
    )
    relaxed = f'"{relaxed_title or title}" "{company}" careers'
    return [exact, relaxed] if relaxed != exact else [exact]


def _search_result_url(href: str) -> str:
    if not href:
        return ""
    if href.startswith("/url?"):
        return (parse_qs(urlparse(href).query).get("q") or [""])[0]
    parsed = urlparse(href)
    if parsed.hostname and "google." in parsed.hostname and parsed.path == "/url":
        return (parse_qs(parsed.query).get("q") or [""])[0]
    # DDG HTML wraps results as protocol-relative redirect: //duckduckgo.com/l/?uddg=<encoded-url>
    if (parsed.hostname or "").endswith("duckduckgo.com") and "uddg" in (parsed.query or ""):
        return parse_qs(parsed.query).get("uddg", [""])[0]
    return href if parsed.scheme in {"http", "https"} else ""


def _role_tokens(value: str) -> set[str]:
    return {
        token for token in re.findall(r"[a-z0-9]+", value.lower())
        if len(token) > 2 and token not in _MATCH_STOP_WORDS
    }


def _page_matches_role(page: Any, candidate: ApplicationCandidate) -> bool:
    try:
        body = page.locator("body").inner_text(timeout=5_000).lower()
    except Exception:
        return False
    title_tokens = _role_tokens(str(candidate.role.get("title") or ""))
    company_tokens = _role_tokens(str(candidate.role.get("company") or ""))
    title_ratio = sum(token in body for token in title_tokens) / max(1, len(title_tokens))
    company_match = not company_tokens or any(token in body for token in company_tokens)
    return title_ratio >= 0.65 and company_match and _is_external_job_url(page.url)


def _search_result_urls_from_page(search_page: Any) -> list[str]:
    urls: list[str] = []
    anchors = search_page.locator("a[href]")
    for index in range(min(anchors.count(), 80)):
        href = _search_result_url(anchors.nth(index).get_attribute("href") or "")
        if href and _is_external_job_url(href) and href not in urls:
            urls.append(href)
    return urls[:10]


_SEARCH_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def _http_search_urls(query: str) -> list[str]:
    """Search via HTTP — no browser, no CAPTCHA, no profile conflicts."""
    import httpx
    engines = [
        f"https://html.duckduckgo.com/html/?q={quote_plus(query)}",
        f"https://www.bing.com/search?q={quote_plus(query)}&count=10",
        f"https://www.google.com/search?q={quote_plus(query)}&num=10",
    ]
    for search_url in engines:
        try:
            r = httpx.get(search_url, headers=_SEARCH_HEADERS, timeout=20, follow_redirects=True)
            if r.status_code != 200:
                log.debug("apply: search %s returned %d for %r", search_url.split("?")[0], r.status_code, query)
                continue
            urls: list[str] = []
            for m in re.finditer(r'href=["\']([^"\']+)["\']', r.text):
                resolved = _search_result_url(m.group(1))
                if resolved and _is_external_job_url(resolved) and resolved not in urls:
                    urls.append(resolved)
            if urls:
                log.debug("apply: search %s found %d candidate URLs for %r", search_url.split("?")[0], len(urls), query)
                return urls[:10]
            log.debug("apply: search %s returned 0 usable URLs for %r", search_url.split("?")[0], query)
        except Exception as exc:
            log.debug("apply: search %s failed: %s", search_url.split("?")[0], exc)
            continue
    return []


def _page_content_matches_role(html: str, candidate: ApplicationCandidate) -> bool:
    """HTTP version of _page_matches_role — checks raw HTML body for title/company tokens."""
    try:
        from bs4 import BeautifulSoup
        body = BeautifulSoup(html, "html.parser").get_text(" ").lower()
    except Exception:
        body = html.lower()
    title_tokens = _role_tokens(str(candidate.role.get("title") or ""))
    company_tokens = _role_tokens(str(candidate.role.get("company") or ""))
    title_ratio = sum(token in body for token in title_tokens) / max(1, len(title_tokens))
    company_match = not company_tokens or any(token in body for token in company_tokens)
    return title_ratio >= 0.65 and company_match


def _resolve_application_url(candidate: ApplicationCandidate) -> str | None:
    """Resolve ATS URL via HTTP search — no browser needed, no profile conflicts."""
    import httpx
    attempted: set[str] = set()
    company = candidate.role.get("company", "")
    title = candidate.role.get("title", "")
    for query in _search_queries(candidate.role)[:2]:
        log.debug("apply: searching ATS URL for %r at %r: %r", title, company, query)
        for url in _http_search_urls(query):
            if url in attempted:
                continue
            attempted.add(url)
            try:
                r = httpx.get(url, headers=_SEARCH_HEADERS, timeout=20, follow_redirects=True)
                host = urlparse(str(r.url)).hostname or ""
                if any(host == agg or host.endswith(f".{agg}") for agg in _AGGREGATOR_DOMAINS):
                    log.debug("apply: skipping aggregator result %s", host)
                    continue
                if r.status_code == 200 and _page_content_matches_role(r.text, candidate):
                    log.debug("apply: ATS URL resolved to %s", r.url)
                    return str(r.url)
                log.debug("apply: page content mismatch for %s (status=%d)", url, r.status_code)
            except Exception as exc:
                log.debug("apply: failed to fetch candidate URL %s: %s", url, exc)
                continue
    log.debug("apply: no ATS URL found for %r at %r", title, company)
    return None


def _write_apply_diagnostic(
    root: Path, candidate: ApplicationCandidate, *, phase: str, error: str, page: Any | None = None,
) -> None:
    """Append privacy-safe browser failure metadata without storing page HTML or form values."""
    record = {
        "ts": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "role_key": candidate.role_key,
        "company": candidate.role.get("company"),
        "title": candidate.role.get("title"),
        "phase": phase,
        "error": error,
        "apply_mode": candidate.role.get("apply_mode"),
    }
    if page is not None:
        try:
            record["url"] = page.url
            record["page_title"] = page.title()[:200]
        except Exception:
            pass
    path = root / "apply_diagnostics.jsonl"
    path.parent.mkdir(mode=0o700, parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")
    path.chmod(0o600)


def _candidate_date(candidate: ApplicationCandidate) -> str:
    for field in ("ts", "date_suggested", "run_date"):
        value = str(candidate.role.get(field) or "")[:10]
        if re.match(r"^\d{4}-\d{2}-\d{2}$", value):
            return value
    if candidate.record_path:
        folder_date = candidate.record_path.parent.parent.name
        if re.match(r"^\d{8}$", folder_date):
            return f"{folder_date[:4]}-{folder_date[4:6]}-{folder_date[6:]}"
        return dt.datetime.fromtimestamp(candidate.record_path.stat().st_mtime).date().isoformat()
    return ""


def _empty_gate_message(
    *, lookback: str | None, include_applied: bool, match_terms: list[str] | None,
) -> str:
    if not RUNS_PATH.exists():
        return "No evaluated roles are available yet. Run `metis` first."
    cutoff = ""
    if lookback:
        from .pipeline import _parse_lookback

        since = _parse_lookback(lookback)
        cutoff = since.date().isoformat() if since else ""
    terms = [term.strip().lower() for term in (match_terms or []) if term.strip()]
    state = load_application_state()
    roles: dict[str, dict[str, Any]] = {}
    for line in RUNS_PATH.read_text(encoding="utf-8").splitlines():
        try:
            role = json.loads(line)
        except (TypeError, json.JSONDecodeError):
            continue
        if (role.get("eval") or {}).get("verdict") not in {"apply", "consider"}:
            continue
        date = str(role.get("ts") or role.get("date_suggested") or "")[:10]
        if cutoff and (not date or date < cutoff):
            continue
        haystack = f"{role.get('company', '')} {role.get('title', '')}".lower()
        if terms and not any(term in haystack for term in terms):
            continue
        roles[_role_key(role)] = role
    unresolved = 0
    final = 0
    for role_key, role in roles.items():
        saved = state.get(role_key) or {}
        if not include_applied and saved.get("status") in _FINAL_STATUSES:
            final += 1
            continue
        if not detect_ats(str(saved.get("application_url") or _start_url(role))):
            unresolved += 1
    return (
        f"{len(roles)} evaluated role(s) matched the window, but none are ready to open.\n"
        f"  {unresolved} have no cached ATS URL yet (web search will run at apply time).\n"
        f"  {final} are excluded by confirmed application outcome.\n"
        "Run `metis apply` interactively to let web search find application URLs."
    )


def select_candidates(
    candidates: list[ApplicationCandidate],
    *,
    preselected: list[ApplicationCandidate] | None = None,
) -> list[ApplicationCandidate]:
    try:
        from InquirerPy import inquirer
        from InquirerPy.base.control import Choice
        from InquirerPy.separator import Separator
        from .theme import INQUIRER_STYLE, console
    except Exception:
        return list(preselected or []) or candidates[:1]
    # Nothing pre-selected unless explicitly given (e.g. via --top N).
    preselected_keys = {item.role_key for item in preselected} if preselected is not None else set()
    console.print()
    hint = (
        f"{len(preselected_keys)} pre-selected" if preselected_keys else f"{len(candidates)} pending"
    )
    console.print(f"[dim]Choose roles to apply to. {hint}, highest match first.[/dim]")
    # "Deselect all" is only shown when items are pre-checked (so there's something to clear).
    # InquirerPy can't relabel choices dynamically, so we include only what's meaningful upfront.
    header_choices: list[Any] = [Choice(_SELECT_ALL, f"Select all {len(candidates)} roles")]
    if preselected_keys:
        header_choices.append(Choice(_DESELECT_ALL, "Deselect all / clear"))
    header_choices.append(Choice(_CANCEL, "Cancel / exit"))
    choices = [
        *header_choices,
        Separator(),
        *[
            Choice(item.role_key, f"{idx:>2}. {_label(item)}", enabled=item.role_key in preselected_keys)
            for idx, item in enumerate(candidates, 1)
        ],
    ]
    selected = inquirer.checkbox(
        message="Applications",
        choices=choices,
        style=INQUIRER_STYLE,
        instruction="↑↓ scrolls · Space toggles · Enter confirms",
        height=min(len(candidates) + len(header_choices) + 2, 20),
        validate=lambda result: bool(result),
        invalid_message="Press Space to select at least one role, or Ctrl-C to cancel.",
    ).execute()
    if _CANCEL in selected:
        raise SystemExit("Cancelled.")
    if _SELECT_ALL in selected:
        return candidates
    # Strip sentinel keys; if real roles remain, use them — user may have toggled
    # "Deselect all" alongside individual picks, which means they want just the picks.
    sentinels = {_SELECT_ALL, _DESELECT_ALL, _CANCEL}
    real_selected = [key for key in selected if key not in sentinels]
    if _DESELECT_ALL in selected and not real_selected:
        raise SystemExit("No roles selected.")
    by_key = {item.role_key: item for item in candidates}
    return [by_key[key] for key in real_selected if key in by_key]


def _looks_submitted(url: str, text: str) -> bool:
    lowered_url = url.lower()
    if any(part in lowered_url for part in ("/confirmation", "/submitted", "application-success")):
        return True
    return any(pattern.search(text or "") for pattern in _SUCCESS_PATTERNS)


def _candidate_values() -> dict[str, str]:
    profile = load_profile_yaml() or {}
    candidate = profile.get("candidate") or {}
    name = str(candidate.get("name") or "").strip()
    first, _, last = name.partition(" ")
    configured = load_application_profile()
    values = {
        "first_name": os.getenv("METIS_FIRST_NAME", first).strip(),
        "last_name": os.getenv("METIS_LAST_NAME", last).strip(),
        "email": os.getenv("GMAIL_ADDRESS", str(candidate.get("email") or "")).strip(),
        "phone": os.getenv("METIS_PHONE", str(candidate.get("phone") or "")).strip(),
        "location": os.getenv("METIS_LOCATION", str(candidate.get("location") or "")).strip(),
        "linkedin": os.getenv("METIS_LINKEDIN_URL", str(candidate.get("linkedin") or "")).strip(),
        "pronouns": os.getenv("METIS_PRONOUNS", "").strip(),
        "current_employer": os.getenv("METIS_CURRENT_EMPLOYER", "").strip(),
        "github": os.getenv("METIS_GITHUB_URL", "").strip(),
        "gender_identity": os.getenv("METIS_GENDER_IDENTITY", "").strip(),
        "hispanic_latino": os.getenv("METIS_HISPANIC_LATINO", "").strip(),
        "race": os.getenv("METIS_RACE", "").strip(),
        "veteran_status": os.getenv("METIS_VETERAN_STATUS", "").strip(),
        "transgender": os.getenv("METIS_TRANSGENDER", "").strip(),
        "disability": os.getenv("METIS_DISABILITY", "").strip(),
        "sexual_orientation": os.getenv("METIS_SEXUAL_ORIENTATION", "").strip(),
        "work_authorized": os.getenv("METIS_WORK_AUTHORIZED", "").strip(),
        "sponsorship_required": os.getenv("METIS_SPONSORSHIP_REQUIRED", "").strip(),
        "willing_to_relocate": os.getenv("METIS_WILLING_TO_RELOCATE", "").strip(),
        "referral_source": os.getenv("METIS_REFERRAL_SOURCE", "LinkedIn").strip(),
    }
    return {key: application_value(key, str(configured.get(key, value) or "")) for key, value in values.items()}


def _fill_visible_form(page: Any, candidate: ApplicationCandidate) -> None:
    """Best-effort deterministic fields; existing browser/autofill values win."""
    page.wait_for_selector("input", state="attached", timeout=15_000)
    values = _candidate_values()
    stable = {
        'input[name*="first" i]': values["first_name"],
        'input[name*="last" i]': values["last_name"],
        "input[type=email]": values["email"],
        "input[type=tel]": values["phone"],
        'input[name*="linkedin" i]': values["linkedin"],
    }
    for selector, value in stable.items():
        if not value:
            continue
        locator = page.locator(selector).first
        if locator.count() and locator.is_visible() and not locator.input_value().strip():
            locator.fill(value)
    resume_upload = page.get_by_label("Resume", exact=True)
    if resume_upload.count():
        resume_upload.first.set_input_files(str(candidate.resume_path))
    if detect_ats(page.url) == "ashby":
        page.wait_for_timeout(1_500)
        _fill_ashby_form(page, candidate, values)
    elif detect_ats(page.url) == "greenhouse":
        _fill_greenhouse_form(page, candidate, values)
    _ensure_resume_attached(page, candidate)
    # Resume uploads can rerender Ashby/Greenhouse forms and clear dependent
    # autocomplete controls. Re-run location and demographics after the final upload settles.
    page.wait_for_timeout(1_000)
    _fill_location_control(page, values["location"])
    _fill_greenhouse_eeoc_by_attr(page, values)
    _llm_fill_remaining(page, values)


def _fill_greenhouse_eeoc_by_attr(page: Any, values: dict[str, str]) -> None:
    """Fallback for Greenhouse EEOC selects whose labels don't match get_by_label.

    Greenhouse EEO forms use predictable id/name attributes (gender_identity, race,
    veteran_status, disability_status, etc.). This function scans all <select> elements
    and tries to match them by attribute keyword, then selects the best matching option.
    It runs after the label-based pass so it only fills still-empty selects.
    """
    gender_choices = (
        ["Female", "Woman"] if re.search(r"woman|female", values["gender_identity"], re.I)
        else ([values["gender_identity"]] if values["gender_identity"] else [])
    )
    race_choices = (
        ["Asian", "Asian (Not Hispanic or Latino)", "Asian or Pacific Islander"]
        if "asian" in values["race"].lower()
        else ([values["race"]] if values["race"] else [])
    )
    veteran_choices = (
        ["No", "Not a protected veteran", "I am not a protected veteran", "None of the above",
         "I am not a protected veteran"]
        if values["veteran_status"].lower() in {"no", "not a protected veteran"}
        else ([values["veteran_status"]] if values["veteran_status"] else [])
    )
    disability_choices = (
        ["No", "No, I don't have a disability", "No, I do not have a disability",
         "I don't have a disability or chronic condition"]
        if values["disability"].lower() == "no"
        else ([values["disability"]] if values["disability"] else [])
    )
    hispanic_choices = (
        ["No", "Not Hispanic or Latino", "No, not of Hispanic, Latino, or Spanish origin"]
        if values["hispanic_latino"].lower() in {"no", "non-hispanic", "not hispanic or latino"}
        else ([values["hispanic_latino"]] if values["hispanic_latino"] else [])
    )
    transgender_choices = (
        ["No", "No, I do not identify as transgender"]
        if values["transgender"].lower() == "no"
        else ([values["transgender"]] if values["transgender"] else [])
    )
    orientation = values.get("sexual_orientation", "")
    orientation_choices = [orientation] if orientation else []

    keyword_choices: list[tuple[list[str], list[str]]] = [
        (["gender"], gender_choices),
        (["race", "ethnicity"], race_choices),
        (["veteran", "military"], veteran_choices),
        (["disability"], disability_choices),
        (["hispanic", "latino"], hispanic_choices),
        (["transgender"], transgender_choices),
        (["orientation"], orientation_choices),
    ]

    try:
        all_selects = page.locator("select").all()
    except Exception:
        return

    for select in all_selects:
        try:
            sel_id = (select.get_attribute("id") or "").lower()
            sel_name = (select.get_attribute("name") or "").lower()
            attr_text = f"{sel_id} {sel_name}"
            for keywords, choices in keyword_choices:
                if not choices or not any(kw in attr_text for kw in keywords):
                    continue
                # Skip if already has a non-empty selection
                current = (select.input_value() or "").strip()
                if current:
                    continue
                options = select.locator("option").all_text_contents()
                for choice in choices:
                    match = next(
                        (opt for opt in options if choice.lower() in opt.lower()),
                        None,
                    )
                    if match:
                        try:
                            select.select_option(label=match, timeout=3_000)
                        except Exception:
                            pass
                        break
        except Exception:
            continue


def _llm_fill_remaining(page: Any, values: dict[str, str]) -> None:
    """Haiku agent pass: fill dropdowns the deterministic pass missed.

    Collects all visible <select> elements that are still empty, asks Haiku which
    option to pick given the candidate profile, and applies the suggestions.
    Only runs if an LLM API key is present — silently skips otherwise.
    """
    try:
        provider = os.getenv("METIS_LLM_PROVIDER", "anthropic")
        from .llm import create_llm_client, normalize_provider, resolve_stage_models
        provider_id = normalize_provider(provider)
        key_env = "OPENAI_API_KEY" if provider_id == "openai" else "ANTHROPIC_API_KEY"
        api_key = os.getenv(key_env, "")
        if not api_key:
            return
        client = create_llm_client(provider=provider_id, api_key=api_key)
        model = resolve_stage_models(provider_id)["extract_model"]
    except Exception:
        return

    # Collect empty visible selects with their labels and options
    fields: list[dict[str, Any]] = []
    try:
        for select in page.locator("select").all():
            try:
                if not select.is_visible():
                    continue
                if (select.input_value() or "").strip():
                    continue  # already filled
                label_text = ""
                sel_id = select.get_attribute("id") or ""
                if sel_id:
                    label_el = page.locator(f'label[for="{sel_id}"]')
                    if label_el.count():
                        label_text = label_el.first.inner_text().strip()
                if not label_text:
                    # Try walking up to find a label-like ancestor text
                    try:
                        label_text = select.evaluate(
                            "el => { let p = el.parentElement; for (let i=0; i<4; i++) {"
                            " if (!p) break; let l = p.querySelector('label,p,span,div.label');"
                            " if (l && l.innerText.trim()) return l.innerText.trim(); p=p.parentElement; }"
                            " return ''; }"
                        )
                    except Exception:
                        pass
                if not label_text:
                    continue
                opts = [o for o in select.locator("option").all_text_contents() if o.strip() and o.strip().lower() not in {"select...", "select", ""}]
                if not opts:
                    continue
                fields.append({"id": sel_id or label_text[:30], "label": label_text, "options": opts})
            except Exception:
                continue
    except Exception:
        return

    if not fields:
        return

    profile_lines = [
        f"Name: {values.get('first_name', '')} {values.get('last_name', '')}",
        f"Gender identity: {values.get('gender_identity', '')}",
        f"Pronouns: {values.get('pronouns', '')}",
        f"Race/ethnicity: {values.get('race', '')}",
        f"Hispanic/Latino: {values.get('hispanic_latino', '')}",
        f"Transgender: {values.get('transgender', '')}",
        f"Sexual orientation: {values.get('sexual_orientation', '')}",
        f"Veteran status: {values.get('veteran_status', '')}",
        f"Disability: {values.get('disability', '')}",
        f"Work authorized in US: {values.get('work_authorized', '')}",
        f"Requires sponsorship: {values.get('sponsorship_required', '')}",
        f"Willing to relocate: {values.get('willing_to_relocate', '')}",
    ]
    profile_text = "\n".join(line for line in profile_lines if line.split(": ", 1)[1])

    prompt = (
        "A job applicant is filling an online application form. "
        "For each dropdown field below, choose the best matching option from the available choices "
        "based on the candidate profile. Only answer fields you can determine with confidence.\n\n"
        f"Candidate profile:\n{profile_text}\n\n"
        f"Form fields:\n{json.dumps(fields, indent=2)}\n\n"
        "Return a JSON object mapping field 'id' to the exact option text to select. "
        "Example: {\"field_id\": \"Female\"}. Return {} if nothing can be determined."
    )

    try:
        from .llm import complete_text
        resp = complete_text(
            client, model=model, system="Return only valid JSON.", user=prompt,
            max_tokens=400, json_mode=True,
        )
        suggestions: dict[str, str] = json.loads(resp.text)
    except Exception as exc:
        log.debug("apply: LLM fill pass failed (%s)", exc)
        return

    for field in fields:
        fid = field["id"]
        chosen = suggestions.get(fid, "")
        if not chosen:
            continue
        try:
            sel = page.locator(f'select[id="{fid}"]') if fid else page.locator("select").filter(has_text=field["label"])
            if sel.count() and not (sel.first.input_value() or "").strip():
                opts = field["options"]
                match = next((o for o in opts if chosen.lower() in o.lower()), None)
                if match:
                    sel.first.select_option(label=match, timeout=3_000)
        except Exception:
            continue


def _fill_empty(locator: Any, value: str) -> None:
    if not value or not locator.count():
        return
    field = locator.first
    if field.is_visible() and not field.input_value().strip():
        field.fill(value, timeout=3_000)


def _fill_ashby_form(page: Any, candidate: ApplicationCandidate, values: dict[str, str]) -> None:
    full_name = " ".join(part for part in (values["first_name"], values["last_name"]) if part).strip()
    fields = {
        r"^(?:Name|Legal Name|Legal First and Last Name|Full Name)\b": full_name,
        r"^Preferred (?:First )?Name\b": values["first_name"],
        r"^Preferred Last Name\b": values["last_name"],
        r"^Preferred Pronouns?\b": values["pronouns"],
        r"^Email\b": values["email"],
        r"^Phone Number\b": values["phone"],
        r"^LinkedIn(?: URL| Profile| Profile Link)?\b": values["linkedin"],
        r"^Current (?:or Most Recent Employer|Company)\b": values["current_employer"],
        r"^GitHub(?: URL)?\b": values["github"],
        r"^How did you hear about (?:us|this opportunity)\b": values["referral_source"],
    }
    for label, value in fields.items():
        _fill_empty(page.get_by_label(re.compile(label, re.I)), value)

    locations = [
        page.get_by_label(re.compile(r"^(?:Current )?Location(?: \(City\))?\b", re.I)),
    ]
    location = next((item for item in locations if item.count()), locations[0])
    if values["location"] and location.count() and location.first.is_visible() and not location.first.input_value().strip():
        _choose_location(page, location.first, values["location"])

    # These fixed answers were explicitly supplied by the candidate. Do not
    # infer them from scoring/profile prose.
    if values["work_authorized"].lower() in {"yes", "true", "1"}:
        _choose_yes_no(page, re.compile(r"(?:legally )?authorized to work", re.I), yes=True)
    no_sponsorship = page.get_by_label(re.compile(r"No, I do not require sponsorship", re.I))
    if values["sponsorship_required"].lower() in {"no", "false", "0"} and no_sponsorship.count():
        try:
            no_sponsorship.first.check(timeout=3_000)
        except Exception:
            pass
    relocate = page.get_by_label(re.compile(r"not based in this location but willing to relocate", re.I))
    if values["willing_to_relocate"].lower() in {"yes", "true", "1"} and relocate.count():
        try:
            relocate.first.check(timeout=3_000)
        except Exception:
            pass
    _choose_custom_select(page, re.compile(r"^(?:Preferred )?Pronouns?\b", re.I), values["pronouns"])
    _choose_radio_group(page, re.compile(r"transgender", re.I), re.compile(r"^No$", re.I) if values["transgender"].lower() == "no" else re.compile(re.escape(values["transgender"]), re.I))
    _choose_radio_group(page, re.compile(r"gender identity", re.I), re.compile(r"woman|female", re.I) if values["gender_identity"] else None)
    _choose_radio_group(page, re.compile(r"disabil", re.I), re.compile(r"no,? i do not|^no$", re.I) if values["disability"].lower() == "no" else None)
    if values["willing_to_relocate"].lower() in {"yes", "true", "1"}:
        _choose_radio_group(page, re.compile(r"located or willing to relocate", re.I), re.compile(r"planning to relocate", re.I))
    if values["sponsorship_required"].lower() in {"no", "false", "0"}:
        _choose_yes_no(page, re.compile(r"require visa sponsorship", re.I), yes=False)
    _fill_or_select(page, re.compile(r"how did you hear", re.I), values["referral_source"])
    _answer_demographics_and_eligibility(page, values)


def _fill_greenhouse_form(page: Any, candidate: ApplicationCandidate, values: dict[str, str]) -> None:
    _fill_empty(page.locator("#first_name"), values["first_name"])
    _fill_empty(page.locator("#last_name"), values["last_name"])
    _fill_empty(page.locator("#preferred_name"), values["first_name"])
    _fill_empty(page.locator("#email"), values["email"])
    _fill_empty(page.locator("#phone"), values["phone"])
    resume = page.locator('input[type="file"]#resume')
    if resume.count():
        resume.first.set_input_files(str(candidate.resume_path))
    _fill_empty(page.get_by_label(re.compile(r"LinkedIn Profile", re.I)), values["linkedin"])
    full_name = " ".join(part for part in (values["first_name"], values["last_name"]) if part).strip()
    aliases = [
        (r"^(?:Full|Legal) Name\b", full_name),
        (r"^First Name\b", values["first_name"]), (r"^Last Name\b", values["last_name"]),
        (r"^Preferred (?:First )?Name\b", values["first_name"]),
        (r"^(?:Preferred )?Pronouns?\b", values["pronouns"]),
        (r"^Email\b", values["email"]), (r"^Phone(?: Number)?\b", values["phone"]),
        (r"^LinkedIn(?: URL| Profile)?\b", values["linkedin"]),
        (r"^Current (?:Company|Employer)\b", values["current_employer"]),
        (r"^GitHub(?: URL)?\b", values["github"]),
    ]
    for label, value in aliases:
        _fill_empty(page.get_by_label(re.compile(label, re.I)), value)
    _choose_labeled_select(page, re.compile(r"^Country", re.I), "United States")
    if values["work_authorized"].lower() in {"yes", "true", "1"}:
        _choose_labeled_select(page, re.compile(r"legally authorized to work", re.I), "Yes")
    if values["sponsorship_required"].lower() in {"no", "false", "0"}:
        _choose_labeled_select(page, re.compile(r"require sponsorship", re.I), "No")
    _fill_or_select(page, re.compile(r"how did you hear", re.I), values["referral_source"])
    location = page.get_by_label(re.compile(r"^(?:Current )?Location(?: \(City\))?\b", re.I))
    if values["location"] and location.count() and not _locator_has_value(location):
        _choose_location(page, location.first, values["location"])
    _answer_demographics_and_eligibility(page, values)


def _fill_location_control(page: Any, value: str) -> None:
    if not value:
        return
    location = page.get_by_label(re.compile(r"^(?:Current )?Location(?: \(City\))?\b", re.I))
    if not location.count():
        location = page.get_by_placeholder(re.compile(r"^Start typing", re.I))
    if location.count() and not _locator_has_value(location):
        _choose_location(page, location.first, value)


def _ensure_resume_attached(page: Any, candidate: ApplicationCandidate) -> None:
    candidates = [
        page.get_by_label("Resume", exact=True),
        page.get_by_label(re.compile(r"^Resume/CV", re.I)),
        page.locator('input[type="file"]#resume'),
    ]
    for locator in candidates:
        if not locator.count():
            continue
        field = locator.first
        try:
            if not field.evaluate("e => e.files && e.files.length"):
                field.set_input_files(str(candidate.resume_path))
            return
        except Exception:
            continue


def _fill_or_select(page: Any, label_pattern: re.Pattern[str], value: str) -> None:
    if not value:
        return
    control = page.get_by_label(label_pattern)
    if not control.count():
        return
    field = control.first
    try:
        tag = field.evaluate("e => e.tagName.toLowerCase()")
        if tag == "select":
            labels = field.locator("option").all_text_contents()
            match = next((label for label in labels if value.lower() in label.lower()), None)
            if match:
                field.select_option(label=match)
            return
        if tag in {"input", "textarea"}:
            is_combobox = (field.get_attribute("role") or "").lower() == "combobox"
            if is_combobox:
                field.fill("", timeout=3_000)
                field.fill(value, timeout=3_000)
            elif not field.input_value().strip():
                field.fill(value, timeout=3_000)
            page.wait_for_timeout(250)
            options = page.locator('[role="option"]:visible')
            if options.count():
                preferred = options.filter(has_text=re.compile(re.escape(value), re.I))
                if preferred.count():
                    preferred.first.click(timeout=3_000)
            return
        field.click(timeout=3_000)
        option = page.get_by_role("option", name=re.compile(re.escape(value), re.I))
        if option.count():
            option.first.click(timeout=3_000)
    except Exception:
        pass


def _answer_known_choice(page: Any, question: str, answers: str | list[str]) -> None:
    """Answer a supplied fixed-value question across text, native, and custom controls."""
    if isinstance(answers, str):
        answers = [answers]
    answers = [answer for answer in answers if answer]
    if not answers:
        return
    pattern = re.compile(question, re.I)
    for answer in answers:
        _fill_or_select(page, pattern, answer)
        _choose_nearby_custom_option(page, pattern, answer)
        _choose_radio_group(page, pattern, re.compile(rf"^{re.escape(answer)}(?:\b|$)", re.I))
        if _question_is_answered(page, pattern):
            return
    if answers[0].lower() in {"yes", "no"}:
        _choose_yes_no(page, pattern, yes=answers[0].lower() == "yes")


def _choose_nearby_custom_option(page: Any, question_pattern: re.Pattern[str], answer: str) -> None:
    """Handle Greenhouse/Ashby custom selects without a label-control association."""
    labels = page.locator("label")
    matches = []
    for index in range(labels.count()):
        label = labels.nth(index)
        text = label.inner_text().strip()
        if question_pattern.search(text):
            matches.append((len(text), label))
    if not matches:
        return
    question = min(matches, key=lambda item: item[0])[1]
    for levels in range(1, 6):
        container = question.locator("xpath=" + "/.." * levels)
        controls = container.locator(
            'input[role="combobox"], [role="combobox"], button[aria-haspopup="listbox"]'
        )
        visible = [controls.nth(i) for i in range(controls.count()) if controls.nth(i).is_visible()]
        if len(visible) != 1:
            continue
        try:
            visible[0].click(timeout=3_000)
            page.wait_for_timeout(250)
            options = page.locator('[role="option"]:visible')
            preferred = options.filter(has_text=re.compile(re.escape(answer), re.I))
            if preferred.count():
                preferred.first.click(timeout=3_000)
                return
        except Exception:
            return


def _answer_demographics_and_eligibility(page: Any, values: dict[str, str]) -> None:
    """Apply only explicit candidate answers; never infer sensitive attributes."""
    authorized = "Yes" if values["work_authorized"].lower() in {"yes", "true", "1"} else ""
    sponsorship = "No" if values["sponsorship_required"].lower() in {"no", "false", "0"} else ""
    disability = ["No", "No, I don't have a disability", "No, I do not have a disability"] if values["disability"].lower() == "no" else values["disability"]
    veteran = ["No", "Not a protected veteran", "I am not a protected veteran"] if values["veteran_status"].lower() in {"no", "not a protected veteran"} else values["veteran_status"]
    gender = ["Female", "Woman"] if re.search(r"woman|female", values["gender_identity"], re.I) else values["gender_identity"]
    hispanic = ["No", "Not Hispanic or Latino"] if values["hispanic_latino"].lower() in {"no", "non-hispanic", "not hispanic or latino"} else values["hispanic_latino"]
    race = ["Asian", "Asian (Not Hispanic or Latino)"] if "asian" in values["race"].lower() else values["race"]
    transgender = ["No", "No, I do not identify as transgender"] if values["transgender"].lower() == "no" else values["transgender"]
    sexual_orientation = values.get("sexual_orientation", "")
    answers = [
        (r"(?:legally |lawfully )?authorized to work", authorized),
        (r"(?:need|require).{0,35}sponsorship|sponsorship.{0,35}(?:need|required)", sponsorship),
        (r"^Gender\b|gender identity", gender),
        (r"transgender|trans experience", transgender),
        (r"Hispanic|Latino", hispanic),
        (r"identify your race|race(?: or| and)? ethnicity|racial/ethnic|ethnicities|^Race\b", race),
        (r"protected veteran|veteran status|military status|veteran or active|served in the military", veteran),
        (r"disability status|have a disability|live with a disability|do you have a disability", disability),
        (r"sexual orientation", sexual_orientation),
    ]
    for question, answer in answers:
        _answer_known_choice(page, question, answer)


def _choose_labeled_select(page: Any, label_pattern: re.Pattern[str], option_text: str) -> None:
    control = page.get_by_label(label_pattern)
    if not control.count():
        return
    try:
        tag = control.first.evaluate("e => e.tagName.toLowerCase()")
        if tag == "select":
            labels = control.first.locator("option").all_text_contents()
            match = next((label for label in labels if option_text.lower() in label.lower()), None)
            if match:
                control.first.select_option(label=match)
            return
        control.first.click(timeout=3_000)
        option = page.get_by_role("option", name=re.compile(rf"^{re.escape(option_text)}$", re.I))
        if option.count():
            option.first.click(timeout=3_000)
            return
        control.first.fill(option_text, timeout=3_000)
        control.first.press("ArrowDown", timeout=3_000)
        control.first.press("Enter", timeout=3_000)
    except Exception:
        pass


def _locator_has_value(locator: Any) -> bool:
    if not locator.count():
        return False
    try:
        field = locator.first
        if (field.get_attribute("type") or "").lower() == "file":
            return bool(field.evaluate("e => e.files && e.files.length"))
        return bool(field.input_value().strip())
    except Exception:
        try:
            text = (locator.first.inner_text() or "").strip()
            return bool(text and text.lower() not in {"select...", "start typing...", "type here..."})
        except Exception:
            return False


def _any_label_has_value(page: Any, labels: list[str]) -> bool:
    return any(_locator_has_value(page.get_by_label(label, exact=True)) for label in labels)


def _question_is_answered(page: Any, pattern: re.Pattern[str]) -> bool:
    direct = page.get_by_label(pattern)
    if direct.count():
        try:
            if direct.first.is_checked():
                return True
        except Exception:
            pass
        if _locator_has_value(direct):
            return True
    labels = page.locator("label")
    matches = []
    for index in range(labels.count()):
        label = labels.nth(index)
        text = label.inner_text().strip()
        if pattern.search(text):
            matches.append((len(text), label))
    if not matches:
        return True
    question = min(matches, key=lambda item: item[0])[1]
    for levels in range(1, 7):
        container = question.locator("xpath=" + "/.." * levels)
        choices = container.locator('input[type="radio"], input[type="checkbox"]')
        if choices.count() and any(choices.nth(i).is_checked() for i in range(choices.count())):
            return True
    return False


def _question_exists(page: Any, pattern: re.Pattern[str]) -> bool:
    if page.get_by_label(pattern).count():
        return True
    labels = page.locator("label")
    return any(pattern.search(labels.nth(index).inner_text().strip()) for index in range(labels.count()))


def _verify_prefill(page: Any) -> list[str]:
    ats = detect_ats(page.url)
    missing: list[str] = []
    if ats == "ashby":
        checks = {
            "name": ["Name", "Full Name", "Legal First and Last Name"],
            "email": ["Email"],
            "resume": ["Resume"],
            "LinkedIn": ["LinkedIn", "LinkedIn URL", "LinkedIn Profile", "Linkedin Profile Link"],
        }
        for field, aliases in checks.items():
            if any(page.get_by_label(alias, exact=True).count() for alias in aliases) and not _any_label_has_value(page, aliases):
                missing.append(field)
        if any(page.get_by_label(alias, exact=True).count() for alias in ["Location", "Current Location"]):
            if not _any_label_has_value(page, ["Location", "Current Location"]):
                missing.append("location")
        if not _question_is_answered(page, re.compile(r"(?:legally )?authorized to work", re.I)):
            missing.append("work authorization")
        if not _question_is_answered(page, re.compile(r"require visa sponsorship", re.I)):
            missing.append("sponsorship")
    elif ats == "greenhouse":
        checks = {
            "first name": page.locator("#first_name"),
            "last name": page.locator("#last_name"),
            "email": page.locator("#email"),
            "resume": page.locator("#resume"),
            "LinkedIn": page.get_by_label(re.compile(r"LinkedIn Profile", re.I)),
        }
        missing.extend(name for name, locator in checks.items() if locator.count() and not _locator_has_value(locator))
        if not _question_is_answered(page, re.compile(r"legally authorized to work", re.I)):
            missing.append("work authorization")
        if not _question_is_answered(page, re.compile(r"require sponsorship", re.I)):
            missing.append("sponsorship")
    common_questions = {
        "gender": re.compile(r"^Gender\b|gender identity", re.I),
        "transgender": re.compile(r"transgender|trans experience", re.I),
        "race/ethnicity": re.compile(r"identify your race|race(?: or| and)? ethnicity|racial/ethnic|ethnicities|^Race\b", re.I),
        "veteran status": re.compile(r"protected veteran|veteran status|military status|veteran or active|served in the military", re.I),
        "disability": re.compile(r"disability status|have a disability|live with a disability|do you have a disability", re.I),
    }
    for name, pattern in common_questions.items():
        if _question_exists(page, pattern) and not _question_is_answered(page, pattern):
            missing.append(name)
    return missing


def _choose_custom_select(page: Any, label_pattern: re.Pattern[str], value: str) -> None:
    if not value:
        return
    control = page.get_by_label(label_pattern)
    if not control.count():
        return
    try:
        control.first.fill(value, timeout=3_000)
        page.wait_for_timeout(300)
        control.first.press("ArrowDown", timeout=3_000)
        control.first.press("Enter", timeout=3_000)
    except Exception:
        pass


def _choose_location(page: Any, field: Any, value: str) -> None:
    city = value.split(",", 1)[0].strip() or value
    try:
        tag = field.evaluate("e => e.tagName.toLowerCase()")
        if tag not in {"input", "textarea"}:
            inputs = field.locator("input")
            if inputs.count():
                field = inputs.first
    except Exception:
        pass
    attempts = [city, value] if city != value else [city]
    for attempt in attempts:
        try:
            field.click(timeout=3_000)
            field.fill("")
            field.press_sequentially(attempt, delay=80, timeout=5_000)
            page.wait_for_timeout(3_500)
            options = page.locator('[role="option"]:visible')
            if options.count():
                preferred = options.filter(has_text=re.compile(re.escape(city), re.I))
                (preferred.first if preferred.count() else options.first).click(timeout=3_000)
                page.wait_for_timeout(500)
                return
        except Exception:
            pass
    # Last resort: arrow-down + enter on whatever is showing
    try:
        field.press("ArrowDown", timeout=3_000)
        field.press("Enter", timeout=3_000)
        page.wait_for_timeout(500)
    except Exception:
        pass


def _choose_yes_no(page: Any, question_pattern: re.Pattern[str], *, yes: bool) -> None:
    labels = page.locator("label")
    matches = []
    for index in range(labels.count()):
        label = labels.nth(index)
        text = label.inner_text().strip()
        if question_pattern.search(text):
            matches.append((len(text), label))
    if not matches:
        return
    question = min(matches, key=lambda item: item[0])[1]
    desired = re.compile(r"^Yes$" if yes else r"^No$", re.I)
    for levels in range(0, 7):
        container = question if levels == 0 else question.locator("xpath=" + "/.." * levels)
        buttons = container.get_by_role("button", name=desired)
        if buttons.count():
            try:
                buttons.first.click(timeout=3_000)
                return
            except Exception:
                pass
    try:
        question.click(timeout=3_000)
    except Exception:
        pass


def _choose_radio_group(page: Any, question_pattern: re.Pattern[str], option_pattern: re.Pattern[str] | None) -> None:
    if option_pattern is None:
        return
    labels = page.locator("label")
    matching_questions = []
    for index in range(labels.count()):
        label = labels.nth(index)
        text = label.inner_text().strip()
        if question_pattern.search(text):
            matching_questions.append((len(text), label))
    if not matching_questions:
        return
    question = min(matching_questions, key=lambda item: item[0])[1]
    for levels in range(1, 7):
        container = question.locator("xpath=" + "/.." * levels)
        if container.locator('input[type="radio"]').count() < 2:
            continue
        labels = container.locator("label")
        for index in range(labels.count()):
            label = labels.nth(index)
            if not option_pattern.search(label.inner_text().strip()):
                continue
            try:
                label.click(timeout=3_000)
            except Exception:
                pass
            return


def _launch_browser_context(playwright: Any, *, headless: bool) -> Any:
    """Launch a fresh Chrome browser context — no persistent profile, no conflict with running Chrome.

    LinkedIn auth is handled via the LINKEDIN_COOKIE (li_at) env var injected as a cookie.
    Google search runs via HTTP (see _resolve_application_url) so no Google session is needed here.
    """
    browser = playwright.chromium.launch(
        channel="chrome",
        headless=headless,
        args=["--no-default-browser-check"],
    )
    context = browser.new_context(
        user_agent=_SEARCH_HEADERS["User-Agent"],
        viewport=None,
    )
    linkedin_cookie = os.getenv("LINKEDIN_COOKIE", "").strip()
    if linkedin_cookie:
        try:
            context.add_cookies([{
                "name": "li_at",
                "value": linkedin_cookie,
                "domain": ".linkedin.com",
                "path": "/",
                "httpOnly": True,
                "secure": True,
                "sameSite": "None",
            }])
        except Exception as exc:
            log.debug("apply: LinkedIn cookie injection failed: %s", exc)
    return context


def _find_linkedin_apply_control(page: Any, *, timeout_seconds: float = 15) -> Any | None:
    # Authenticated LinkedIn commonly uses an accessible name such as
    # "Apply to <role> on company website", and hydrates the button after
    # DOMContentLoaded. Wait for those real variants instead of checking once.
    candidates = [
        page.locator("button.jobs-apply-button"),
        page.locator('button[aria-label^="Apply to" i]'),
        page.locator('a[aria-label^="Apply to" i]'),
        page.get_by_role("button", name=re.compile(r"\b(?:easy )?apply\b", re.I)),
        page.get_by_role("link", name=re.compile(r"\b(?:easy )?apply\b", re.I)),
        page.locator('button[data-modal="job-details-subnav-apply-modal"]'),
    ]
    control = None
    deadline = time.monotonic() + timeout_seconds
    while control is None and time.monotonic() < deadline:
        for controls in candidates:
            for index in range(controls.count()):
                item = controls.nth(index)
                try:
                    if item.is_visible():
                        control = item
                        break
                except Exception:
                    continue
            if control is not None:
                break
        if control is None:
            page.wait_for_timeout(250)
    return control


def _navigate_to_application(context: Any, page: Any) -> tuple[Any, str]:
    """Follow LinkedIn's Apply control, returning (destination, easy|external)."""
    if detect_ats(page.url):
        return page, "external"
    path = urlparse(page.url).path.lower()
    is_auth_route = any(marker in path for marker in ("/login", "/signup", "/authwall"))
    if is_auth_route:
        raise LinkedInAuthenticationError(
            "LinkedIn is not signed in in Chrome's Default profile. Sign into LinkedIn in that profile, "
            "quit Chrome, then retry."
        )
    control = _find_linkedin_apply_control(page)
    if control is None:
        raise RuntimeError("Could not find LinkedIn's Apply control on the posting.")
    text = " ".join(filter(None, [
        (control.inner_text() or "").strip(),
        (control.get_attribute("aria-label") or "").strip(),
    ]))
    if "easy apply" in text.lower():
        return page, "easy"
    before = list(context.pages)
    control.click(timeout=15_000)
    page.wait_for_timeout(1_500)
    new_pages = [item for item in context.pages if item not in before]
    destination = new_pages[-1] if new_pages else page
    destination.wait_for_load_state("domcontentloaded", timeout=60_000)
    if "linkedin.com" in (urlparse(destination.url).hostname or "").lower():
        raise RuntimeError("LinkedIn Apply did not open an external career page.")
    if not _is_external_job_url(destination.url):
        raise RuntimeError(f"LinkedIn Apply opened an unsupported destination: {destination.url}")
    return destination, "external"


def _navigate_external_to_form(context: Any, page: Any) -> Any:
    """Follow an employer job-detail Apply control to the actual application form."""
    if detect_ats(page.url):
        return page
    controls = [
        page.get_by_role("link", name=re.compile(r"^(?:apply|apply now|apply for this job)\b", re.I)),
        page.get_by_role("button", name=re.compile(r"^(?:apply|apply now|apply for this job)\b", re.I)),
    ]
    control = None
    for group in controls:
        for index in range(group.count()):
            item = group.nth(index)
            try:
                text = (item.inner_text() or "").strip().lower()
                if item.is_visible() and "submit" not in text:
                    control = item
                    break
            except Exception:
                continue
        if control is not None:
            break
    if control is None:
        return page
    before = list(context.pages)
    original = page
    control.click(timeout=15_000)
    page.wait_for_timeout(1_000)
    new_pages = [item for item in context.pages if item not in before]
    destination = new_pages[-1] if new_pages else page
    destination.wait_for_load_state("domcontentloaded", timeout=60_000)
    if destination is not original:
        try:
            original.close()
        except Exception:
            pass
    return destination


def _sync_tracker_applied(candidate: ApplicationCandidate) -> None:
    import openpyxl
    from . import xlsx
    from .track_write import find_tracker_row, update_confirmation

    xlsx.write_to_tracker([candidate.role])
    if not xlsx.TRACKER_PATH.exists():
        return
    workbook = openpyxl.load_workbook(xlsx.TRACKER_PATH)
    sheet = workbook.active
    row = find_tracker_row(sheet, str(candidate.role.get("company") or ""), str(candidate.role.get("title") or ""))
    if row is None:
        return
    update_confirmation(sheet, row, dt.date.today().isoformat())
    xlsx.format_tracker_sheet(sheet)
    workbook.save(xlsx.TRACKER_PATH)
    xlsx.TRACKER_PATH.chmod(0o600)


def prepare_in_browser(candidate: ApplicationCandidate, *, root: Path | None = None, headless: bool = False) -> str:
    return prepare_batch_in_browser([candidate], root=root, headless=headless)[0]["status"]


def prepare_batch_in_browser(
    candidates: list[ApplicationCandidate], *, root: Path | None = None, headless: bool = False,
) -> list[dict[str, str]]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise SystemExit("Browser support is not installed. Run `pip install 'metis-job[browser]'`.") from exc

    root = root or data_dir()
    results: dict[str, dict[str, str]] = {}
    with sync_playwright() as playwright:
        context = _launch_browser_context(playwright, headless=headless)
        active: dict[Any, tuple[ApplicationCandidate, dict[str, Any]]] = {}
        print("Launched Chrome for form filling (fresh session, LinkedIn cookie injected).")
        # Only probe LinkedIn if Easy Apply roles are in the batch — search runs via HTTP.
        needs_linkedin = any(
            str(item.role.get("apply_mode") or "").lower() == "easy_apply" for item in candidates
        )
        linkedin_ok = True
        if needs_linkedin:
            probe_page = context.new_page()
            linkedin_ok = _probe_linkedin_session(context, probe_page)
            probe_page.close()
            if not linkedin_ok:
                print(
                    "⚠  LinkedIn session not valid — Easy Apply roles will be blocked. "
                    "Update LINKEDIN_COOKIE in ~/.job_pipeline/.env with a fresh li_at value "
                    "from Chrome DevTools (Application → Cookies → linkedin.com)."
                )
        for candidate_index, candidate in enumerate(candidates):
            start_url = _start_url(candidate.role)
            ats = detect_ats(start_url) or "unresolved"
            common = {
                "role": candidate.role,
                "tailoring_record": str(candidate.record_path) if candidate.record_path else None,
                "resume_used": str(candidate.resume_path),
                "ats": ats,
            }
            update_application_state(candidate.role_key, status="opened", root=root, **common)
            page = None
            try:
                apply_mode_val = str(candidate.role.get("apply_mode") or "").lower()

                def _mark_blocked(reason: str) -> None:
                    common.update({
                        "resolution_status": "unresolved",
                        "resolution_attempted_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
                    })
                    update_application_state(candidate.role_key, status="blocked", root=root, **common)
                    print(f"Could not prepare {candidate.role.get('title')} at {candidate.role.get('company')}: {reason}.")
                    results[candidate.role_key] = {
                        "role": f"{candidate.role.get('title')} at {candidate.role.get('company')}",
                        "status": "blocked",
                    }

                # ── Step 1: Easy Apply path (LinkedIn inline form) ────────────────
                if apply_mode_val == "easy_apply":
                    if not linkedin_ok:
                        _mark_blocked("Easy Apply requires a valid LINKEDIN_COOKIE in .env")
                        continue
                    li_page = context.new_page()
                    li_page.goto(start_url, wait_until="domcontentloaded", timeout=60_000)
                    try:
                        destination, route = _navigate_to_application(context, li_page)
                    except LinkedInAuthenticationError as linkedin_exc:
                        message = str(linkedin_exc)
                        _write_apply_diagnostic(root, candidate, phase="linkedin_auth", error=message, page=li_page)
                        linkedin_ok = False
                        log.warning("LinkedIn auth lost mid-batch for %s: %s", candidate.role_key, message)
                        update_application_state(candidate.role_key, status="blocked", root=root, error=message, **common)
                        results[candidate.role_key] = {
                            "role": f"{candidate.role.get('title')} at {candidate.role.get('company')}",
                            "status": "blocked",
                        }
                        li_page.close()
                        print(f"LinkedIn cookie invalid — skipping {candidate.role.get('company')}. Update LINKEDIN_COOKIE in .env.")
                        continue
                    except Exception as linkedin_exc:
                        log.info("LinkedIn Apply routing failed for %s: %s", candidate.role_key, linkedin_exc)
                        _write_apply_diagnostic(root, candidate, phase="linkedin_routing", error=str(linkedin_exc), page=li_page)
                        destination = None
                        route = "unresolved"
                    if route == "easy":
                        page = li_page
                        common.update({"ats": "linkedin", "application_url": page.url, "resolution_status": "easy_apply"})
                        update_application_state(candidate.role_key, status="opened_linkedin", root=root, **common)
                        results[candidate.role_key] = {
                            "role": f"{candidate.role.get('title')} at {candidate.role.get('company')}",
                            "status": "opened_linkedin",
                        }
                        active[page] = (candidate, common)
                        continue
                    elif route == "external" and destination is not None:
                        # Edge case: clicking Apply on LinkedIn opened an external ATS page.
                        page = destination
                        if destination is not li_page:
                            li_page.close()
                        common.update({
                            "application_url": page.url,
                            "resolution_status": "confirmed",
                            "resolution_method": "linkedin_apply",
                            "resolved_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
                        })
                    else:
                        li_page.close()
                        _mark_blocked("LinkedIn Easy Apply navigation failed")
                        continue

                # ── Step 2: Offsite / unknown path (HTTP search for ATS URL) ─────
                else:
                    # Recruiter aggregators hide the employer; no discoverable ATS URL.
                    source_host = (urlparse(start_url).hostname or "").lower().removeprefix("www.")
                    company_lower = str(candidate.role.get("company") or "").lower()
                    if (
                        source_host in _AGGREGATOR_DOMAINS
                        or any(source_host.endswith(f".{d}") for d in _AGGREGATOR_DOMAINS)
                        or any(sub in company_lower for sub in _AGGREGATOR_COMPANY_SUBSTRINGS)
                    ):
                        _mark_blocked(f"role posted by recruiter aggregator ({company_lower or source_host}) — no employer identity discoverable")
                        continue

                    direct_external = _is_external_job_url(start_url)
                    if direct_external:
                        # apply_url or direct employer URL already known — skip search.
                        page = context.new_page()
                        page.goto(start_url, wait_until="domcontentloaded", timeout=60_000)
                    else:
                        # HTTP search — no browser, no profile, no CAPTCHA.
                        if candidate_index > 0:
                            time.sleep(1.0)
                        google_url: str | None = None
                        try:
                            google_url = _resolve_application_url(candidate)
                        except Exception:
                            pass
                        if google_url:
                            page = context.new_page()
                            page.goto(google_url, wait_until="domcontentloaded", timeout=60_000)
                            common.update({
                                "application_url": google_url,
                                "resolution_status": "confirmed",
                                "resolution_method": "google_search",
                                "resolved_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
                            })
                        else:
                            _mark_blocked("web search found no ATS URL")
                            continue
                page = _navigate_external_to_form(context, page)
                common["ats"] = detect_ats(page.url) or ats
                common["application_url"] = page.url
                _fill_visible_form(page, candidate)
                missing = _verify_prefill(page)
                status = "needs_review" if missing else "prefilled"
                update_application_state(
                    candidate.role_key, status=status, root=root, prefill_missing=missing, **common,
                )
                if missing:
                    print(f"Needs review: {candidate.role.get('title')} at {candidate.role.get('company')} — missing {', '.join(missing)}")
                results[candidate.role_key] = {
                    "role": f"{candidate.role.get('title')} at {candidate.role.get('company')}", "status": status,
                }
                active[page] = (candidate, common)
            except Exception as exc:
                _write_apply_diagnostic(
                    root, candidate, phase="prepare_form", error=str(exc), page=page,
                )
                update_application_state(candidate.role_key, status="blocked", root=root, error=str(exc), **common)
                print(f"Could not prepare {candidate.role.get('title')} at {candidate.role.get('company')}: {exc}")
                results[candidate.role_key] = {"role": f"{candidate.role.get('title')} at {candidate.role.get('company')}", "status": "blocked"}
                if page is not None:
                    page.close()
        # Close any pages that were opened but are not active application tabs.
        for page in list(context.pages):
            if page not in active:
                try:
                    page.close()
                except Exception:
                    pass
        if headless or not active:
            context.close()
            return [results[item.role_key] for item in candidates]
        print(f"Prepared {len(active)} application tab(s). Review each and click Submit; close the tabs to stop monitoring.")
        try:
            while active:
                for page, (candidate, common) in list(active.items()):
                    if page.is_closed():
                        active.pop(page, None)
                        continue
                    if common.get("ats") == "linkedin":
                        continue
                    try:
                        if _looks_submitted(page.url, page.locator("body").inner_text(timeout=2_000)):
                            update_application_state(
                                candidate.role_key, status="applied", root=root,
                                submission_evidence={"url": page.url, "kind": "browser_success"}, **common,
                            )
                            _sync_tracker_applied(candidate)
                            results[candidate.role_key]["status"] = "applied"
                            active.pop(page, None)
                    except Exception:
                        if page.is_closed():
                            active.pop(page, None)
                if active:
                    time.sleep(1)
        except Exception as exc:
            # Closing the last tab or the entire Chrome window tears down Playwright's
            # page/context while the monitor may be between calls. That is a normal exit.
            if "closed" not in str(exc).lower():
                raise
        try:
            context.close()
        except Exception:
            pass
    final = [results[item.role_key] for item in candidates]
    counts: dict[str, int] = {}
    for r in final:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    summary_lines = []
    if counts.get("applied") or counts.get("prefilled") or counts.get("needs_review"):
        summary_lines.append(f"  Prepared externally:  {counts.get('prefilled', 0) + counts.get('needs_review', 0) + counts.get('applied', 0)}")
    if counts.get("opened_linkedin"):
        summary_lines.append(f"  Opened in LinkedIn:   {counts['opened_linkedin']}")
    if counts.get("blocked"):
        summary_lines.append(f"  Could not resolve:    {counts['blocked']}")
    if counts.get("auth_required"):
        summary_lines.append(f"  Auth required:        {counts['auth_required']}")
    if summary_lines:
        print("\nBatch summary:")
        print("\n".join(summary_lines))
    return final


def run_apply(
    *, apply_all: bool = False, top_n: int | None = None, include_applied: bool = False,
    match_terms: list[str] | None = None, latest_n: int | None = None,
    lookback: str | None = None, resume_path: str | None = None,
    force_default_resume: bool = False,
) -> list[dict[str, str]]:
    configured_default = _fallback_resume()
    override: Path | None = None
    if resume_path:
        override = Path(resume_path).expanduser()
        if not override.is_file() or override.suffix.lower() != ".docx":
            raise SystemExit("--resume must point to an existing DOCX file.")
    elif force_default_resume:
        if configured_default is None:
            raise SystemExit("No configured default resume. Run `metis config autofill` or pass --resume DOCX.")
        override = configured_default
    candidates = load_application_candidates(include_applied=include_applied, resume_override=override)
    if override is not None:
        kind = "custom" if resume_path else "default"
        candidates = [replace(candidate, resume_path=override, tailored=False, resume_kind=kind) for candidate in candidates]

    if lookback:
        from .pipeline import _parse_lookback

        since = _parse_lookback(lookback)
        if since is None:
            raise SystemExit(f"Could not parse --lookback '{lookback}'. Try: '7d', '30d', '2026-07-01'.")
        cutoff = since.date().isoformat()
        candidates = [candidate for candidate in candidates if _candidate_date(candidate) >= cutoff]
    terms = [term.strip().lower() for term in (match_terms or []) if term.strip()]
    if terms:
        candidates = [
            candidate for candidate in candidates
            if any(term in f"{candidate.role.get('company', '')} {candidate.role.get('title', '')}".lower() for term in terms)
        ]
    if not candidates:
        raise SystemExit(_empty_gate_message(
            lookback=lookback, include_applied=include_applied, match_terms=match_terms,
        ))
    if top_n is not None:
        if top_n <= 0:
            raise SystemExit("--top must be a positive integer.")
        preselected = candidates[:top_n]
        selected = select_candidates(candidates, preselected=preselected) if os.isatty(0) else preselected
    elif latest_n is not None:
        if latest_n <= 0:
            raise SystemExit("--latest must be a positive integer.")
        preselected = sorted(
            candidates,
            key=lambda candidate: (_candidate_date(candidate), int((candidate.role.get("eval") or {}).get("score") or 0)),
            reverse=True,
        )[:latest_n]
        selected = select_candidates(candidates, preselected=preselected) if os.isatty(0) else preselected
    elif apply_all:
        selected = candidates
    elif not os.isatty(0):
        raise SystemExit("Run `metis apply --all`, or run interactively to choose roles.")
    else:
        selected = select_candidates(candidates)
    return prepare_batch_in_browser(selected)
