"""scorerole/track.py — thin orchestrator for the track pipeline stage.

Usage:
    metis track                   # parse emails from last 7 days
    metis track --lookback 30d    # extend lookback
    metis track --no-excel        # print matches to stdout, no xlsx write or open

Sub-modules own the actual logic:
  track_imap.py   — IMAP fetch + email decode helpers
  track_parse.py  — email classification + entity extraction
  track_write.py  — tracker fuzzy match + xlsx write operations

This module wires them together and exposes run_track() and backfill_from_digests().
"""
from __future__ import annotations

import datetime
import email as email_lib
import imaplib
import logging
import re
import sys
import time

log = logging.getLogger(__name__)

# Re-exports for backward-compat (tests import from scorerole.track)
from .track_parse import classify_email, _LLM_VALID_CLASSES  # noqa: F401
from .track_write import update_recruiter_screen              # noqa: F401

from .track_imap  import (
    fetch_candidate_emails,
    _safe_decode,
    _IMAP_MAX_RETRIES,
    _IMAP_RETRY_DELAY,
)
from .track_parse import parse_email
from .track_write import (
    find_tracker_row,
    update_confirmation,
    update_rejection,
    create_skipped_row,
    create_backfill_row,
    _parse_digest_html,
)

# Subject patterns that strongly imply recruiter screen (used as body tiebreaker)
_SUBJECT_IMPLIES_RECRUITER_SCREEN = re.compile(
    r"next steps with |"                  # "Next Steps with Descript"
    r"\w[\w ]+ next steps$|"             # "SeekOut Next Steps", "Klaviyo - Next Steps"
    r"let's set up your phone screen|"   # Microsoft eightfold
    r"phone screen with |"               # generic
    r"hello from \w",                    # Datadog "Hello from Datadog"
    re.IGNORECASE,
)

# Subject patterns that strongly imply rejection (used as body tiebreaker)
_SUBJECT_IMPLIES_REJECTION = re.compile(
    r"thank you from |"           # NVIDIA, Elastic, Workato
    r"following up on your (application|recent application)|"
    r"application (status )?update|"
    r"update on your .+ application|"
    r"application feedback from |"  # Google DeepMind
    r"important information about your application",  # Altruist
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Company extraction — tried in order, first match wins
# ---------------------------------------------------------------------------

_COMPANY_FROM_SUBJECT = [
    # "applying to/at {COMPANY}[!, comma, end]"  — Yelp uses "at", most use "to"
    re.compile(r"applying (?:to|at) ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "application to the {ROLE} role at {COMPANY}" — must come before bare "application to"
    re.compile(r"application to the .+? (?:role|position) at ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "your application to {COMPANY}"
    re.compile(r"your application to ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "your application for {ROLE} at {COMPANY}" — must come before bare "applying to"
    re.compile(r"your application for .+ at ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "Thanks for applying for the role of {ROLE} at {COMPANY}" (Workday)
    re.compile(r"applying for the role of .+ at ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "interest in {COMPANY}" / "interest with {COMPANY}"
    re.compile(r"interest (?:in|with) ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "thank you from {COMPANY}" (NVIDIA, Workato)
    re.compile(r"thank you from ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "thanks for applying to {COMPANY}"
    re.compile(r"thanks for applying to ([A-Za-z0-9][^!,\n]+?)(?:[!,\s]|$)", re.IGNORECASE),
    # "{COMPANY} | We Got It!" / "Luma | Thanks for..."
    re.compile(r"^([A-Za-z0-9][^|]+?)\s*\|", re.IGNORECASE),
    # "following up on your (recent) application to {COMPANY}" (Google rejection)
    re.compile(r"following up on your (?:recent )?application to ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "following up on your application with {COMPANY}" (Front)
    re.compile(r"application with ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "{COMPANY} - Thank You For..." or "Google DeepMind - Thank You..."
    re.compile(r"^([A-Za-z0-9][^-|]+?)\s+[-–]", re.IGNORECASE),
    # "Update on your {COMPANY} Application"
    re.compile(r"update on your ([A-Za-z0-9][^a-z]+?) application", re.IGNORECASE),
    # "Application Update from {COMPANY}" (Snorkel AI)
    re.compile(r"application update from ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "{Name}, we have received your application for {ROLE} at {COMPANY}" (Netflix)
    re.compile(r"we have received your application for .+ at ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
    # "we have received your application for {COMPANY}!" (Circle — company, not role, after "for")
    re.compile(r"we have received your application for ([A-Za-z0-9][A-Za-z0-9 &.,]+?)(?:[!,]|$)", re.IGNORECASE),
    # "{COMPANY}: Thanks for..." (Atlassian)
    re.compile(r"^([A-Za-z0-9][^:]+?):\s+thanks for", re.IGNORECASE),
    # "Alex, Thank you from {COMPANY}" — leading name prefix with candidate first name
    re.compile(r"^[A-Z][a-z]+,\s+thank you from ([A-Za-z0-9][^!,\n]+?)(?:[!,]|$)", re.IGNORECASE),
]

# Trailing noise to strip from extracted company names
_COMPANY_TRAILING_NOISE = re.compile(
    r"\s*[-–]\s*(?:[A-Z][a-z]+ [A-Z][a-z]+|[A-Z][a-z]+,? [A-Z][a-z]+)$"  # "- Alex Chen" — trailing candidate name
)
_COMPANY_LEADING_THE   = re.compile(r"^the\s+", re.IGNORECASE)
_COMPANY_SENDER_SUFFIX = re.compile(
    r"\s+(?:recruiting|talent|hiring team|hr|careers|talent acquisition)$", re.IGNORECASE
)
# Words that indicate the extracted string is a role title, not a company name
_ROLE_TITLE_WORDS = re.compile(
    r"\b(?:manager|director|engineer|product|principal|staff|senior|lead|analyst|"
    r"scientist|designer|architect|specialist|coordinator|associate|vp|head of)\b",
    re.IGNORECASE,
)

_COMPANY_FROM_BODY = [
    re.compile(r"applying (?:for|to) .+? (?:role|position)(?: here)? at ([A-Z][A-Za-z0-9 &.,]+?)[\.\!,\n]"),
    re.compile(r"apply for .+ (?:role|position) at ([A-Z][A-Za-z0-9 &.,]+?)[\.\!,\n]"),
    re.compile(r"interest in ([A-Z][A-Za-z0-9 &.,]+?) and"),
    re.compile(r"(?:joining|with) ([A-Z][A-Za-z0-9 &.,]+?)[\.\!,\n]"),
]

# ---------------------------------------------------------------------------
# Role extraction — body is primary; subject fallback
# ---------------------------------------------------------------------------

_ROLE_FROM_BODY = [
    re.compile(r"apply(?:ing)? (?:for|to) the (.+?) (?:role|position)", re.IGNORECASE),
    re.compile(r"your application for the (.+?) (?:role|position)", re.IGNORECASE),
    re.compile(r"application for the (.+?) (?:role|position)", re.IGNORECASE),
    re.compile(r"the (.+?) (?:role|position)(?: here)? at [A-Z]", re.IGNORECASE),
    re.compile(r"(?:role|position) (?:of|for) (.+?) (?:at|here)", re.IGNORECASE),
    re.compile(r"position of (.+?)(?:\s*\(|,|\.|$)", re.IGNORECASE),
    re.compile(r"for the (.+?) (?:role|position|opportunity)", re.IGNORECASE),
]

# Role strings longer than this are almost certainly boilerplate leakage
_ROLE_MAX_LEN = 120

_ROLE_FROM_SUBJECT = [
    # "Your application for {ROLE} at {COMPANY}"
    re.compile(r"your application for (.+?) at [A-Za-z]", re.IGNORECASE),
    # "applying to {ROLE} - {COMPANY}" or "Applying to {ROLE}!"
    re.compile(r"applying to (.+?)(?:\s*[-–]|\s*[!,]|$)", re.IGNORECASE),
    # "Update on your {COMPANY} Application - {ROLE}"
    re.compile(r"application\s*[-–]\s*(.+?)$", re.IGNORECASE),
]

# ---------------------------------------------------------------------------
# Email fetching
# ---------------------------------------------------------------------------

def _decode_header_value(raw: str | bytes) -> str:
    if isinstance(raw, bytes):
        parts = decode_header(raw.decode("utf-8", errors="replace"))
    else:
        parts = decode_header(raw)
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(part)
    return " ".join(decoded)


def _strip_html(html: str) -> str:
    """Strip HTML tags and decode entities to plain text."""
    import html as html_lib
    text = re.sub(r"<style[^>]*>.*?</style>", " ", html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<script[^>]*>.*?</script>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html_lib.unescape(text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _safe_decode(payload: bytes, charset: str | None) -> str:
    """Decode bytes to str, falling back gracefully on unknown/binary charsets."""
    for enc in (charset, "utf-8", "latin-1"):
        if not enc or enc.lower() in ("binary", "unknown", "x-unknown"):
            continue
        try:
            return payload.decode(enc, errors="replace")
        except (LookupError, UnicodeDecodeError):
            continue
    return payload.decode("latin-1", errors="replace")


def _extract_body(msg) -> str:
    """Extract plain-text body from an email.Message object.

    Prefers text/plain parts. Falls back to stripping text/html when no plain
    text is available — required for Amazon, Netflix, Google DeepMind, etc.
    """
    plain_parts = []
    html_parts  = []

    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            decoded = _safe_decode(payload, part.get_content_charset())
            if ct == "text/plain":
                plain_parts.append(decoded)
            elif ct == "text/html":
                html_parts.append(decoded)
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            decoded = _safe_decode(payload, msg.get_content_charset())
            if msg.get_content_type() == "text/html":
                html_parts.append(decoded)
            else:
                plain_parts.append(decoded)

    if plain_parts:
        return "\n".join(plain_parts)
    if html_parts:
        return _strip_html("\n".join(html_parts))
    return ""


def fetch_candidate_emails(
    gmail_address: str,
    app_password: str,
    since_dt: datetime.datetime,
    applied_companies: "set[str] | None" = None,
) -> list[dict]:
    """Fetch emails from Gmail that might be confirmations, rejections, or recruiter screens.

    Three-tier server-side fetch gate (all run inside one IMAP connection, results unioned):
      1. Subject keyword search — fast path for ATS-generated emails with predictable subjects.
      2. ATS sender domain search — catches ATS emails with non-standard subjects.
      3. Company name FROM search — catches direct recruiter outreach from company domains,
         which have no predictable subject pattern.
    Full RFC822 bodies are only fetched for the union of matched message IDs.
    """
    from email.utils import parsedate_to_datetime

    since_str = since_dt.strftime("%d-%b-%Y")

    # Server-side subject filters — each maps to one IMAP SEARCH call.
    # Gmail evaluates these server-side so no per-message round trips needed.
    # Each term is a COMPLETE WORD or phrase as it appears in real subjects.
    # Gmail IMAP does word-based search — "apply" does NOT match "applying".
    # All terms derived from real email samples in metis_confEmail/.
    _SUBJECT_SEARCHES = [
        # Confirmation — "Thank you for applying to {Company}" (Greenhouse, direct)
        "thank you for applying",
        "thanks for applying",
        # Confirmation — "Thank you for your application to {Company}" (Ashby, Lever)
        "thank you for your application",
        "thanks for your application",
        # Confirmation — "Your application for {Role} at {Company}" (Ashby/Superhuman)
        "your application for",
        # Confirmation — "Thanks for your interest in {Company}" (Greenhouse/Carta)
        "thanks for your interest",
        "thank you for your interest",
        # Confirmation — "Keep track of your application" (Amazon jobs)
        "keep track of your application",
        # Confirmation — "We've Received Your Application!" (Console/Ashby)
        "we've received your application",
        # Confirmation — "We received your application" (various)
        "we received your application",
        # Confirmation — "We Got It!" (Gen Digital)
        "we got it",
        # Confirmation — "Next Steps with {Company}" (Descript/Klaviyo)
        "next steps",
        # Rejection — "Following up on your application" (Google, Front)
        "following up on your",
        # Rejection/Update — "Application Update", "Update on your application"
        "application update",
        "update on your",
        # Rejection — "Thank you from {Company}" (NVIDIA, Workato)
        "thank you from",
        # Rejection — "Application feedback from {Company}" (Google DeepMind)
        "application feedback",
        # Confirmation/Rejection — "Your Application to {Company}" (Anduril direct)
        "your application to",
        # Recruiter screen — "Action Requested: Let's set up your phone screen" (Microsoft)
        "phone screen",
        # Recruiter screen — "Hello from Datadog" (named recruiter outreach)
        "hello from",
        # Update — "Application Follow Up" (GitHub/iCIMS)
        "application follow up",
        # Update — "We look forward to" (various)
        "we look forward to",
    ]

    emails = []

    for attempt in range(1, _IMAP_MAX_RETRIES + 1):
        try:
            with imaplib.IMAP4_SSL("imap.gmail.com") as imap:
                imap.login(gmail_address, app_password)
                imap.select("INBOX")

                # Union of all searches — deduplicated across all three tiers
                matching_ids: set[bytes] = set()

                # Tier 1: subject keywords — fast path for ATS-generated emails
                for term in _SUBJECT_SEARCHES:
                    _, data = imap.search(None, f'SINCE {since_str} SUBJECT "{term}"')
                    if data and data[0]:
                        matching_ids.update(data[0].split())

                # Tier 2: ATS sender domains — catches ATS emails with non-standard subjects
                for domain in _ATS_FROM_DOMAINS:
                    _, data = imap.search(None, f'SINCE {since_str} FROM "@{domain}"')
                    if data and data[0]:
                        matching_ids.update(data[0].split())

                # Tier 3: company names — catches direct recruiter outreach from company domains
                if applied_companies:
                    for company in applied_companies:
                        term = _normalize_company_search(company)
                        if len(term) >= 4:  # skip very short names to avoid noise
                            _, data = imap.search(None, f'SINCE {since_str} FROM "{term}"')
                            if data and data[0]:
                                matching_ids.update(data[0].split())

                if not matching_ids:
                    log.info("track: no candidate emails found since %s", since_str)
                    return []

                log.info("track: %d subject-matched emails — fetching bodies…", len(matching_ids))

                for msg_id in sorted(matching_ids):
                    _, raw = imap.fetch(msg_id, "(RFC822)")
                    if not raw or not raw[0]:
                        continue
                    raw_bytes = raw[0][1] if isinstance(raw[0], tuple) else raw[0]
                    msg = email_lib.message_from_bytes(raw_bytes)

                    subject     = _decode_header_value(msg.get("Subject", ""))
                    sender      = _decode_header_value(msg.get("From", ""))
                    date_header = msg.get("Date", "")

                    # Skip out-of-office auto-replies — they're triggered by our own emails
                    if re.search(r"\bout of office\b|auto.?reply|automatic reply|autoreply", subject, re.IGNORECASE):
                        log.debug("track: skipping OOO/auto-reply — %s", subject[:80])
                        continue
                    # Also check X-Auto-Submitted header (RFC 3834)
                    if msg.get("X-Auto-Submitted", "").lower() not in ("", "no"):
                        log.debug("track: skipping auto-submitted — %s", subject[:80])
                        continue

                    try:
                        email_date = parsedate_to_datetime(date_header).date().isoformat()
                    except Exception:
                        email_date = datetime.date.today().isoformat()

                    body = _extract_body(msg)
                    log.debug("track: fetched — %s | from: %s", subject[:80], sender[:50])
                    emails.append({
                        "subject": subject,
                        "sender":  sender,
                        "body":    body,
                        "date":    email_date,
                    })
            break   # success
        except OSError as e:
            if attempt < _IMAP_MAX_RETRIES:
                log.warning(
                    "track: IMAP connect failed (attempt %d/%d): %s — retrying in %ds…",
                    attempt, _IMAP_MAX_RETRIES, e, _IMAP_RETRY_DELAY,
                )
                time.sleep(_IMAP_RETRY_DELAY)
            else:
                log.error("track: IMAP connect failed after %d attempts: %s", _IMAP_MAX_RETRIES, e)
                return []

    log.info("track: fetched %d candidate emails since %s", len(emails), since_str)
    return emails


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def _clean_company(raw: str) -> str | None:
    """Normalize and validate a raw company string extracted from subject or body.

    Returns None if the extracted string looks like a job title rather than a company name.
    """
    company = raw.strip().rstrip("!").strip()
    company = re.sub(r",\s+[A-Z][a-z]+$", "", company)       # trailing ", FirstName"
    company = _COMPANY_TRAILING_NOISE.sub("", company)         # trailing "- Candidate Name"
    company = _COMPANY_LEADING_THE.sub("", company)            # leading "the "
    company = _COMPANY_SENDER_SUFFIX.sub("", company)          # trailing "Recruiting", "Talent"
    company = company.strip()
    if not company or len(company) < 2:
        return None
    # If the string contains colon or looks like a role title, discard it — we caught too much
    if ":" in company:
        company = company.split(":")[0].strip()
    if _ROLE_TITLE_WORDS.search(company) and len(company.split()) >= 2:
        return None
    return company


def extract_company(subject: str, body: str) -> str | None:
    """Extract company name from subject (primary) or body (fallback)."""
    # Normalize line-folds in subject (MIME headers can wrap mid-phrase)
    subject = " ".join(subject.split())

    # ATS-specific patterns that would otherwise trip up generic patterns
    # "{COMPANY} Job Application Update: {ID} {ROLE}" (Workday / GE HealthCare)
    m = re.search(r"^([A-Za-z0-9][^:]+?) job application update", subject, re.IGNORECASE)
    if m:
        company = _clean_company(m.group(1))
        if company:
            return company

    for pattern in _COMPANY_FROM_SUBJECT:
        m = pattern.search(subject)
        if m:
            company = _clean_company(m.group(1))
            if company and len(company) > 1:
                return company

    for pattern in _COMPANY_FROM_BODY:
        m = pattern.search(body)
        if m:
            result = _clean_company(m.group(1))
            if result:
                return result

    return None


_ROLE_BOILERPLATE = re.compile(
    r"if you are not|keep an eye|we will contact|please continue|"
    r"for this role|for this position|you are not selected|"
    r"time to apply for our|taking the time",
    re.IGNORECASE,
)


def _clean_role(raw: str) -> str | None:
    """Normalize and validate a raw role string."""
    role = raw.strip()
    # Drop ATS job IDs: "(ID: 10443018)", "JR2018175 Senior...", "200022543"
    role = re.sub(r"\s*[\(\[]?(?:ID|JR|REQ)[:\s]\S+[\)\]]?", "", role, flags=re.IGNORECASE)
    role = re.sub(r"^\w{2,}\d{6,}\s+", "", role)   # leading alphanumeric job codes
    role = re.sub(r"\s+\d{6,}$", "", role)           # trailing numeric job IDs
    role = role.strip().rstrip(".,")
    if len(role) < 4 or len(role) > _ROLE_MAX_LEN:
        return None
    if _ROLE_BOILERPLATE.search(role):
        return None
    return role


def extract_role(subject: str, body: str) -> str | None:
    """Extract role title from body (primary) or subject (fallback)."""
    subject = " ".join(subject.split())  # normalize line-folds
    for pattern in _ROLE_FROM_BODY:
        m = pattern.search(body)
        if m:
            role = _clean_role(m.group(1))
            if role:
                return role

    for pattern in _ROLE_FROM_SUBJECT:
        m = pattern.search(subject)
        if m:
            role = _clean_role(m.group(1))
            if role:
                return role

    return None


_ROLE_EXTRACT_PROMPT = """\
You are extracting a job role title from an application confirmation email.

Subject: {subject}

Email body (first 1500 chars):
{body}

Reply with ONLY the job title/role name the candidate applied for.
Examples of valid replies: "Staff Product Manager", "Principal PM – AI Platform", "Senior Software Engineer"
If no specific role title is mentioned anywhere in the email, reply with exactly: NONE
No punctuation, no explanation, no quotes."""


def _extract_role_llm(subject: str, body: str, client) -> str | None:
    """Haiku fallback when regex role extraction returns nothing."""
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=30,
            messages=[{"role": "user", "content": _ROLE_EXTRACT_PROMPT.format(
                subject=subject,
                body=body[:1500],
            )}],
        )
        result = resp.content[0].text.strip()
        if result.upper() == "NONE" or not result:
            return None
        return _clean_role(result)
    except Exception as exc:
        log.debug("track: LLM role extraction failed (%s)", exc)
        return None


def parse_email(email_dict: dict, llm_client=None) -> dict:
    """Classify and extract structured fields from a candidate email."""
    subject = email_dict["subject"]
    body    = email_dict["body"]

    classification = classify_email(body, subject, llm_client=llm_client)
    company = extract_company(subject, body)

    # Sender domain fallback: "no-reply@databricks.com" → "Databricks"
    # Only used when subject/body extraction failed, and domain isn't a generic ESP
    if not company:
        company = _company_from_sender(email_dict.get("sender", ""))

    role = extract_role(subject, body)
    # Claude fallback when regex comes up empty — many ATS templates don't use
    # standard "applying for the X role" phrasing but do mention the title somewhere
    if role is None and llm_client is not None:
        role = _extract_role_llm(subject, body, llm_client)
    log.debug("track: parse  [%s] company=%r | %s", classification[:4], company, subject[:70])

    return {
        "classification": classification,
        "company":        company,
        "role":           role,
        "date":           email_dict["date"],
        "subject":        subject,
        "sender":         email_dict["sender"],
    }


_GENERIC_SENDER_DOMAINS = {
    # Email providers
    "gmail", "googlemail", "yahoo", "hotmail", "outlook", "icloud",
    # ATS platforms — emails come "from" these but the company is in the display name
    "ashbyhq", "ashby", "greenhouse-mail", "greenhouse",
    "lever", "smartrecruiters", "icims", "jobvite", "taleo",
    "bamboohr", "myworkdayjobs", "myworkday", "successfactors", "applytojob",
    "talentplatform", "workday",
    # Generic sender names
    "notifications", "noreply", "no-reply", "mailer", "bounce", "mail",
}

# ATS platform domains to search by sender — catches ATS emails regardless of subject.
# These are the @domain portions used in FROM IMAP searches.
_ATS_FROM_DOMAINS = [
    "greenhouse.io", "greenhouse-mail.io",
    "lever.co",
    "myworkdayjobs.com",
    "smartrecruiters.com",
    "ashbyhq.com",
    "icims.com",
    "jobvite.com",
    "bamboohr.com",
    "workday.com",
    "applytojob.com",
    "successfactors.com",
    "taleo.net",
]


def _normalize_company_search(company: str) -> str:
    """Normalize a company name for use in an IMAP FROM substring search.

    Strips legal suffixes, punctuation, and excess whitespace so that
    'Sigma Computing, Inc.' becomes 'sigma computing' and matches
    any sender whose address or display name contains that string.
    """
    term = re.sub(r",?\s+(?:inc|llc|corp|ltd|co)\.?$", "", company, flags=re.IGNORECASE)
    term = re.sub(r"[^\w\s-]", "", term).strip().lower()
    return term

def _company_from_sender(sender: str) -> str | None:
    """Extract company name from sender display name or email domain.

    Tries in order:
    1. Display name pattern: "Console Hiring Team" → "Console"
    2. Email domain: "careers@databricks.com" → "Databricks"
    Ignores generic ATS and ESP domains.
    """
    # 1. Display name: "Console Hiring Team <...>" or "EvenUp Talent Team <...>"
    display_m = re.match(r'^"?([^"<]+?)"?\s*(?:<|$)', sender)
    if display_m:
        display = display_m.group(1).strip()
        # Strip trailing " Hiring Team", " Talent Team", " Careers", " Recruiting", etc.
        company = re.sub(
            r'\s+(?:hiring team|talent team|talent acquisition|recruiting team|'
            r'careers|recruiter|hr|talent|@ \w+)$',
            "", display, flags=re.IGNORECASE
        ).strip()
        # Strip trailing legal suffixes and punctuation
        company = re.sub(r',?\s+(?:inc|llc|corp|ltd)\.?$', "", company, flags=re.IGNORECASE).strip()
        # Reject personal names: "Laura Otto", "John Smith" (2 title-case words, no digits)
        words = company.split()
        is_personal_name = (
            len(words) == 2
            and all(w[0].isupper() and w[1:].islower() and w.isalpha() for w in words)
        )
        if (company
                and len(company) > 2
                and not is_personal_name
                and not re.search(r'\b(?:noreply|no.reply|mailer|notifications|do.not.reply)\b',
                                  company, re.IGNORECASE)
                and "@" not in company):
            return company

    # 2. Email domain fallback
    m = re.search(r"@([a-zA-Z0-9][-a-zA-Z0-9.]*\.[a-zA-Z]{2,})", sender)
    if not m:
        return None
    domain = m.group(1).lower()
    parts = domain.split(".")
    if len(parts) >= 2:
        slug = parts[-2]
    else:
        return None
    if slug in _GENERIC_SENDER_DOMAINS or len(slug) <= 2:
        return None
    return slug.title()


# ---------------------------------------------------------------------------
# Fuzzy matching against tracker rows
# ---------------------------------------------------------------------------

def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _similarity(a: str, b: str) -> float:
    from difflib import SequenceMatcher
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


def find_tracker_row(ws, company: str, role: str | None) -> int | None:
    """Return 1-based row index of the best-matching row, or None.

    Matching strategy:
      1. Company must match above COMPANY_THRESHOLD (fuzzy)
      2. If role is provided, title must match above ROLE_THRESHOLD (fuzzy)
      3. Return the highest-scoring match
    """
    COMPANY_THRESHOLD = 0.70
    ROLE_THRESHOLD    = 0.55   # lower because ATS titles drift from LinkedIn titles

    best_row   = None
    best_score = 0.0

    for row_idx in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row_idx, 3).value or "")
        row_title   = str(ws.cell(row_idx, 2).value or "")

        company_score = _similarity(company, row_company)
        if company_score < COMPANY_THRESHOLD:
            continue

        if role:
            role_score = _similarity(role, row_title)
            if role_score < ROLE_THRESHOLD:
                continue
            combined = (company_score + role_score) / 2
        else:
            combined = company_score

        if combined > best_score:
            best_score = combined
            best_row   = row_idx

    return best_row


# ---------------------------------------------------------------------------
# Tracker updates
# ---------------------------------------------------------------------------

def _apply_status_fill(ws, row_idx: int, col: int, value: str) -> None:
    from openpyxl.styles import PatternFill
    _STATUS_FILL = {
        "Applied":          "C6EFCE",
        "Not Applied":      "D9D9D9",
        "Pending":          "FFEB9C",
        "Proceeding":       "C6EFCE",
        "Rejected":         "FFC7CE",
        "Limited Match":    "D9D9D9",
        # "External" intentionally omitted — no fill (white/blank)
        "Recruiter Screen": "BDD7EE",  # light blue
    }
    color = _STATUS_FILL.get(value)
    if color:
        ws.cell(row_idx, col).fill = PatternFill(fill_type="solid", fgColor=color)


def update_confirmation(ws, row_idx: int, date_applied: str) -> None:
    """Flip action_taken → Applied and set date_applied + application_status."""
    ws.cell(row_idx, 6).value = "Applied"
    ws.cell(row_idx, 7).value = date_applied
    ws.cell(row_idx, 8).value = "Pending"
    _apply_status_fill(ws, row_idx, 6, "Applied")
    _apply_status_fill(ws, row_idx, 8, "Pending")


def update_rejection(ws, row_idx: int) -> None:
    """Set application_status → Rejected."""
    ws.cell(row_idx, 8).value = "Rejected"
    _apply_status_fill(ws, row_idx, 8, "Rejected")


def update_recruiter_screen(ws, row_idx: int) -> None:
    """Set action_taken → Applied (if not already) and application_status → Recruiter Screen."""
    if ws.cell(row_idx, 6).value != "Applied":
        ws.cell(row_idx, 6).value = "Applied"
        _apply_status_fill(ws, row_idx, 6, "Applied")
    ws.cell(row_idx, 8).value = "Recruiter Screen"
    _apply_status_fill(ws, row_idx, 8, "Recruiter Screen")


def _write_row_from_email(ws, parsed: dict, suggestion_status: str,
                          date_suggested: str | None = None,
                          match_score: float | None = None,
                          url: str = "") -> None:
    """Shared helper: append one row to ws from parsed email data."""
    from openpyxl.styles import Alignment
    from .xlsx import _set_hyperlink

    next_row = ws.max_row + 1
    values = [
        date_suggested or parsed["date"],
        parsed.get("role") or "",
        parsed.get("company") or "",
        match_score,
        suggestion_status,
        "Applied",
        parsed["date"],
        "Pending",
        None,
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(next_row, col_idx, value).alignment = Alignment(vertical="top")

    if url:
        _set_hyperlink(ws.cell(next_row, 2), url, values[1])

    ws.cell(next_row, 4).number_format = "0%"
    for col, val in [(5, suggestion_status), (6, "Applied"), (8, "Pending")]:
        _apply_status_fill(ws, next_row, col, val)


def create_skipped_row(ws, parsed: dict, skipped_meta: dict) -> None:
    """Create a new row for a skipped role the user applied to anyway."""
    score = skipped_meta.get("match_score")
    _write_row_from_email(
        ws, parsed,
        suggestion_status="Limited Match",
        date_suggested=skipped_meta.get("date_suggested"),
        match_score=(score / 100.0) if score else None,
        url=skipped_meta.get("url", ""),
    )
    # Prefer stored role title / company over email-extracted ones
    next_row = ws.max_row
    if skipped_meta.get("role_title"):
        ws.cell(next_row, 2).value = skipped_meta["role_title"]
    if skipped_meta.get("company"):
        ws.cell(next_row, 3).value = skipped_meta["company"]


def create_backfill_row(ws, parsed: dict) -> None:
    """Create a row for a confirmed application with no prior tracker entry.

    Used for roles applied to before the tracker existed, or applied to outside
    the metis digest. suggestion_status='External' marks these as applied outside
    the digest flow. date_suggested is left blank.
    """
    _write_row_from_email(ws, parsed, suggestion_status="External")


# ---------------------------------------------------------------------------
# Digest backfill — parse past "Personalized Job Alert Digest" emails
# ---------------------------------------------------------------------------

def _parse_digest_html(html: str, email_date: str) -> list[dict]:
    """Extract Apply/Consider job rows from a rendered digest HTML email.

    Returns a list of job dicts compatible with write_to_tracker():
    {'title', 'company', 'url', 'eval': {'score', 'verdict'}}
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")

    # ── Prefer embedded JSON data island (added to future digests) ──────────
    tag = soup.find("script", {"type": "application/json", "id": "metis-data"})
    if tag and tag.string:
        import json as _json
        try:
            payload = _json.loads(tag.string)
            jobs = []
            for j in payload.get("jobs", []):
                verdict = j.get("verdict", "").lower()
                if verdict not in ("apply", "consider"):
                    continue
                jobs.append({
                    "title":   j.get("title", ""),
                    "company": j.get("company", ""),
                    "url":     j.get("postingUrl", ""),
                    "eval":    {"score": j.get("score", 0), "verdict": verdict},
                })
            return jobs
        except Exception:
            pass  # fall through to HTML parsing

    # ── HTML fallback: parse job cards by "View posting" links ──────────────
    # React Email renders: title in <td font-size:15px font-weight:500>,
    # company in <p color:#72716d>, score in <span border-radius:20px>.
    # Python fallback uses slightly different colors but same structure.
    # Both use "View posting" anchor text to identify job cards.

    def _nstyle(tag) -> str:
        """Normalize a tag's style for robust substring matching."""
        return re.sub(r"\s+", "", (tag.get("style") or "").lower())

    # Known muted colors across React Email and Python fallback renderers.
    _MUTED = ("72716d", "888780", "726f6a", "6b6b6b", "757575", "666666")

    jobs = []
    for anchor in soup.find_all("a"):
        if "View posting" not in anchor.get_text():
            continue
        url = anchor.get("href", "")
        # Walk up to find the enclosing card table — the card has border-radius:8px.
        # find_parent("table") would grab the inner footer table, not the card.
        card = anchor.find_parent(
            "table",
            style=lambda s: s and "border-radius:8px" in s.replace(" ", "")
        )
        if not card:
            continue

        # Title: font-size 15px with bold/500/600 weight.
        # _nstyle() strips spaces so "font-size: 15px" and "font-size:15px" both match.
        # Expanded tag names cover <td>, <div>, <p>, <span> across renderer variants.
        title_td = card.find(
            lambda t: t.name in ("td", "div", "p", "span")
            and "15px" in _nstyle(t)
            and any(w in _nstyle(t) for w in ("font-weight:500", "font-weight:600",
                                               "font-weight:bold"))
        )
        title = title_td.get_text(strip=True) if title_td else ""

        # Company+location: muted color element (primary), or "·" separator (fallback).
        # Primary handles both React Email and Python fallback color palettes.
        # Fallback catches any renderer variant where the color doesn't match exactly.
        co_tag = card.find(
            lambda t: t.name in ("p", "div", "td", "span")
            and any(c in _nstyle(t) for c in _MUTED)
            and t is not title_td
        )
        if not co_tag or not co_tag.get_text(strip=True):
            # Structural fallback: find a leaf-ish element containing "·".
            # Two guards prevent matching a card container:
            #   - t not in title_td.parents → skip ancestors of the title element
            #   - len < 120 → skip container elements whose text spans the whole card
            co_tag = card.find(
                lambda t: t.name in ("p", "div", "td", "span")
                and "·" in t.get_text()
                and t is not title_td
                and (title_td is None or t not in title_td.parents)
                and len(t.get_text(strip=True)) < 120
            )
        company_raw = co_tag.get_text(strip=True) if co_tag else ""
        company = company_raw.split("·")[0].strip()

        # Score: span with border-radius:20px containing "\d+%"
        score_span = card.find(
            lambda t: t.name == "span"
            and "border-radius:20px" in (t.get("style") or "").replace(" ", "")
            and re.search(r"\d+%", t.get_text())
        )
        score = 0
        if score_span:
            m = re.search(r"(\d+)%", score_span.get_text())
            if m:
                score = int(m.group(1))

        # Verdict: from pill background color
        # React Email:    apply=#eef2ee  consider=#f4f0e8
        # Python fallback: apply=#EAF3DE consider=#FAEEDA
        style = (score_span.get("style", "") if score_span else "").lower()
        if "eef2ee" in style or "eaf3de" in style:
            verdict = "apply"
        elif "f4f0e8" in style or "faeeda" in style:
            verdict = "consider"
        else:
            verdict = "consider"   # safe default

        if title and company:
            jobs.append({
                "title":   title,
                "company": company,
                "url":     url,
                "eval":    {"score": score, "verdict": verdict},
            })

    return jobs


def backfill_from_digests(
    gmail_address: str,
    app_password: str,
    since_dt: datetime.datetime,
) -> int:
    """Fetch past digest emails from Gmail and write any new rows to the tracker.

    Only writes Apply/Consider roles not already in the tracker (dedup by
    normalized title+company). Returns the number of new rows added.
    """
    from email.utils import parsedate_to_datetime
    from .xlsx import write_to_tracker

    since_str = since_dt.strftime("%d-%b-%Y")
    added_total = 0
    n_digests = 0

    for attempt in range(1, _IMAP_MAX_RETRIES + 1):
        try:
            with imaplib.IMAP4_SSL("imap.gmail.com") as imap:
                imap.login(gmail_address, app_password)
                imap.select("INBOX")

                _, data = imap.search(None, f'SINCE {since_str} SUBJECT "Metis Digest"')
                if not data or not data[0]:
                    log.info("track: no digest emails found since %s", since_str)
                    return 0

                digest_ids = data[0].split()
                n_digests = len(digest_ids)
                log.info("track: found %d digest email(s) to backfill from", n_digests)

                for msg_id in digest_ids:
                    _, raw = imap.fetch(msg_id, "(RFC822)")
                    if not raw or not raw[0]:
                        continue
                    raw_bytes = raw[0][1] if isinstance(raw[0], tuple) else raw[0]
                    msg = email_lib.message_from_bytes(raw_bytes)

                    date_header = msg.get("Date", "")
                    try:
                        email_date = parsedate_to_datetime(date_header).date().isoformat()
                    except Exception:
                        email_date = datetime.date.today().isoformat()

                    # Extract HTML body from digest email
                    html_body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/html":
                                payload = part.get_payload(decode=True)
                                if payload:
                                    html_body = _safe_decode(payload, part.get_content_charset())
                                    break
                    else:
                        if msg.get_content_type() == "text/html":
                            payload = msg.get_payload(decode=True)
                            if payload:
                                html_body = _safe_decode(payload, msg.get_content_charset())

                    if not html_body:
                        log.debug("track: digest %s has no HTML body — skipping", email_date)
                        continue

                    jobs = _parse_digest_html(html_body, email_date)
                    if jobs:
                        log.info("track: digest %s → %d role(s) parsed", email_date, len(jobs))
                        write_to_tracker(jobs, run_date=email_date)
                        added_total += len(jobs)
                    else:
                        log.warning("track: digest %s → 0 roles parsed (HTML structure mismatch?)", email_date)
            break   # success
        except OSError as e:
            if attempt < _IMAP_MAX_RETRIES:
                log.warning(
                    "track: IMAP connect failed (attempt %d/%d): %s — retrying in %ds…",
                    attempt, _IMAP_MAX_RETRIES, e, _IMAP_RETRY_DELAY,
                )
                time.sleep(_IMAP_RETRY_DELAY)
            else:
                log.error("track: IMAP connect failed after %d attempts: %s — skipping backfill",
                          _IMAP_MAX_RETRIES, e)
                return 0

    log.info("track: digest backfill complete — %d role(s) eligible across %d digest(s)",
             added_total, n_digests)
    return added_total


# ---------------------------------------------------------------------------
# LLM client factory
# ---------------------------------------------------------------------------

def _build_llm_client(api_key: str | None):
    """Return an Anthropic client if profile.yaml has track.llm_fallback: true.

    Returns None when the flag is absent, false, or the api_key is missing —
    so classify_email() runs phrase-only (safe for OSS users with no key).
    """
    if not api_key:
        return None
    try:
        from .profile import load_profile_yaml
        profile = load_profile_yaml() or {}
    except Exception:
        return None
    if not profile.get("track", {}).get("llm_fallback", False):
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        log.info("track: LLM fallback enabled (Haiku) for unknown emails")
        return client
    except Exception as exc:
        log.warning("track: could not build LLM client (%s) — running phrase-only", exc)
        return None


# ---------------------------------------------------------------------------
# Digest backfill
# ---------------------------------------------------------------------------

def backfill_from_digests(
    gmail_address: str,
    app_password: str,
    since_dt: datetime.datetime,
) -> int:
    """Fetch past digest emails from Gmail and write any new rows to the tracker.

    Only writes Apply/Consider roles not already in the tracker (dedup by
    normalized title+company). Returns the number of new rows added.
    """
    from email.utils import parsedate_to_datetime
    from .xlsx import write_to_tracker

    since_str    = since_dt.strftime("%d-%b-%Y")
    added_total  = 0
    n_digests    = 0

    for attempt in range(1, _IMAP_MAX_RETRIES + 1):
        try:
            with imaplib.IMAP4_SSL("imap.gmail.com") as imap:
                imap.login(gmail_address, app_password)
                imap.select("INBOX")

                _, data = imap.search(None, f'SINCE {since_str} SUBJECT "Personalized Job Alert Digest"')
                if not data or not data[0]:
                    log.info("track: no digest emails found since %s", since_str)
                    return 0

                digest_ids = data[0].split()
                n_digests  = len(digest_ids)
                log.info("track: found %d digest email(s) to backfill from", n_digests)

                for msg_id in digest_ids:
                    _, raw = imap.fetch(msg_id, "(RFC822)")
                    if not raw or not raw[0]:
                        continue
                    raw_bytes = raw[0][1] if isinstance(raw[0], tuple) else raw[0]
                    msg = email_lib.message_from_bytes(raw_bytes)

                    date_header = msg.get("Date", "")
                    try:
                        email_date = parsedate_to_datetime(date_header).date().isoformat()
                    except Exception:
                        email_date = datetime.date.today().isoformat()

                    html_body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/html":
                                payload = part.get_payload(decode=True)
                                if payload:
                                    html_body = _safe_decode(payload, part.get_content_charset())
                                    break
                    else:
                        if msg.get_content_type() == "text/html":
                            payload = msg.get_payload(decode=True)
                            if payload:
                                html_body = _safe_decode(payload, msg.get_content_charset())

                    if not html_body:
                        log.debug("track: digest %s has no HTML body — skipping", email_date)
                        continue

                    jobs = _parse_digest_html(html_body, email_date)
                    if jobs:
                        log.info("track: digest %s → %d role(s) parsed", email_date, len(jobs))
                        write_to_tracker(jobs, run_date=email_date)
                        added_total += len(jobs)
                    else:
                        log.warning("track: digest %s → 0 roles parsed (HTML structure mismatch?)", email_date)
            break
        except OSError as e:
            if attempt < _IMAP_MAX_RETRIES:
                log.warning(
                    "track: IMAP connect failed (attempt %d/%d): %s — retrying in %ds…",
                    attempt, _IMAP_MAX_RETRIES, e, _IMAP_RETRY_DELAY,
                )
                time.sleep(_IMAP_RETRY_DELAY)
            else:
                log.error("track: IMAP connect failed after %d attempts: %s — skipping backfill",
                          _IMAP_MAX_RETRIES, e)
                return 0

    log.info("track: digest backfill complete — %d role(s) eligible across %d digest(s)",
             added_total, n_digests)
    return added_total


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_track(
    gmail_address: str,
    app_password: str,
    since_dt: datetime.datetime,
    dry_run: bool = False,
    api_key: str | None = None,
) -> None:
    """Parse confirmation/rejection emails and update the Applications tracker."""
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    from .xlsx  import TRACKER_PATH
    from .state import lookup_skipped_role, promote_skipped_role

    llm_client = _build_llm_client(api_key)

    log.info("track: step 1 — backfilling tracker from digest emails…")
    backfill_from_digests(gmail_address, app_password, since_dt)

    applied_companies: set[str] = set()
    if TRACKER_PATH.exists():
        _wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
        _ws = _wb.active
        for _r in range(2, _ws.max_row + 1):
            if _ws.cell(_r, 6).value == "Applied":
                co = _ws.cell(_r, 3).value
                if co:
                    applied_companies.add(str(co))
        _wb.close()
    log.info("track: %d applied companies loaded for direct-outreach search", len(applied_companies))

    emails = fetch_candidate_emails(gmail_address, app_password, since_dt, applied_companies=applied_companies)
    if not emails:
        log.info("track: no candidate emails found in lookback window.")
        return

    parsed_emails = [parse_email(e, llm_client=llm_client) for e in emails]

    actionable = [p for p in parsed_emails
                  if p["classification"] in ("confirmation", "rejection", "recruiter_screen")
                  and p["company"]]
    unknown    = [p for p in parsed_emails if p["classification"] == "unknown"]

    log.info("track: %d actionable (%d confirmation, %d rejection, %d recruiter_screen), %d unknown",
             len(actionable),
             sum(1 for p in actionable if p["classification"] == "confirmation"),
             sum(1 for p in actionable if p["classification"] == "rejection"),
             sum(1 for p in actionable if p["classification"] == "recruiter_screen"),
             len(unknown))

    if unknown:
        for u in unknown:
            log.warning("track: unclassified — '%s' (company=%s)", u["subject"], u["company"])

    if dry_run:
        for p in parsed_emails:
            print(f"[{p['classification'].upper():12}] {p['company'] or '?':25} | {p['role'] or '?'}")
        return

    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        from .xlsx import _write_header, _set_column_widths
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        _write_header(ws)
        _set_column_widths(ws)
        ws.freeze_panes = "A2"
        log.info("track: created new tracker at %s", TRACKER_PATH)

    changed = 0

    for parsed in actionable:
        company = parsed["company"]
        role    = parsed["role"]
        kind    = parsed["classification"]

        row_idx = find_tracker_row(ws, company, role)

        if row_idx is None:
            if kind == "confirmation":
                skipped_meta = lookup_skipped_role(company, role or "")
                if skipped_meta:
                    log.info("track: backfill (skipped) — %s / %s", company, role)
                    create_skipped_row(ws, parsed, skipped_meta)
                    promote_skipped_role(company, role or "")
                else:
                    log.info("track: backfill (external) — %s / %s", company, role or "?")
                    create_backfill_row(ws, parsed)
                changed += 1
            else:
                log.info("track: skip rejection (no tracker row) — %s", company)
            continue

        current_action = ws.cell(row_idx, 6).value
        current_status = ws.cell(row_idx, 8).value

        if kind == "confirmation" and current_action != "Applied":
            log.info("track: ✓ confirmation — %s / %s → Applied + Pending", company, role or "?")
            update_confirmation(ws, row_idx, parsed["date"])
            changed += 1
        elif kind == "confirmation" and current_action == "Applied":
            log.info("track: skip confirmation (already Applied) — %s", company)
        elif kind == "rejection" and current_status != "Rejected":
            log.info("track: ✗ rejection  — %s / %s → Rejected", company, role or "?")
            update_rejection(ws, row_idx)
            changed += 1
        elif kind == "rejection" and current_status == "Rejected":
            log.info("track: skip rejection (already Rejected) — %s", company)
        elif kind == "recruiter_screen" and current_status not in ("Recruiter Screen", "Rejected"):
            log.info("track: ★ recruiter screen — %s / %s → Recruiter Screen", company, role or "?")
            update_recruiter_screen(ws, row_idx)
            changed += 1
        elif kind == "recruiter_screen":
            log.info("track: skip recruiter screen (already %s) — %s", current_status, company)

    from .xlsx import _sort_rows_by_date
    _sort_rows_by_date(ws)

    if changed:
        wb.save(TRACKER_PATH)
        TRACKER_PATH.chmod(0o600)
        log.info("track: updated %d row(s) in %s", changed, TRACKER_PATH)
        print(f"  Tracker → {TRACKER_PATH}")
        if sys.stdout.isatty():
            import subprocess
            subprocess.Popen(["open", str(TRACKER_PATH)])
    else:
        wb.save(TRACKER_PATH)
        TRACKER_PATH.chmod(0o600)
        log.info("track: no updates needed.")
