"""metis summary — generate and deliver the Metis market report.

Usage:
    metis summary                      # generate + send to email
    metis summary --lookback 60d       # scope market intel to 60 days (default: 30d)
    metis summary --output report.html # save as HTML instead of sending
    metis summary --output report.pdf  # save as PDF instead of sending
"""

from __future__ import annotations

import datetime
import math
import os
import smtplib
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import json
import re
from collections import Counter, defaultdict

import openpyxl

from metis.state import DATA_DIR

# ---------------------------------------------------------------------------
# Design tokens — Ink & Slate palette (matches Job Alert Digest)
# ---------------------------------------------------------------------------
FONT      = "-apple-system, BlinkMacSystemFont, 'Segoe UI', 'Helvetica Neue', Arial, sans-serif"
C_HEADING = "#1f2118"
C_MUTED   = "#72716d"
C_SUBTLE  = "#aaaaaa"
APPLY_BG  = "#eef2ee"; APPLY_NUM  = "#2d5a2d"
CONSID_BG = "#faeeda"; CONSID_NUM = "#854f0b"
TOTAL_BG  = "#f5f5f3"; TOTAL_NUM  = "#1f2118"
RED_BG    = "#f2eeee"; RED_NUM    = "#8b2e2e"
CARD_BG   = "#ffffff"
BORDER    = "#eeece5"
BODY_BG   = "#f0f0ef"

# suggestion_status values — support both old and new naming
_SOLID    = {"Apply", "Solid", "Solid Match"}
_MODERATE = {"Consider", "Moderate", "Moderate Match"}
_PARTIAL  = {"Skipped", "Partial", "Partial Match"}
_SCORED   = _SOLID | _MODERATE | _PARTIAL


# ---------------------------------------------------------------------------
# Data layer
# ---------------------------------------------------------------------------

def _coerce_date(val) -> datetime.date | None:
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        try:
            return datetime.date.fromisoformat(str(val)[:10])
        except ValueError:
            return None
    return None


_SKIPPED_PATH = DATA_DIR / "skipped_roles.json"


def load_report_data(tracker_path: Path) -> dict:
    """Read applications.xlsx and compute all metrics for the report."""
    if not tracker_path.exists():
        return {
            "roles_scored": 0, "total_applied": 0, "time_saved_h": 0.0,
            "alignment": 0.0, "TP": 0, "TN": 0, "align_base": 0,
            "solid": 0, "moderate": 0, "partial": 0,
            "pending": 0, "screens": 0, "rejections": 0, "daily": {},
        }
    wb = openpyxl.load_workbook(tracker_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = [
        {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        for r in range(2, ws.max_row + 1)
    ]
    wb.close()

    scored     = [r for r in rows if r.get("suggestion_status") in _SCORED]
    solid_rows = [r for r in scored if r.get("suggestion_status") in _SOLID]
    mod_rows   = [r for r in scored if r.get("suggestion_status") in _MODERATE]
    part_rows  = [r for r in scored if r.get("suggestion_status") in _PARTIAL]
    applied    = [r for r in rows if r.get("action_taken") == "Applied"]

    pending    = [r for r in applied if r.get("application_status") == "Pending"]
    screens    = [r for r in applied if r.get("application_status") == "Recruiter Screen"]
    rejections = [r for r in applied if r.get("application_status") == "Rejected"]

    # Daily application volume
    daily: dict[datetime.date, int] = defaultdict(int)
    for r in applied:
        d = _coerce_date(r.get("date_applied"))
        if d:
            daily[d] += 1

    # Skipped count: skipped_roles.json is the authoritative source —
    # skipped roles are never written to xlsx so xlsx always shows 0.
    skipped_count = 0
    if _SKIPPED_PATH.exists():
        try:
            skipped_count = len(json.loads(_SKIPPED_PATH.read_text()))
        except Exception:
            pass

    total_scored = len(solid_rows) + len(mod_rows) + skipped_count

    # Alignment: what fraction of metis's scoring calls turned out to be correct?
    #   TP = recommended (Solid or Moderate) and user applied
    #   TN = not recommended (Partial/skipped) and user did not apply
    #       skipped roles are always not-applied by definition
    #   base = all scored roles (Solid + Moderate + Partial + skipped)
    recommended = solid_rows + mod_rows
    TP = sum(1 for r in recommended if r.get("action_taken") == "Applied")
    TN = skipped_count + sum(1 for r in part_rows if r.get("action_taken") != "Applied")
    base      = len(recommended) + len(part_rows) + skipped_count
    alignment = (TP + TN) / base if base else 0.0

    return {
        "roles_scored":  total_scored,
        "total_applied": len(applied),
        "time_saved_h":  round(total_scored / 60, 1),
        "alignment":     alignment,
        "TP": TP, "TN": TN, "align_base": base,
        "solid":    len(solid_rows),
        "moderate": len(mod_rows),
        "partial":  skipped_count,
        "pending":    len(pending),
        "screens":    len(screens),
        "rejections": len(rejections),
        "daily": dict(daily),
    }


# ---------------------------------------------------------------------------
# HTML rendering helpers
# ---------------------------------------------------------------------------

def _pct(n: int, total: int) -> str:
    if not total:
        return "—"
    return f"{n} / {round(n / total * 100)}%"


def _section_h(t: str) -> str:
    return (
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:6px;">'
        f'<tr><td style="font-family:{FONT};font-size:14px;font-weight:600;color:{C_HEADING};'
        f'padding-bottom:6px;border-bottom:1px solid {BORDER};">{t}</td></tr></table>'
    )


def _sub_h(t: str) -> str:
    return (
        f'<p style="font-family:{FONT};font-size:10px;font-weight:600;text-transform:uppercase;'
        f'letter-spacing:0.05em;color:{C_SUBTLE};margin:0 0 8px 0;">{t}</p>'
    )


def _spacer(h: int = 20) -> str:
    return f'<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td height="{h}"></td></tr></table>'


def _tiles(items: list) -> str:
    cells = ""
    for label, num, sub, bg, color in items:
        sub_html = (
            f'<div style="font-family:{FONT};font-size:10px;color:{color};margin-top:2px;opacity:0.6;">{sub}</div>'
            if sub else ""
        )
        cells += (
            f'<td width="25%" style="padding:0 4px;" valign="top">'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="background:{bg};border-radius:6px;table-layout:fixed;">'
            f'<tr><td style="padding:10px 10px 8px;text-align:center;">'
            f'<div style="font-family:{FONT};font-size:28px;font-weight:700;color:{color};line-height:1.1;">{num}</div>'
            f'<div style="font-family:{FONT};font-size:11px;font-weight:500;color:{color};margin-top:3px;opacity:0.75;">{label}</div>'
            f'{sub_html}'
            f'</td></tr></table></td>'
        )
    return (
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        f'style="table-layout:fixed;border-collapse:separate;border-spacing:0;">'
        f'<tr>{cells}</tr></table>'
    )


def _green_th(cols: list, aligns: list | None = None) -> str:
    if aligns is None:
        aligns = ["left"] * len(cols)
    cells = "".join(
        f'<td style="font-family:{FONT};font-size:12px;font-weight:600;color:{APPLY_NUM};'
        f'background:{APPLY_BG};padding:8px 10px;text-transform:uppercase;'
        f'letter-spacing:0.04em;text-align:{a};">{c}</td>'
        for c, a in zip(cols, aligns)
    )
    return f"<tr>{cells}</tr>"


def _td(content: str, width: str | None = None, align: str = "left",
        muted: bool = False, size: int = 13) -> str:
    color = C_MUTED if muted else C_HEADING
    w = f'width="{width}"' if width else ""
    return (
        f'<td {w} style="font-family:{FONT};font-size:{size}px;font-weight:400;color:{color};'
        f'padding:7px 10px;line-height:1.4;text-align:{align};border-bottom:1px solid {BORDER};">'
        f'{content}</td>'
    )


def _fit_pill(label: str) -> str:
    if "Solid" in label or "Strong" in label:
        return f'<span style="background:#eef2ee;color:#2d5a2d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Solid</span>'
    if "Moderate" in label or "Partial" in label:
        return f'<span style="background:#faeeda;color:#854f0b;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Moderate</span>'
    return f'<span style="background:#f2eeee;color:#8b2e2e;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Gap</span>'


def _trend_pill(label: str) -> str:
    if "Trending" in label:
        return f'<span style="background:#eef2ee;color:#2d5a2d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Trending ↑</span>'
    if "Niche" in label:
        return f'<span style="background:#f5f5f3;color:#72716d;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Niche</span>'
    return f'<span style="background:#e8f0f8;color:#185FA5;font-size:10px;font-weight:500;padding:2px 7px;border-radius:10px;">Established</span>'


def _bar_chart(daily: dict) -> str:
    if not daily:
        return '<p style="font-family:{FONT};font-size:12px;color:{C_SUBTLE};">No application data.</p>'

    start_d = min(daily.keys())
    end_d   = max(daily.keys())
    all_days: list[datetime.date] = []
    d = start_d
    while d <= end_d:
        all_days.append(d)
        d += datetime.timedelta(days=1)

    max_v = max(daily.values(), default=1)
    H     = 8
    col_w = 15

    bar_rows = ""
    for level in range(H, 0, -1):
        cells = ""
        for day in all_days:
            val = daily.get(day, 0)
            bh  = max(1, math.ceil(val / max_v * H)) if val > 0 else 0
            filled   = val > 0 and level <= bh
            top_cell = val > 0 and level == bh
            bg     = APPLY_NUM if filled else "transparent"
            radius = "border-radius:2px 2px 0 0;" if top_cell else ""
            cells += (
                f'<td width="{col_w}" style="width:{col_w}px;height:10px;'
                f'background:{bg};{radius}padding:0;"></td>'
            )
        bar_rows += f"<tr>{cells}</tr>\n"

    label_cells = ""
    for i, day in enumerate(all_days):
        if day == start_d or day.day == 1:
            lbl = day.strftime("%b %-d")
        elif i % 4 == 0:
            lbl = str(day.day)
        else:
            lbl = ""
        label_cells += (
            f'<td width="{col_w}" style="width:{col_w}px;font-family:{FONT};font-size:9px;'
            f'color:{C_SUBTLE};text-align:center;padding-top:3px;white-space:nowrap;'
            f'overflow:hidden;">{lbl}</td>'
        )

    peak_day  = max(daily, key=daily.get)
    peak_val  = daily[peak_day]
    span_days = (end_d - start_d).days + 1

    chart = (
        f'<table cellpadding="0" cellspacing="2" border="0" '
        f'style="border-collapse:separate;border-spacing:2px 0;">'
        f'{bar_rows}<tr>{label_cells}</tr></table>'
        f'<p style="font-family:{FONT};font-size:10px;color:{C_SUBTLE};'
        f'margin:4px 0 0;text-align:right;">'
        f'{start_d.strftime("%b %-d")} – {end_d.strftime("%b %-d")} · '
        f'{span_days} days · peak {peak_day.strftime("%b %-d")} ({peak_val})</p>'
    )
    return chart


# ---------------------------------------------------------------------------
# Market intelligence (§5–§6) — aggregated from runs.jsonl
# ---------------------------------------------------------------------------

_RUNS_PATH = DATA_DIR / "runs.jsonl"

# Canonical profile signal buckets: (label, keyword patterns)
_SIGNAL_BUCKETS: list[tuple[str, list[str]]] = [
    ("AI / LLM Product Strategy",  ["LLM", "RAG", "agentic", "agent", "multi-model", "LangChain", "AutoGen", "AI feature", "AI platform", "generative"]),
    ("0→1 Product Development",    ["0-to-1", "zero-to-one", "earnings call", "shipped", "launched", "full lifecycle", "pioneered"]),
    ("Enterprise Scale & GTM",     ["enterprise", "350+", "paying customer", "Fortune 1000", "large-enterprise", "CLM", "IAM"]),
    ("Eval & Model Operations",    ["eval", "2,000+", "automated eval", "precision", "recall", "benchmark", "telemetry", "reliability metric"]),
    ("Technical Depth",            ["patent", "API", "CLI", "Azure CLI", "CSAT", "developer", "backend", "infrastructure", "technical spec"]),
]

# Human-readable labels for structured extraction fields
_STACK_LABELS: dict[str, str] = {
    "roadmap":         "Product Roadmap & Strategy",
    "technical_specs": "Technical Specifications",
    "user_research":   "User Research & Discovery",
    "data_analysis":   "Data Analysis & Metrics",
    "gtm":             "GTM & Launch",
    "ml_ai":           "ML / AI Systems",
    "platform":        "Platform Architecture",
    "growth":          "Growth & Experimentation",
}
_SURFACE_LABELS: dict[str, str] = {
    "web_app": "Web App Products",
    "api":     "API / Developer Products",
    "platform": "Platform Products",
    "ml_ai":   "ML / AI Products",
    "data":    "Data Products",
    "mobile":  "Mobile Products",
    "internal_tools": "Internal Tooling",
}

# Lomis's profile fit for skill domains (Solid / Moderate / Gap)
_SKILL_FIT: dict[str, str] = {
    "roadmap": "Solid", "technical_specs": "Solid", "ml_ai": "Solid",
    "platform": "Solid", "api": "Solid",
    "user_research": "Moderate", "gtm": "Moderate", "data_analysis": "Moderate",
    "data": "Moderate",
    "growth": "Gap", "mobile": "Gap", "hardware": "Gap", "internal_tools": "Moderate",
}

# Vertical display names keyed by (company_tier, customer_type)
# fit_areas = Lomis's relevant skills for this vertical type
_VERTICAL_META: dict[tuple, tuple[str, str, str]] = {
    # (company_tier, customer_type) -> (display_name, fit_areas, fit)
    ("large_private", "b2b"):      ("Enterprise SaaS (Large Private)", "AI Platform, CLM, IAM",  "Solid"),
    ("large_public",  "b2b"):      ("Enterprise SaaS (Large Public)",  "Enterprise Platform",    "Solid"),
    ("growth",        "b2b"):      ("Growth-Stage B2B",                "0→1 Product",            "Solid"),
    ("early",         "b2b"):      ("Early-Stage B2B",                 "0→1 Scope",              "Moderate"),
    ("large_private", "developer"):("Developer Tools",                  "Azure CLI, API Design",  "Solid"),
    ("large_public",  "b2c"):      ("Consumer (Large Public)",         "Consumer UX",            "Moderate"),
    ("large_private", "b2c"):      ("Consumer (Large Private)",        "Consumer UX",            "Moderate"),
}


def _classify_lp(text: str) -> str | None:
    """Assign a leveragePoint string to a canonical signal bucket via keyword matching."""
    for label, keywords in _SIGNAL_BUCKETS:
        if any(kw.lower() in text.lower() for kw in keywords):
            return label
    return None


def load_market_intel(
    runs_path: Path | None = None,
    lookback_days: int = 30,
    normalize_fn=None,
) -> dict:
    """
    Aggregate market intelligence from runs.jsonl.

    normalize_fn: optional Callable[[list[str]], list[str]] that maps raw
    leveragePoint strings to canonical signal labels. When None, keyword
    bucketing (_classify_lp) is used. This is the extension point for an
    LLM normalization layer.

    Returns:
        {
            "strengths": [{"signal", "evidence", "companies"}, ...],  # top 4
            "skills_a":  [{"name", "freq", "freq_pct", "trend", "fit"}, ...],
            "verticals_b": [{"name", "freq", "freq_pct", "expertise", "fit"}, ...],
        }
    """
    path = runs_path or _RUNS_PATH
    if not path.exists():
        return {"strengths": [], "skills_a": [], "verticals_b": []}

    cutoff = datetime.datetime.now() - datetime.timedelta(days=lookback_days)
    with open(path) as f:
        runs = [json.loads(l) for l in f if l.strip()]

    recent = [
        r for r in runs
        if datetime.datetime.fromisoformat(r["ts"]) >= cutoff
        and r.get("eval", {}).get("verdict") in ("apply", "consider")
    ]
    n = len(recent)
    if not n:
        return {"strengths": [], "skills_a": [], "verticals_b": []}

    # --- §5 Core Strengths --------------------------------------------------
    # signal -> {count, companies_set, example_lp}
    signal_data: dict[str, dict] = {}

    for r in recent:
        lps = r.get("eval", {}).get("leveragePoints") or []
        company = r.get("company", "")

        if normalize_fn is not None:
            labels = normalize_fn(lps)
        else:
            labels = [_classify_lp(lp) for lp in lps]
            labels = [l for l in labels if l]

        for label in set(labels):  # dedupe per run
            if label not in signal_data:
                signal_data[label] = {"count": 0, "companies": set(), "example": ""}
            signal_data[label]["count"] += 1
            if company:
                signal_data[label]["companies"].add(company)

        # Store one example LP per signal (first unseen)
        for lp in lps:
            label = _classify_lp(lp) if normalize_fn is None else None
            if label and label in signal_data and not signal_data[label]["example"]:
                signal_data[label]["example"] = lp

    strengths = []
    for label, sd in sorted(signal_data.items(), key=lambda x: -x[1]["count"]):
        companies_str = ", ".join(sorted(sd["companies"])[:5])
        if len(sd["companies"]) > 5:
            companies_str += f" +{len(sd['companies'])-5} more"
        ex = sd["example"]  # no truncation — full sentence displayed, cell wraps
        strengths.append({
            "signal":    label,
            "evidence":  f"{sd['count']} JDs matched · e.g. {ex}" if ex else f"{sd['count']} JDs matched",
            "companies": companies_str or "—",
        })

    # --- §6A Skills & Domains -----------------------------------------------
    stack_ct: Counter = Counter()
    surface_ct: Counter = Counter()
    for r in recent:
        ext = r.get("extraction") or {}
        for s in (ext.get("primary_execution_stack") or []):
            stack_ct[s] += 1
        for s in (ext.get("product_surface") or []):
            surface_ct[s] += 1

    skills_a = []
    combined = {k: v for k, v in stack_ct.items() if k in _STACK_LABELS}
    combined.update({k: v for k, v in surface_ct.items() if k in _SURFACE_LABELS and k not in combined})
    for key, cnt in sorted(combined.items(), key=lambda x: -x[1]):
        pct = cnt / n * 100
        trend = "Trending ↑" if pct > 50 else ("Established" if pct > 20 else "Niche")
        label = _STACK_LABELS.get(key) or _SURFACE_LABELS.get(key, key)
        fit   = _SKILL_FIT.get(key, "Moderate")
        skills_a.append({
            "name": label, "freq": cnt, "freq_pct": f"{round(pct)}%",
            "trend": trend, "fit": fit,
        })

    # --- §6B Verticals & Markets --------------------------------------------
    vert_ct: Counter = Counter()
    for r in recent:
        ext = r.get("extraction") or {}
        key = (ext.get("company_tier"), ext.get("customer_type"))
        if any(k is not None for k in key):
            vert_ct[key] += 1

    verticals_b = []
    for key, cnt in vert_ct.most_common(6):
        meta = _VERTICAL_META.get(key)
        if meta:
            name, fit_areas, fit = meta
        else:
            parts = [p for p in key if p]
            name = " / ".join(p.replace("_", " ").title() for p in parts) or "Unknown"
            fit_areas = "—"
            fit = "Moderate"
        verticals_b.append({
            "name": name, "freq": cnt,
            "fit_areas": fit_areas, "fit": fit,
        })

    # --- §7 Level Distribution ----------------------------------------------
    _LEVEL_ORDER = ["staff", "senior", "principal", None]
    _LEVEL_LABELS = {"staff": "Staff / Lead (target band)", "senior": "Senior (one below target)",
                     "principal": "Principal (stretch)", None: "Level Not Specified"}
    level_ct: Counter = Counter()
    level_comp: dict[str | None, list[tuple]] = {}
    for r in recent:
        ext = r.get("extraction") or {}
        lvl = ext.get("inferred_structural_level")
        level_ct[lvl] += 1
        if ext.get("salary_disclosed") and ext.get("salary_min") and ext.get("salary_max"):
            level_comp.setdefault(lvl, []).append((ext["salary_min"], ext["salary_max"]))

    levels = []
    for lvl in _LEVEL_ORDER:
        cnt = level_ct.get(lvl, 0)
        if cnt == 0:
            continue
        comp_list = level_comp.get(lvl, [])
        if comp_list:
            mins = [c[0] for c in comp_list]
            maxs = [c[1] for c in comp_list]
            comp_str = f"${min(mins)//1000:,}K–${max(maxs)//1000:,}K base"
        else:
            comp_str = "—"
        levels.append({
            "label": _LEVEL_LABELS.get(lvl, str(lvl)),
            "count": cnt,
            "pct": f"{round(cnt/n*100)}%",
            "comp": comp_str,
        })

    # --- §8 Comp Snapshot ---------------------------------------------------
    comp_disclosed_total = sum(1 for r in recent if (r.get("extraction") or {}).get("salary_disclosed"))
    comp_pct_withheld = round((1 - comp_disclosed_total / n) * 100) if n else 0
    comp_by_level = []
    for lvl in ["staff", "principal", "senior"]:
        data_pts = level_comp.get(lvl, [])
        if data_pts:
            mins = [c[0] for c in data_pts]
            maxs = [c[1] for c in data_pts]
            comp_by_level.append({
                "level": _LEVEL_LABELS.get(lvl, lvl).split(" (")[0],
                "range": f"${min(mins)//1000:,}K–${max(maxs)//1000:,}K",
                "note": f"{len(data_pts)} role{'s' if len(data_pts)>1 else ''} disclosed",
            })
        else:
            comp_by_level.append({
                "level": _LEVEL_LABELS.get(lvl, lvl).split(" (")[0],
                "range": "—",
                "note": "No salary disclosed in batch",
            })

    return {
        "strengths":        strengths[:4],
        "skills_a":         skills_a[:8],
        "verticals_b":      verticals_b[:6],
        "levels":           levels,
        "comp_by_level":    comp_by_level,
        "comp_pct_withheld": comp_pct_withheld,
    }


# ---------------------------------------------------------------------------
# Full HTML assembly
# ---------------------------------------------------------------------------

def _render_strengths(strengths: list) -> str:
    if not strengths:
        return (
            f'<tr><td colspan="3" style="font-family:{FONT};font-size:12px;color:{C_MUTED};'
            f'padding:16px 10px;text-align:center;font-style:italic;">'
            f'No leverage points found in the lookback window.</td></tr>'
        )
    rows = ""
    for s in strengths:
        rows += (
            f"<tr>"
            + _td(f'<strong>{s["signal"]}</strong>', align="left")
            + _td(s["evidence"], align="left", muted=True, size=12)
            + _td(s["companies"], align="left", muted=True, size=12)
            + "</tr>"
        )
    return rows


_COLGROUP_4 = '<colgroup><col style="width:40%"><col style="width:15%"><col style="width:25%"><col style="width:20%"></colgroup>'


def _render_skills_a(skills: list) -> str:
    if not skills:
        return (
            f'<tr><td colspan="4" style="font-family:{FONT};font-size:12px;color:{C_MUTED};'
            f'padding:16px 10px;text-align:center;font-style:italic;">'
            f'No skill data found in the lookback window.</td></tr>'
        )
    rows = ""
    for s in skills:
        rows += (
            f"<tr>"
            + _td(s["name"], align="left")
            + _td(str(s["freq"]), align="center", muted=True)
            + f'<td style="font-family:{FONT};font-size:12px;padding:7px 10px;text-align:center;border-bottom:1px solid {BORDER};">{_trend_pill(s["trend"])}</td>'
            + f'<td style="font-family:{FONT};font-size:12px;padding:7px 10px;text-align:center;border-bottom:1px solid {BORDER};">{_fit_pill(s["fit"])}</td>'
            + "</tr>"
        )
    return rows


def _render_verticals_b(verts: list) -> str:
    if not verts:
        return (
            f'<tr><td colspan="4" style="font-family:{FONT};font-size:12px;color:{C_MUTED};'
            f'padding:16px 10px;text-align:center;font-style:italic;">'
            f'No vertical data found in the lookback window.</td></tr>'
        )
    rows = ""
    for v in verts:
        rows += (
            f"<tr>"
            + _td(v["name"], align="left")
            + _td(str(v["freq"]), align="center", muted=True)
            + _td(v["fit_areas"], align="center", muted=True, size=12)
            + f'<td style="font-family:{FONT};font-size:12px;padding:7px 10px;text-align:center;border-bottom:1px solid {BORDER};">{_fit_pill(v["fit"])}</td>'
            + "</tr>"
        )
    return rows


def _render_levels(levels: list) -> str:
    if not levels:
        return (
            f'<tr><td colspan="4" style="font-family:{FONT};font-size:12px;color:{C_MUTED};'
            f'padding:16px 10px;text-align:center;font-style:italic;">No level data.</td></tr>'
        )
    rows = ""
    for lv in levels:
        rows += (
            f"<tr>"
            + _td(lv["label"], align="left")
            + _td(str(lv["count"]), align="center", muted=True)
            + _td(lv["pct"], align="center", muted=True)
            + _td(lv["comp"], align="center", muted=True, size=12)
            + "</tr>"
        )
    return rows


def _render_comp(comp_by_level: list, pct_withheld: int) -> str:
    rows = ""
    for c in comp_by_level:
        rows += (
            f"<tr>"
            + _td(c["level"], align="left")
            + _td(c["range"], align="center", muted=True)
            + _td(c["note"], align="center", muted=True, size=12)
            + "</tr>"
        )
    note = (
        f'<tr><td colspan="3" style="font-family:{FONT};font-size:11px;color:{C_MUTED};'
        f'padding:8px 10px;font-style:italic;border-top:1px solid {BORDER};">'
        f'{pct_withheld}% of roles withheld compensation. Total comp at AI-native companies '
        f'may run 1.5–2× base when equity is included — a $220K Staff base at a Series B '
        f'can out-earn a $280K base at a large public company.'
        f'</td></tr>'
    )
    return rows + note


def render_html(data: dict, first_name: str = "", lookback_days: int = 30) -> str:
    now_h = datetime.datetime.now().hour
    if now_h < 12:
        greeting = "Good morning"
    elif now_h < 18:
        greeting = "Good afternoon"
    else:
        greeting = "Good evening"

    intel = load_market_intel(lookback_days=lookback_days)

    daily = data["daily"]
    if daily:
        start_d = min(daily.keys())
        end_d   = max(daily.keys())
        date_range = f'{start_d.strftime("%b %-d")} &ndash; {end_d.strftime("%b %-d, %Y")}'
    else:
        date_range = "No data"

    n_scored  = data["roles_scored"]
    n_applied = data["total_applied"]
    alignment = data["alignment"]
    align_pct = f"{alignment:.1%}"
    align_bg  = APPLY_BG  if alignment >= 0.80 else CONSID_BG
    align_fg  = APPLY_NUM if alignment >= 0.80 else CONSID_NUM

    # Alignment banner (shown when < 80%)
    diag_banner = ""
    if alignment < 0.80:
        diag_banner = (
            f'{_spacer(8)}'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0">'
            f'<tr><td style="border-left:3px solid {CONSID_NUM};background:{CONSID_BG};'
            f'padding:8px 12px;border-radius:0 4px 4px 0;">'
            f'<span style="font-family:{FONT};font-size:12px;color:{CONSID_NUM};">'
            f'Recommendation alignment is {align_pct} — below the 80% threshold. '
            f'Your application choices diverged from scorer suggestions more than expected. '
            f'Consider reviewing your profile weights.</span>'
            f'</td></tr></table>'
        )

    solid    = data["solid"]
    moderate = data["moderate"]
    partial  = data["partial"]
    pending  = data["pending"]
    screens  = data["screens"]
    rejects  = data["rejections"]

    html = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:{BODY_BG};font-family:{FONT};">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:{BODY_BG};">
<tr><td align="center" style="padding:24px 16px;">
<table width="600" cellpadding="0" cellspacing="0" border="0" style="background:{CARD_BG};border-radius:10px;overflow:hidden;">

<!-- HEADER -->
<tr><td style="padding:28px 28px 20px;">
  <p style="font-family:{FONT};font-size:18px;font-weight:600;color:{C_HEADING};margin:0 0 4px;line-height:1.3;">{greeting}{', ' + first_name if first_name else ''}</p>
  <p style="font-family:{FONT};font-size:13px;color:{C_MUTED};margin:0;line-height:1.5;">Metis report &middot; {date_range} &middot; {n_scored} roles evaluated</p>
</td></tr>

<!-- §2 EXECUTIVE ROI -->
<tr><td style="padding:0 28px 0;">
{_section_h("Executive ROI")}
{_tiles([
    ("Roles Scored",  str(n_scored),                  None, TOTAL_BG,  TOTAL_NUM),
    ("Total Applied", str(n_applied),                 None, TOTAL_BG,  TOTAL_NUM),
    ("Time Saved",    f"{data['time_saved_h']}h",     None, APPLY_BG,  APPLY_NUM),
    ("Alignment",     align_pct,                      None, align_bg,  align_fg),
])}
{diag_banner}
{_spacer(6)}
<p style="font-family:{FONT};font-size:10px;color:{C_SUBTLE};margin:0;">
  Time saved estimated at 1 min/role vs. manual JD review &middot;
  Alignment = (recommended &amp; applied + not recommended &amp; not applied) / total scored &middot; base = {data['align_base']} roles
</p>
</td></tr>

{_spacer(20)}

<!-- §3 SCORE EVAL -->
<tr><td style="padding:0 28px 0;">
{_section_h("Score Eval Distribution")}
{_tiles([
    ("Total Scored",   str(n_scored), "100%",              TOTAL_BG,  TOTAL_NUM),
    ("Solid Match",    str(solid),    _pct(solid,    n_scored), APPLY_BG,  APPLY_NUM),
    ("Moderate Match", str(moderate), _pct(moderate, n_scored), CONSID_BG, CONSID_NUM),
    ("Partial Match",  str(partial),  _pct(partial,  n_scored), TOTAL_BG,  TOTAL_NUM),
])}
</td></tr>

{_spacer(20)}

<!-- §4 APPLICATION PIPELINE -->
<tr><td style="padding:0 28px 0;">
{_section_h("Application Pipeline")}
{_tiles([
    ("Total Applied", str(n_applied), "100%",                TOTAL_BG,  TOTAL_NUM),
    ("Pending",       str(pending),   _pct(pending,  n_applied), CONSID_BG, CONSID_NUM),
    ("Screens",       str(screens),   _pct(screens,  n_applied), APPLY_BG,  APPLY_NUM),
    ("Rejections",    str(rejects),   _pct(rejects,  n_applied), RED_BG,    RED_NUM),
])}
{_spacer(16)}
{_sub_h("Daily Application Volume")}
{_bar_chart(daily)}
</td></tr>

{_spacer(20)}

<!-- §5 CORE STRENGTHS -->
<tr><td style="padding:0 28px 0;">
{_section_h("Core Strengths Alignment")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};table-layout:fixed;">
<colgroup><col style="width:28%"><col style="width:44%"><col style="width:28%"></colgroup>
{_green_th(["Profile Signal", "Evidence Found in JDs", "Companies in Batch Hiring This"])}
{_render_strengths(intel["strengths"])}
</table>
</td></tr>

{_spacer(20)}

<!-- §6 MARKET LANDSCAPE -->
<tr><td style="padding:0 28px 0;">
{_section_h("Market Landscape")}
{_sub_h("Table A — Skills & Domains")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};table-layout:fixed;">
{_COLGROUP_4}
{_green_th(["Skill / Domain", "Role Count", "Trend", "Profile Fit"], ["left","center","center","center"])}
{_render_skills_a(intel["skills_a"])}
</table>
{_spacer(14)}
{_sub_h("Table B — Verticals & Markets")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};table-layout:fixed;">
{_COLGROUP_4}
{_green_th(["Company Vertical", "Role Count", "Your Fit Areas", "Profile Fit"], ["left","center","center","center"])}
{_render_verticals_b(intel["verticals_b"])}
</table>
</td></tr>

{_spacer(20)}

<!-- §7 ROLE LEVEL DISTRIBUTION -->
<tr><td style="padding:0 28px 0;">
{_section_h("Role Level Distribution")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};table-layout:fixed;">
{_COLGROUP_4}
{_green_th(["Level", "Role Count", "% of Batch", "Avg Comp Range"], ["left","center","center","center"])}
{_render_levels(intel["levels"])}
</table>
</td></tr>

{_spacer(20)}

<!-- §8 COMP SNAPSHOT -->
<tr><td style="padding:0 28px 0;">
{_section_h("Comp Snapshot")}
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:1px solid {BORDER};table-layout:fixed;">
<colgroup><col style="width:30%"><col style="width:25%"><col style="width:45%"></colgroup>
{_green_th(["Level", "Base Range", "Notes"], ["left","center","center"])}
{_render_comp(intel["comp_by_level"], intel["comp_pct_withheld"])}
</table>
</td></tr>

{_spacer(28)}

<!-- FOOTER -->
<tr><td style="padding:0 28px 28px;">
  <p style="font-family:{FONT};font-size:10px;color:{C_SUBTLE};margin:0;line-height:1.6;border-top:1px solid {BORDER};padding-top:14px;">
    Generated by metis &middot; {datetime.date.today().strftime("%b %-d, %Y")} &middot;
    {n_scored} roles evaluated
  </p>
</td></tr>

</table></td></tr></table>
</body></html>"""

    return html


# ---------------------------------------------------------------------------
# Output — send email or save file
# ---------------------------------------------------------------------------

def _send_email(html: str, gmail_address: str, app_password: str, preview: bool) -> None:
    prefix = "[DRAFT PREVIEW] " if preview else ""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"{prefix}Metis Progress Report — Insights from your recent searches"
    msg["From"]    = gmail_address
    msg["To"]      = gmail_address
    msg.attach(MIMEText(html, "html"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(gmail_address, app_password)
        smtp.sendmail(gmail_address, gmail_address, msg.as_string())


def _save_file(html: str, output_path: Path) -> None:
    suffix = output_path.suffix.lower()
    if suffix == ".pdf":
        try:
            from weasyprint import HTML as WP_HTML
        except ImportError:
            raise SystemExit(
                "weasyprint is required for PDF output: pip install weasyprint\n"
                "On macOS also run: brew install pango"
            )
        WP_HTML(string=html).write_pdf(str(output_path))
    else:
        output_path.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run_report(
    tracker_path: Path,
    gmail_address: str,
    app_password: str,
    output: str | None = None,
    preview: bool = True,
    lookback_days: int = 30,
) -> None:
    """Generate the Metis report and send or save it."""
    from .profile import load_profile_yaml
    profile = load_profile_yaml() or {}
    candidate_name = profile.get("candidate", {}).get("name", "")
    first_name = candidate_name.split()[0] if candidate_name else ""
    data = load_report_data(tracker_path)
    html = render_html(data, first_name=first_name, lookback_days=lookback_days)

    if output:
        out_path = Path(output)
        _save_file(html, out_path)
        print(f"Report saved to {out_path.resolve()}")
    else:
        _send_email(html, gmail_address, app_password, preview=preview)
        label = "draft preview" if preview else "report"
        print(f"Metis {label} sent to {gmail_address}")
