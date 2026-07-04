from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any


_VALID_VERDICTS = {"apply", "consider", "skipped", "filtered", "prescreened"}
_RECOMMENDED_VERDICTS = {"apply", "consider"}
_VERDICT_RANK = {"apply": 0, "consider": 1}


def _default_data_dir(env: dict[str, str] | None = None) -> Path:
    env_map = env if env is not None else os.environ
    return Path(env_map.get("METIS_DATA_DIR", str(Path.home() / ".job_pipeline"))).expanduser()


def _default_profile_path(
    *,
    data_dir: Path | None = None,
    env: dict[str, str] | None = None,
) -> Path:
    env_map = env if env is not None else os.environ
    if env_map.get("METIS_PROFILE"):
        return Path(env_map["METIS_PROFILE"]).expanduser()
    return (data_dir or _default_data_dir(env_map)) / "profile.yaml"


def _default_tracker_path(
    *,
    data_dir: Path | None = None,
    env: dict[str, str] | None = None,
) -> Path:
    env_map = env if env is not None else os.environ
    if env_map.get("TRACKER_PATH"):
        return Path(env_map["TRACKER_PATH"]).expanduser()
    return (data_dir or _default_data_dir(env_map)) / "applications.xlsx"


def _is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.expanduser().resolve(strict=False).relative_to(parent.expanduser().resolve(strict=False))
        return True
    except ValueError:
        return False


def _safe_json(path: Path, fallback: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return fallback


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return []
    for line in lines:
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            rows.append(value)
    return rows


def _role_id(row: dict[str, Any]) -> str:
    role_hash = str(row.get("role_hash") or "").strip()
    if role_hash:
        return role_hash
    key = f"{row.get('title', '')}{row.get('company', '')}".lower()
    return re.sub(r"[^a-z0-9]", "", key)[:24]


def _role_key(title: Any, company: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", f"{title or ''}{company or ''}".lower())


def _role_summary(row: dict[str, Any]) -> dict[str, Any]:
    eval_data = row.get("eval") if isinstance(row.get("eval"), dict) else {}
    return {
        "role_id": _role_id(row),
        "title": row.get("title", ""),
        "company": row.get("company", ""),
        "location": row.get("location", ""),
        "url": row.get("url", ""),
        "source": row.get("source", ""),
        "score": eval_data.get("score"),
        "verdict": eval_data.get("verdict"),
        "friction_points": eval_data.get("frictionPoints") or [],
        "leverage_points": eval_data.get("leveragePoints") or [],
        "tags": eval_data.get("tags") or [],
        "tailoring_available": bool(row.get("url") or row.get("jd")),
        "timestamp": row.get("ts", ""),
    }


def get_metis_status(
    *,
    data_dir: str | Path | None = None,
    profile_path: str | Path | None = None,
    tracker_path: str | Path | None = None,
    env: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Return setup and state status without mutating local files."""
    env_map = env if env is not None else os.environ
    resolved_data_dir = Path(data_dir).expanduser() if data_dir else _default_data_dir(env_map)
    resolved_profile = (
        Path(profile_path).expanduser()
        if profile_path
        else _default_profile_path(data_dir=resolved_data_dir, env=env_map)
    )
    resolved_tracker = (
        Path(tracker_path).expanduser()
        if tracker_path
        else _default_tracker_path(data_dir=resolved_data_dir, env=env_map)
    )
    last_run_path = resolved_data_dir / "last_run.json"
    runs_path = resolved_data_dir / "runs.jsonl"
    schedule_path = resolved_data_dir / "schedule.json"
    feedback_path = resolved_data_dir / "feedback.md"

    llm_provider = env_map.get("METIS_LLM_PROVIDER", env_map.get("LLM_PROVIDER", "anthropic"))
    provider_key = "OPENAI_API_KEY" if llm_provider.lower().replace("-", "_") in {"openai", "open_ai"} else "ANTHROPIC_API_KEY"
    email_oauth_provider = _safe_json(resolved_data_dir / "email_provider.json", {}).get("provider")
    email_has_oauth = bool(
        email_oauth_provider
        and (resolved_data_dir / f"{email_oauth_provider.split('_')[0]}_token.json").exists()
    )
    email_has_legacy = bool(env_map.get("GMAIL_ADDRESS") and env_map.get("GMAIL_APP_PASSWORD"))

    checks = {
        "data_dir_exists": resolved_data_dir.exists(),
        "profile_exists": resolved_profile.exists(),
        "llm_api_key_configured": bool(env_map.get(provider_key)),
        "email_access_configured": email_has_oauth or email_has_legacy,
        "tracker_exists": resolved_tracker.exists(),
        "runs_log_exists": runs_path.exists(),
        "last_run_exists": last_run_path.exists(),
        "feedback_exists": feedback_path.exists(),
        "schedule_exists": schedule_path.exists(),
    }
    missing = []
    if not checks["profile_exists"]:
        missing.append("profile")
    if not checks["llm_api_key_configured"]:
        missing.append("llm_api_key")
    if not checks["email_access_configured"]:
        missing.append("email_access")
    warnings = []
    if resolved_profile.exists() and not _is_relative_to(resolved_profile, resolved_data_dir):
        warnings.append(
            "profile_path is outside data_dir; confirm this is an intentional persona/config split."
        )
    if resolved_tracker.exists() and not _is_relative_to(resolved_tracker, resolved_data_dir):
        warnings.append(
            "tracker_path is outside data_dir; tracker writes may affect a different persona/state directory."
        )

    return {
        "configured": not missing,
        "missing": missing,
        "warnings": warnings,
        "checks": checks,
        "paths": {
            "data_dir": str(resolved_data_dir),
            "profile": str(resolved_profile),
            "tracker": str(resolved_tracker),
            "runs": str(runs_path),
            "last_run": str(last_run_path),
            "feedback": str(feedback_path),
            "schedule": str(schedule_path),
        },
        "llm": {
            "provider": llm_provider,
            "api_key_env": provider_key,
            "api_key_configured": checks["llm_api_key_configured"],
        },
        "email": {
            "oauth_provider": email_oauth_provider,
            "oauth_configured": email_has_oauth,
            "legacy_gmail_configured": email_has_legacy,
        },
        "last_run": _safe_json(last_run_path, None) if last_run_path.exists() else None,
        "next_steps": _setup_next_steps(missing),
    }


def _setup_next_steps(missing: list[str]) -> list[str]:
    steps = []
    if "profile" in missing:
        steps.append("Run `metis init` to create a scoring profile.")
    if "email_access" in missing:
        steps.append("Run `metis config access` to connect Gmail or Outlook, or configure legacy Gmail credentials.")
    if "llm_api_key" in missing:
        steps.append("Set the API key for the configured LLM provider.")
    return steps


def list_recommended_roles(
    *,
    data_dir: str | Path | None = None,
    limit: int = 20,
    verdicts: set[str] | None = None,
    latest_run_only: bool = True,
) -> dict[str, Any]:
    """Return recent Solid/Moderate roles from runs.jsonl, newest first."""
    resolved_data_dir = Path(data_dir).expanduser() if data_dir else _default_data_dir()
    allowed = verdicts or _RECOMMENDED_VERDICTS
    invalid = allowed - _VALID_VERDICTS
    if invalid:
        raise ValueError(f"Unknown verdict(s): {', '.join(sorted(invalid))}")

    rows = _read_jsonl(resolved_data_dir / "runs.jsonl")
    allowed_keys: set[str] | None = None
    if latest_run_only:
        last_run = _safe_json(resolved_data_dir / "last_run.json", None)
        if isinstance(last_run, dict) and isinstance(last_run.get("roles"), list):
            allowed_keys = {
                _role_key(role.get("title"), role.get("company"))
                for role in last_run["roles"]
                if isinstance(role, dict)
            }

    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        eval_data = row.get("eval") if isinstance(row.get("eval"), dict) else {}
        if eval_data.get("verdict") not in allowed:
            continue
        if allowed_keys is not None and _role_key(row.get("title"), row.get("company")) not in allowed_keys:
            continue
        deduped[_role_id(row)] = row

    roles = [_role_summary(row) for row in deduped.values()]
    roles.sort(
        key=lambda role: (
            _VERDICT_RANK.get(str(role.get("verdict")), 99),
            -(role.get("score") or 0),
            str(role.get("company") or ""),
            str(role.get("title") or ""),
        )
    )
    if limit >= 0:
        roles = roles[:limit]
    return {
        "count": len(roles),
        "roles": roles,
        "data_dir": str(resolved_data_dir),
    }


def get_role_details(
    role_id: str,
    *,
    data_dir: str | Path | None = None,
) -> dict[str, Any] | None:
    """Return the latest trace record for a role id/hash."""
    resolved_data_dir = Path(data_dir).expanduser() if data_dir else _default_data_dir()
    rows = _read_jsonl(resolved_data_dir / "runs.jsonl")
    for row in reversed(rows):
        if _role_id(row) == role_id:
            detail = dict(row)
            detail["role_id"] = _role_id(row)
            return detail
    return None


_COMMENT_RE = re.compile(r"<!--\s*id:(?P<id>\S+?)\s*\|(?P<meta>.*?)-->")
_HEADER_RE = re.compile(r"^## (?P<header>.+?)$", re.MULTILINE)


def list_scoring_feedback(
    *,
    data_dir: str | Path | None = None,
    limit: int = 5,
) -> dict[str, Any]:
    """Return recent feedback.md entries without importing feedback.py."""
    resolved_data_dir = Path(data_dir).expanduser() if data_dir else _default_data_dir()
    feedback_path = resolved_data_dir / "feedback.md"
    if not feedback_path.exists():
        return {"count": 0, "entries": [], "path": str(feedback_path)}

    try:
        content = feedback_path.read_text(encoding="utf-8")
    except OSError:
        return {"count": 0, "entries": [], "path": str(feedback_path)}

    comment_positions = [
        (m.start(), m.end(), m.group("id"), m.group("meta"))
        for m in _COMMENT_RE.finditer(content)
    ]
    headers = [(m.start(), m.group("header").strip()) for m in _HEADER_RE.finditer(content)]
    entries: list[dict[str, Any]] = []
    for idx, (pos, header) in enumerate(headers):
        end = headers[idx + 1][0] if idx + 1 < len(headers) else len(content)
        prev = headers[idx - 1][0] if idx else 0
        feedback_id = None
        meta = ""
        for c_start, c_end, c_id, c_meta in comment_positions:
            if c_start >= prev and c_end <= pos:
                feedback_id = c_id
                meta = c_meta

        body_lines = [
            line.strip()
            for line in content[pos:end].splitlines()[1:]
            if line.strip() and not line.strip().startswith("<!--")
        ]
        entries.append({
            "feedback_id": feedback_id,
            "header": header,
            "date": re.sub(r"\[(?:user|auto)\]\s*", "", header).strip(),
            "source": "auto" if "[auto]" in header else "user",
            "summary": body_lines[0] if body_lines else "",
            "metadata": _parse_feedback_meta(meta),
        })

    recent = list(reversed(entries))
    if limit >= 0:
        recent = recent[:limit]
    return {"count": len(recent), "entries": recent, "path": str(feedback_path)}


def _parse_feedback_meta(meta: str) -> dict[str, Any]:
    parsed: dict[str, Any] = {}
    for part in meta.split("|"):
        if ":" not in part:
            continue
        key, value = part.split(":", 1)
        key = key.strip()
        value = value.strip()
        if key in {"roles", "dims"}:
            parsed[key] = [item for item in value.split(",") if item]
        else:
            parsed[key] = value
    return parsed


def list_application_activity(
    *,
    tracker_path: str | Path | None = None,
    data_dir: str | Path | None = None,
    limit: int = 20,
) -> dict[str, Any]:
    """Return recent rows from applications.xlsx in tracker-column terms."""
    resolved_data_dir = Path(data_dir).expanduser() if data_dir else _default_data_dir()
    resolved_tracker = Path(tracker_path).expanduser() if tracker_path else _default_tracker_path(data_dir=resolved_data_dir)
    if not resolved_tracker.exists():
        return {"count": 0, "applications": [], "path": str(resolved_tracker)}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(resolved_tracker, read_only=True, data_only=True)
    except Exception:
        return {"count": 0, "applications": [], "path": str(resolved_tracker), "error": "tracker_unreadable"}

    ws = wb.active
    headers = [
        str(cell.value or "").strip()
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    applications: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row):
            continue
        item = {
            headers[idx] if idx < len(headers) and headers[idx] else f"column_{idx + 1}": value
            for idx, value in enumerate(row)
        }
        applications.append(item)

    if limit >= 0:
        applications = applications[:limit]
    return {
        "count": len(applications),
        "applications": applications,
        "path": str(resolved_tracker),
    }
