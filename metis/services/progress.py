from __future__ import annotations

import datetime
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from .read import _default_data_dir, _default_tracker_path


_SOLID = {"Apply", "Solid", "Solid Match"}
_MODERATE = {"Consider", "Moderate", "Moderate Match"}
_PARTIAL = {"Skipped", "Partial", "Partial Match"}
_SCORED = _SOLID | _MODERATE | _PARTIAL


def _coerce_date(value: Any) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        try:
            return datetime.date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


def _load_tracker_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    ws = wb.active
    try:
        headers = [
            str(cell.value or "").strip()
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
    except StopIteration:
        wb.close()
        return []

    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in raw):
            continue
        rows.append({
            headers[idx] if idx < len(headers) and headers[idx] else f"column_{idx + 1}": value
            for idx, value in enumerate(raw)
        })
    wb.close()
    return rows


def _load_skipped_count(data_dir: Path) -> int:
    path = data_dir / "skipped_roles.json"
    if not path.exists():
        return 0
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return 0
    return len(data) if isinstance(data, dict) else 0


def generate_progress_summary(
    *,
    data_dir: str | Path | None = None,
    tracker_path: str | Path | None = None,
    lookback_days: int = 30,
) -> dict[str, Any]:
    """Return structured progress metrics from tracker and skipped-role state."""
    resolved_data_dir = Path(data_dir).expanduser() if data_dir else _default_data_dir()
    resolved_tracker = (
        Path(tracker_path).expanduser()
        if tracker_path
        else _default_tracker_path(data_dir=resolved_data_dir)
    )
    rows = _load_tracker_rows(resolved_tracker)
    skipped_count = _load_skipped_count(resolved_data_dir)
    market_intel = _load_market_intel(resolved_data_dir / "runs.jsonl", lookback_days)

    cutoff = datetime.date.today() - datetime.timedelta(days=lookback_days)
    recent_rows = [
        row for row in rows
        if (date_value := _coerce_date(row.get("date_suggested"))) is None or date_value >= cutoff
    ]

    scored = [row for row in recent_rows if row.get("suggestion_status") in _SCORED]
    solid_rows = [row for row in scored if row.get("suggestion_status") in _SOLID]
    moderate_rows = [row for row in scored if row.get("suggestion_status") in _MODERATE]
    partial_rows = [row for row in scored if row.get("suggestion_status") in _PARTIAL]
    applied = [row for row in recent_rows if row.get("action_taken") == "Applied"]
    pending = [row for row in applied if row.get("application_status") == "Pending"]
    screens = [row for row in applied if row.get("application_status") == "Recruiter Screen"]
    rejections = [row for row in applied if row.get("application_status") == "Rejected"]

    recommended = solid_rows + moderate_rows
    true_positive = sum(1 for row in recommended if row.get("action_taken") == "Applied")
    true_negative = skipped_count + sum(1 for row in partial_rows if row.get("action_taken") != "Applied")
    alignment_base = len(recommended) + len(partial_rows) + skipped_count
    alignment = (true_positive + true_negative) / alignment_base if alignment_base else 0.0
    roles_scored = len(solid_rows) + len(moderate_rows) + skipped_count

    daily: dict[str, int] = defaultdict(int)
    for row in applied:
        applied_date = _coerce_date(row.get("date_applied"))
        if applied_date:
            daily[applied_date.isoformat()] += 1

    not_applied_recommendations = [
        {
            "date_suggested": row.get("date_suggested"),
            "role_title": row.get("role_title"),
            "company": row.get("company"),
            "match_score": row.get("match_score"),
            "suggestion_status": row.get("suggestion_status"),
        }
        for row in recommended
        if row.get("action_taken") != "Applied"
    ]
    not_applied_recommendations.sort(
        key=lambda row: (-(row.get("match_score") or 0), str(row.get("company") or ""))
    )

    return {
        "lookback_days": lookback_days,
        "paths": {
            "data_dir": str(resolved_data_dir),
            "tracker": str(resolved_tracker),
            "skipped_roles": str(resolved_data_dir / "skipped_roles.json"),
        },
        "metrics": {
            "roles_scored": roles_scored,
            "total_applied": len(applied),
            "time_saved_h": round(roles_scored / 60, 1),
            "alignment": alignment,
            "true_positive": true_positive,
            "true_negative": true_negative,
            "alignment_base": alignment_base,
        },
        "score_distribution": {
            "solid": len(solid_rows),
            "moderate": len(moderate_rows),
            "partial": skipped_count,
        },
        "application_pipeline": {
            "pending": len(pending),
            "screens": len(screens),
            "rejections": len(rejections),
            "daily": dict(sorted(daily.items())),
        },
        "market_intel": market_intel,
        "not_applied_recommendations": not_applied_recommendations[:20],
    }


def _load_market_intel(runs_path: Path, lookback_days: int) -> dict[str, Any]:
    """Load report-style market intelligence with explicit runs path."""
    try:
        from metis.report_cmd import load_market_intel
    except Exception:
        return _empty_market_intel()

    intel = load_market_intel(runs_path=runs_path, lookback_days=lookback_days)
    return {
        "profile_signals": intel.get("strengths", []),
        "skills_in_demand": intel.get("skills_a", []),
        "verticals_and_markets": intel.get("verticals_b", []),
        "role_level_distribution": intel.get("levels", []),
        "comp_snapshot": {
            "by_level": intel.get("comp_by_level", []),
            "pct_withheld": intel.get("comp_pct_withheld", 0),
        },
    }


def _empty_market_intel() -> dict[str, Any]:
    return {
        "profile_signals": [],
        "skills_in_demand": [],
        "verticals_and_markets": [],
        "role_level_distribution": [],
        "comp_snapshot": {"by_level": [], "pct_withheld": 0},
    }
