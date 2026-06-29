#!/usr/bin/env python3
"""One-off backfill: patch Pre-tracker rows to External + fill blank role titles.

Fetches one batch of emails (last 30 days), then matches against tracker rows —
no per-company IMAP calls. Tests extraction on up to --sample N rows (default 30).

Usage:
    python backfill_external_roles.py              # run on up to 30 Pre-tracker rows
    python backfill_external_roles.py --sample 10  # smaller sample
    python backfill_external_roles.py --all        # all Pre-tracker rows
    python backfill_external_roles.py --dry-run    # preview only, no writes
"""
import sys, os, datetime, logging
from pathlib import Path

import importlib.util
if importlib.util.find_spec("metis") is None:
    sys.path.insert(0, str(Path(__file__).parent))

import anthropic
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

DRY_RUN  = "--dry-run" in sys.argv
RUN_ALL  = "--all" in sys.argv
SAMPLE_N = 30
for i, arg in enumerate(sys.argv):
    if arg == "--sample" and i + 1 < len(sys.argv):
        try: SAMPLE_N = int(sys.argv[i + 1])
        except ValueError: pass

_GARBAGE = ["unique skills deemed necessary", "we received your application",
            "thank you", "application", "job id", "req"]

def _is_bad_title(t: str) -> bool:
    if not t or len(t.strip()) < 5: return True
    return any(p in t.lower() for p in _GARBAGE)

def _match_email(emails: list, company: str, date_applied) -> dict | None:
    """Find best email match for a company from a pre-fetched list."""
    co = company.lower().replace(" inc.", "").replace(" llc", "").strip()
    if isinstance(date_applied, datetime.datetime): target = date_applied.date()
    elif isinstance(date_applied, datetime.date):   target = date_applied
    else: target = None

    candidates = []
    for e in emails:
        text = " ".join([e.get("sender",""), e.get("subject",""), e.get("body","")[:400]]).lower()
        if co in text:
            d = e.get("date")
            delta = abs((d.date() - target).days) if target and d and hasattr(d, "date") else 999
            candidates.append((delta, e))
    return min(candidates, key=lambda x: x[0])[1] if candidates else None


def main():
    load_dotenv(override=True)
    from metis.track import fetch_candidate_emails, extract_role, _extract_role_llm
    from metis.xlsx import TRACKER_PATH

    api_key      = os.environ.get("ANTHROPIC_API_KEY", "")
    gmail_addr   = os.environ.get("GMAIL_ADDRESS", "")
    app_password = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not api_key or not gmail_addr or not app_password:
        log.error("Missing env vars — ANTHROPIC_API_KEY, GMAIL_ADDRESS, GMAIL_APP_PASSWORD")
        return

    client = anthropic.Anthropic(api_key=api_key)

    if not TRACKER_PATH.exists():
        log.error("Tracker not found: %s", TRACKER_PATH)
        return

    # Single IMAP fetch — one round trip for all rows
    since_dt = datetime.datetime.now() - datetime.timedelta(days=30)
    log.info("Fetching emails since %s (one batch)…", since_dt.strftime("%Y-%m-%d"))
    emails = fetch_candidate_emails(gmail_addr, app_password, since_dt)
    log.info("Fetched %d candidate emails.", len(emails))

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active

    COL_ROLE_TITLE        = 2
    COL_COMPANY           = 3
    COL_SUGGESTION_STATUS = 5
    COL_DATE_APPLIED      = 7

    patched = title_filled = processed = 0

    for row_idx in range(2, ws.max_row + 1):
        if not RUN_ALL and processed >= SAMPLE_N:
            break
        if ws.cell(row_idx, COL_SUGGESTION_STATUS).value != "Pre-tracker":
            continue

        company      = ws.cell(row_idx, COL_COMPANY).value or ""
        role_title   = ws.cell(row_idx, COL_ROLE_TITLE).value or ""
        date_applied = ws.cell(row_idx, COL_DATE_APPLIED).value
        processed += 1

        log.info("Row %d — %s / %s", row_idx, company or "?", role_title or "(blank)")

        if not DRY_RUN:
            ws.cell(row_idx, COL_SUGGESTION_STATUS).value = "External"
            ws.cell(row_idx, COL_SUGGESTION_STATUS).fill = PatternFill(fill_type=None)
        patched += 1

        if _is_bad_title(role_title) and company:
            email_dict = _match_email(emails, company, date_applied)
            if email_dict:
                subject   = email_dict.get("subject", "")
                body      = email_dict.get("body", "")
                new_title = extract_role(subject, body) or _extract_role_llm(subject, body, client)
                if new_title and not _is_bad_title(new_title):
                    log.info("  → extracted: %r", new_title)
                    if not DRY_RUN:
                        ws.cell(row_idx, COL_ROLE_TITLE).value = new_title
                    title_filled += 1
                else:
                    log.info("  → no title signal in matched email")
            else:
                log.info("  → no email match in 30-day batch")

    if DRY_RUN:
        log.info("[DRY RUN] Would patch %d rows -> External, fill %d titles. No writes.", patched, title_filled)
        return

    wb.save(TRACKER_PATH)
    TRACKER_PATH.chmod(0o600)
    log.info("Done. Patched %d rows -> External, filled %d titles.", patched, title_filled)
    print(f"  Tracker -> {TRACKER_PATH}")


if __name__ == "__main__":
    main()
